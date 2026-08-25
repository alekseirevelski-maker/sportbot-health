#!/usr/bin/env python3
"""Telegram бот ЧБК v3.0 — Мониторинг здоровья спортсменов."""

import os, json, csv, io, re, logging, time, asyncio
from datetime import datetime, date, timedelta
from typing import Dict, Any, Optional, List, Tuple
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes
)

from config import (
    BOT_TOKEN, AGE_GROUPS, SIMPLE_PROTOCOLS,
    HRV_NORMS, HR_NORMS, EMOJIS,
    MOTIVATIONAL_MESSAGES, PROXY_URL,
    RATE_LIMIT_MESSAGES, RATE_LIMIT_WINDOW,
    MENSTRUAL_ENABLED_GROUPS,
    ADMIN_IDS, REMINDER_HOUR_DEFAULT, REMINDER_MINUTE_DEFAULT,
    BMI_PERCENTILES, AGE_GROUP_YEARS,
    BC_NORMS, BC_RED_FLAGS,
)
from cycle_medicine import (
    get_cycle_phase_info, get_cycle_phase, PHASES,
    AGE_SPECIFIC, CYCLE_LENGTH_MEDIAN,
)
from watch_parser import parse_watch
from scale_csv_parser import parse_scale_csv_dated

from validator import sanitize_text, validate_heart_rate, validate_age_ge14

if PROXY_URL:
    os.environ["http_proxy"] = PROXY_URL
    os.environ["https_proxy"] = PROXY_URL

from database import Database, TEAMS, ACTIVE_TEAMS

# TESTING flag — disables all outgoing notifications to athletes
TESTING = os.environ.get("TESTING", "").lower() in ("true", "1", "yes")

class _TokenFilter(logging.Filter):
    def filter(self, record):
        if BOT_TOKEN:
            # Маскируем токен в самом сообщении и в его аргументах
            for attr in ("msg", "message"):
                val = getattr(record, attr, None)
                if isinstance(val, str) and BOT_TOKEN in val:
                    setattr(record, attr, val.replace(BOT_TOKEN, "***TOKEN***"))
        return True

from logging.handlers import RotatingFileHandler
_log_handler = RotatingFileHandler("bot.log", maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
_log_handler.addFilter(_TokenFilter())
_console = logging.StreamHandler()
_console.addFilter(_TokenFilter())
logging.basicConfig(format="%(asctime)s [%(levelname)s] %(message)s", level=logging.INFO, handlers=[_log_handler, _console])
logger = logging.getLogger(__name__)

# Подавляем шумный HTTP-лог (python-telegram-bot/httpx печатают URL с токеном на INFO)
for _noisy in ("httpx", "httpcore", "urllib3", "apscheduler", "telegram._httpx"):
    logging.getLogger(_noisy).setLevel(logging.WARNING)

ADMIN_TELEGRAM_IDS = ADMIN_IDS

# Настройки напоминаний (по умолчанию из .env; переопределяются из БД при старте)
REMINDER_HOUR = REMINDER_HOUR_DEFAULT
REMINDER_MINUTE = REMINDER_MINUTE_DEFAULT
REMINDER_TZ = "Asia/Yekaterinburg"

# Сколько профилей весов поддерживаем (весы GARLYN Bodyscan Master / MovingLife).
# Физически прибор может держать меньше — тогда взвешивают подгруппами, перепривязывая.
SCALE_PROFILE_COUNT = 50
# Кнопок профилей на одном экране (Telegram не любит >~100), пагинация
BC_PROFILE_PAGE = 25


# ==================== УТИЛИТЫ ====================

def sparkline(values, width=7):
    clean = [v for v in values if v is not None]
    if not clean:
        return "▁" * width
    mn, mx = min(clean), max(clean)
    rng = mx - mn if mx != mn else 1
    chars = ["▁", "▂", "▃", "▄", "▅", "▆", "▇", "█"]
    return "".join(chars[max(0, min(7, int(round((v - mn) / rng * 7)))) if v is not None else 0] for v in values[-width:])


def sparkline_colored(values, width=7, invert=False):
    """Цветной спарклайн: 🟢🟡🔴 вместо монохромных символов.
    Если invert=True — 1=хорошо, 7=плохо."""
    clean = [v for v in values if v is not None]
    if not clean:
        return "—" * width
    result = ""
    for v in values[-width:]:
        if v is None:
            result += "—"
            continue
        if invert:
            if v >= 6: result += "🔴"
            elif v >= 4: result += "🟡"
            else: result += "🟢"
        else:
            if v <= 2: result += "🔴"
            elif v <= 4: result += "🟡"
            else: result += "🟢"
    return result




def score_bar(score, maximum=7):
    if score is None:
        return "—"
    filled = max(0, min(maximum, int(score)))
    return "▓" * filled + "░" * (maximum - filled) + f" {score}/{maximum}"


def trend_arrow(current, previous, invert=False):
    if current is None or previous is None:
        return "—"
    diff = current - previous
    if abs(diff) < 0.3:
        return "→ стабильно"
    if invert:
        diff = -diff
    return f"↑ +{diff:.1f}" if diff > 0 else f"↓ {diff:.1f}"


def get_rank(streak):
    if streak >= 30:
        return "🥇 Золото"
    elif streak >= 14:
        return "🥈 Серебро"
    elif streak >= 7:
        return "🥉 Бронза"
    elif streak >= 3:
        return "⭐ Старт"
    return "🌱 Новичок"


def get_score_emoji(score):
    """Эмодзи для оценки. 7=отлично: 6-7🟢, 4-5🟡, 1-3🔴."""
    if score is None:
        return "❓"
    s = int(score)
    if s >= 6: return "🟢"
    elif s >= 4: return "🟡"
    return "🔴"


# ==================== ДЕКОРАТОРЫ ====================

def admin_only(method):
    """Декоратор: доступ только для admin/doctor. Показывает alert при отказе."""
    async def wrapper(self, update, ctx):
        uid = None
        if update.callback_query:
            uid = update.callback_query.from_user.id
        elif update.effective_user:
            uid = update.effective_user.id
        if uid and not self._is_full_access(uid):
            q = update.callback_query
            if q:
                try:
                    await q.answer("🔒 Нет доступа", show_alert=True)
                except Exception:
                    pass
            return
        return await method(self, update, ctx)
    wrapper.__name__ = method.__name__
    wrapper.__doc__ = method.__doc__
    return wrapper


# ==================== БОТ ====================

class SportHealthBot:
    def __init__(self):
        self.db = Database()
        self.user_states: Dict[int, Dict[str, Any]] = {}
        self.rate_limiter = defaultdict(list)
        self.job_queue = None
        self._state_ttl = 3 * 24 * 3600  # TTL состояний/сессий (3 дня) — защита от утечки памяти

    def get_state(self, uid):
        if uid not in self.user_states:
            self.user_states[uid] = {"step": None, "data": {}, "_seen": time.time()}
        else:
            self.user_states[uid]["_seen"] = time.time()
        return self.user_states[uid]

    def clear_state(self, uid):
        self.user_states.pop(uid, None)
        # асинхронно удаляем из БД (fire-and-forget)
        try:
            import asyncio as _asio
            loop = _asio.get_event_loop()
            if loop.is_running():
                loop.run_in_executor(None, lambda: self.db.delete_user_state(uid))
        except Exception:
            pass

    async def _db_run(self, fn, *a, **k):
        """Выполнить синхронный БД-вызов в executor, не блокируя event loop."""
        import asyncio
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, lambda: fn(*a, **k))

    async def restore_user_state(self, uid):
        """При старте (cmd_start) восстановить сессию из БД в RAM, если её ещё нет."""
        if uid in self.user_states:
            return self.user_states[uid]
        try:
            import json as _json
            payload = await self._db_run(self.db.load_user_state, uid)
            if payload:
                data = _json.loads(payload)
                data["_seen"] = time.time()
                self.user_states[uid] = data
                return self.user_states[uid]
        except Exception as e:
            logger.warning(f"restore_user_state: {e}")
        self.user_states[uid] = {"step": None, "data": {}, "_seen": time.time()}
        return self.user_states[uid]

    async def persist_user_state(self, uid):
        """Сохранить RAM-сессию в БД (защита от потери при рестарте)."""
        st = self.user_states.get(uid)
        if not st:
            return
        # не храним служебные поля
        payload = {k: v for k, v in st.items() if k != "_seen"}
        try:
            await self._db_run(self.db.save_user_state, uid, payload)
        except Exception as e:
            logger.warning(f"persist_user_state: {e}")

    async def _cleanup_stale_sessions(self, context=None):
        """Периодическая очистка неактивных сессий (TTL = _state_ttl)."""
        # JobQueue (PTB 22.8) вызывает callback через await — метод обязан быть async,
        # иначе каждые 30 мин падает TypeError: object NoneType can't be used in 'await' expression.
        # Внутри тело sync-совместимое: все операции локальные (dict/time), await не нужен.
        now = time.time()
        stale = [uid for uid, st in list(self.user_states.items())
                 if now - st.get("_seen", now) > self._state_ttl]
        for uid in stale:
            self.user_states.pop(uid, None)
        # rate_limiter: отсекаем записи, чей последний запрос старше 2x window (по сути TTL)
        for uid in list(self.rate_limiter.keys()):
            lst = [t for t in self.rate_limiter[uid] if now - t < RATE_LIMIT_WINDOW]
            if not lst:
                self.rate_limiter.pop(uid, None)
            else:
                self.rate_limiter[uid] = lst
        if stale or len(self.rate_limiter) > 5000:
            logger.info(f"TTL cleanup: удалено сессий={len(stale)}")

    def check_rate_limit(self, uid):
        now = time.time()
        self.rate_limiter[uid] = [t for t in self.rate_limiter[uid] if now - t < RATE_LIMIT_WINDOW]
        if len(self.rate_limiter[uid]) >= RATE_LIMIT_MESSAGES:
            return False
        self.rate_limiter[uid].append(now)
        return True

    def _first_name(self, athlete):
        """Имя для приветствия: хранимое first_name (Telegram) или эвристика по full_name.

        full_name в БД смешанный: импортированные «Фамилия Имя», через бота «Имя Фамилия»,
        поэтому эвристика ненадёжна — приоритет у колонки first_name (из профиля Telegram)."""
        if not athlete:
            return ""
        if isinstance(athlete, dict):
            fn = athlete.get("first_name")
            if fn:
                return fn
            full = athlete.get("full_name", "")
        else:
            full = str(athlete)
        if not full:
            return ""
        parts = str(full).split()
        return parts[-1] if len(parts) > 1 else str(full)

    def _looks_like_name(self, name):
        """Имя из Telegram похоже на настоящее (не ник-хэндл типа mfilkov).

        Принимаем: содержит кириллицу (любой регистр) либо начинается с заглавной буквы."""
        if not name:
            return False
        return bool(re.search(r"[А-Яа-яЁё]", name)) or (name[0].isupper() and len(name) >= 2)

    def _logo_file(self):
        """Открывает файл клубной эмблемы для отправки фото. None — если файла нет."""
        for p in ("logo_cbk.png",):
            if os.path.exists(p):
                try:
                    return open(p, "rb")
                except OSError:
                    return None
        return None

    def _add_logo_to_ws(self, ws, anchor="A1", size=90):
        """Вставляет эмблему клуба на лист Excel. Нет файла — ничего не делает."""
        try:
            if not os.path.exists("logo_cbk.png"):
                return
            img = XLImage("logo_cbk.png")
            # эмблема 480x480 квадратная — уменьшаем до size px, сохраняя пропорции
            img.width = size
            img.height = size
            img.anchor = anchor
            ws.add_image(img)
        except Exception as e:
            logger.warning(f"logo insert failed: {e}")

    def kb(self, buttons):
        # Отфильтровываем пустые строки кнопок
        buttons = [row for row in buttons if row]
        return InlineKeyboardMarkup([
            [b if isinstance(b, InlineKeyboardButton) else InlineKeyboardButton(b[0], callback_data=b[1]) for b in row]
            for row in buttons
        ])

    def score_buttons(self, prefix, mn=1, mx=7, add_cancel=True):
        buttons, row = [], []
        for i in range(mn, mx + 1):
            row.append((f"{get_score_emoji(i)} {i}", f"srv_{prefix}_{i}"))
            if len(row) == 4:
                buttons.append(row)
                row = []
        if row:
            buttons.append(row)
        if add_cancel:
            buttons.append([("❌ Отмена", "main_menu")])
        return buttons

    # ==================== ГЛАВНОЕ МЕНЮ ====================

    async def _ask_gender_mandatory(self, update, ctx):
        """Запросить пол у пользователя — без этого не пускать в меню."""
        user = update.effective_user
        if not user:
            return
        text = (
            "⚠️ *Для продолжения нужно указать пол*\n\n"
            "Выбери ниже 👇"
        )
        buttons = self.kb([
            [("♂ Мужской", "gender_mandatory_male")],
            [("♀ Женский", "gender_mandatory_female")],
        ])
        if update.callback_query:
            await update.callback_query.edit_message_text(text, reply_markup=buttons, parse_mode="Markdown")
        else:
            await update.message.reply_text(text, reply_markup=buttons, parse_mode="Markdown")

    async def _gender_mandatory_chosen(self, update, ctx, gender):
        """Сохранить пол и показать главное меню."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if athlete:
            self.db.update_athlete_gender(athlete["id"], gender)
        await self.show_main_menu(update, ctx)

    async def show_main_menu(self, update, ctx):
        user = update.effective_user

        # Тренер — админ своей команды, опрос не проходит
        if self.db.is_coach(user.id):
            text = (
                f"🏀 *ЧБК — Мониторинг здоровья*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"👕 Тренерская панель — команда и отчёты:\n\n"
            )
            buttons = [
                [(f"👕 Моя команда", "coach_team_view")],
                [(f"📊 Мой отчёт (Excel)", "report_export_menu")],
            ]
            if self._is_full_access(user.id):
                buttons.insert(0, [(f"📋 Отчет для врача", "admin_report")])
                buttons.insert(1, [(f"⚙️ Управление", "admin_manage")])
                buttons.insert(2, [(f"⚖️ Весы и состав тела", "bc_menu")])
            buttons.append([(f"❓ Помощь", "help_menu")])
            if update.callback_query:
                await update.callback_query.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
            else:
                _logo = self._logo_file()
                if _logo is not None:
                    try:
                        await update.message.reply_photo(_logo)
                    except Exception as e:
                        logger.warning(f"logo in menu failed, fallback to text: {e}")
                    finally:
                        try: _logo.close()
                        except Exception: pass
                await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
            return

        athlete = self.db.get_athlete_by_telegram_id(user.id)
        if not athlete:
            await self.cmd_start(update, ctx)
            return

        if self.db.is_athlete_banned(athlete["id"]):
            text = "🔒 Твой аккаунт заблокирован. Обратись к администратору."
            if update.callback_query:
                await update.callback_query.edit_message_text(text, reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            else:
                await update.message.reply_text(text, reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            return

        # Если пол не указан — блокируем меню
        if not athlete.get("gender"):
            await self._ask_gender_mandatory(update, ctx)
            return

        today = self.db.get_survey_today(athlete["id"])
        has_survey = today is not None
        streak = athlete.get("survey_streak", 0)
        hour = datetime.now().hour
        greet = "Доброе утро" if hour < 12 else ("Добрый день" if hour < 18 else "Добрый вечер")

        text = (
            f"🏀 *ЧБК — Мониторинг здоровья*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👤 *{athlete['full_name']}* | {athlete['age_group']}\n"
            f"🏀 Команда: *{athlete.get('team', '?')}*\n"
            f"🔥 Серия: *{streak} дней* | {get_rank(streak)}\n"
            f"📊 Всего опросов: *{athlete.get('total_surveys', 0)}*\n\n"
        )

        if has_survey:
            text += (
                f"✅ *Опрос за сегодня пройден!*\n\n"
                f"😴 Сон: {score_bar(today.get('sleep_score', 0))}\n"
                f"😩 Утомление: {score_bar(today.get('fatigue_score', 0))}\n"
            )
            if today.get("resting_hr"):
                text += f"❤️ Пульс: {today['resting_hr']} уд/мин\n"
        else:
            text += f"⚠️ *Опрос за сегодня ещё не пройден!*\n"

        text += f"\n{greet}, {self._first_name(athlete)}! 💪"

        # Минимальное меню спортсмена (простые, интуитивные действия)
        q_done = self.db.has_questionnaire(athlete["id"]) and not self.db.has_incomplete_questionnaire(athlete["id"])
        buttons = [
            [(f"📝 Пройти опрос" if not has_survey else f"✅ Сегодня пройден",
              "do_survey" if not has_survey else "view_today")],
            [(f"📈 Мой прогресс", "my_progress")],
            [(f"🎯 Мои цели", "my_goals")],
            [(f"📊 Мои показатели", "my_stats")],
            [(f"⌚ Данные часов", "watch_data")],
            ([] if q_done else [(f"📋 Заполнить анкету", "questionnaire")]),
            [(f"📅 Запись к врачу", "consultation_start")],
            [(f"❓ Помощь", "help_menu")],
        ]

        # Админ-блок (виден только врачу/тебе)
        if self._is_full_access(user.id):
            buttons.insert(0, [(f"📋 Отчет для врача", "admin_report")])
            buttons.insert(1, [(f"📊 Экспорт в Excel", "report_export_menu")])
            buttons.insert(2, [(f"⚖️ Весы и состав тела", "bc_menu")])
            buttons.insert(3, [(f"⚙️ Управление", "admin_manage")])

        if update.callback_query:
            await update.callback_query.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        else:
            # Вход из /start: эмблема отдельной карточкой + текстовое меню (которое можно редактировать)
            _logo = self._logo_file()
            if _logo is not None:
                try:
                    await update.message.reply_photo(_logo)
                except Exception as e:
                    logger.warning(f"logo in menu failed, fallback to text: {e}")
                finally:
                    try: _logo.close()
                    except Exception: pass
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    # ==================== СТАРТ / РЕГИСТРАЦИЯ ====================

    async def cmd_start(self, update, ctx):
        user = update.effective_user
        logger.info(f"cmd_start called by {user.id} ({user.username})")
        # Восстановить сессию из БД (переживает рестарт), если существует
        try:
            await self.restore_user_state(user.id)
        except Exception:
            pass
        # Тренер — не спортсмен: сразу панель тренера, без регистрации спортсмена
        if self.db.is_coach(user.id):
            await self.show_main_menu(update, ctx)
            return
        athlete = self.db.get_athlete_by_telegram_id(user.id)
        if athlete:
            if self.db.is_athlete_banned(athlete["id"]):
                if update.callback_query:
                    await update.callback_query.edit_message_text("🔒 Твой аккаунт заблокирован. Обратись к администратору.", reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]))
                else:
                    await update.message.reply_text("🔒 Твой аккаунт заблокирован. Обратись к администратору.", reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]))
                return
            # Само-синхронизация имени из Telegram (только если похоже на имя, не ник)
            if user.first_name and self._looks_like_name(user.first_name) and athlete.get("first_name") != user.first_name:
                self.db.set_athlete_first_name(athlete["id"], user.first_name)
                athlete["first_name"] = user.first_name
            await self.show_main_menu(update, ctx)
            return

        # Согласие на обработку ПДн (152-ФЗ) — обязательно до регистрации
        if not self.db.has_consent(user.id):
            consent_text = (
                "📄 *Согласие на обработку персональных данных*\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "Бот собирает данные о твоём здоровье (опросы самочувствия, анкета, данные часов) "
                "для медицинского сопровождения команды. Данные хранятся на сервере и доступны только тренеру и врачу.\n\n"
                "Нажимая кнопку ниже, ты подтверждаешь согласие на их обработку."
            )
            consent_kb = self.kb([[(f"✅ Согласен", "consent_accept")],
                                  [(f"❌ Не согласен", "consent_decline")]])
            if update.callback_query:
                await update.callback_query.edit_message_text(consent_text, reply_markup=consent_kb, parse_mode="Markdown")
            else:
                await update.message.reply_text(consent_text, reply_markup=consent_kb, parse_mode="Markdown")
            return

        await self._ask_age_group(update, ctx)

    async def _ask_age_group(self, update, ctx):
        """Показать выбор возрастной группы при регистрации.

        Работает и из команды /start (update.message), и из callback
        (consent_accept / start_reg — update.message is None)."""
        text = (
            f"🏀 *Добро пожаловать в ЧБК СпортМед!*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Я — твой помощник для мониторинга здоровья.\n"
            f"Каждый день опрос → тренды → рекомендации врача.\n\n"
            f"*Что я умею:*\n"
            f"📝 Ежедневный опрос\n"
            f"📈 Графики и тренды\n"
            f"🏥 Рекомендации спортивного врача\n"
            f"⌚ Импорт данных с часов\n"
            f"🏆 Достижения\n\n"
            f"Выбери свою возрастную группу:"
        )
        buttons = [[(desc, f"reg_{key}")] for key, desc in AGE_GROUPS.items()]

        if update.callback_query:
            # Пришли из кнопки «Согласен» — редактируем текущее сообщение, не плодя новые
            await update.callback_query.edit_message_text(
                text, reply_markup=self.kb(buttons), parse_mode="Markdown"
            )
            return

        _logo = self._logo_file()
        if _logo is not None:
            try:
                # Эмблема — отдельной карточкой (без кнопок), чтобы текстовое меню было редактируемым
                await update.message.reply_photo(_logo)
            except Exception as e:
                logger.warning(f"logo in /start failed, fallback to text: {e}")
            finally:
                try: _logo.close()
                except Exception: pass
        await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def reg_callback(self, update, ctx):
        q = update.callback_query
        await q.answer()
        age = q.data.replace("reg_", "")
        state = self.get_state(q.from_user.id)
        state["step"] = "reg_team"
        state["data"]["age_group"] = age

        protocol = "Simple" if age in SIMPLE_PROTOCOLS else "Full"
        buttons = [[(team, f"team_{team}")] for team in TEAMS]
        await q.edit_message_text(
            f"✅ Группа: *{AGE_GROUPS[age]}*\n📋 Протокол: {protocol}\n\nВыбери *команду*:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def team_callback(self, update, ctx):
        q = update.callback_query
        await q.answer()
        team = q.data.replace("team_", "")
        state = self.get_state(q.from_user.id)
        state["step"] = "reg_gender"
        state["data"]["team"] = team

        await q.edit_message_text(
            f"✅ Команда: *{team}*\n\n👤 *Укажите пол:*",
            reply_markup=self.kb([
                [("♂ Мужской", "reg_gender_male")],
                [("♀ Женский", "reg_gender_female")],
            ]), parse_mode="Markdown"
        )

    async def reg_gender_callback(self, update, ctx):
        q = update.callback_query
        await q.answer()
        gender = q.data.replace("reg_gender_", "")
        state = self.get_state(q.from_user.id)
        state["step"] = "reg_name"
        state["data"]["gender"] = gender
        await q.edit_message_text(
            f"✅ Пол: {'♂ Мужской' if gender == 'male' else '♀ Женский'}\n\nВведи *Фамилию Имя*:",
            parse_mode="Markdown"
        )

    # ==================== СБРОС АККАУНТА ====================

    async def reset_account(self, update, ctx):
        q = update.callback_query
        await q.answer()
        await q.edit_message_text(
            "⚠️ *Сбросить аккаунт?*\n\nВсе данные будут удалены. Можно пройти регистрацию заново.",
            reply_markup=self.kb([
                [("✅ Да, сбросить", "confirm_reset")],
                [("❌ Нет", "main_menu")]
            ]), parse_mode="Markdown"
        )

    async def confirm_reset(self, update, ctx):
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if athlete:
            self.db.conn.execute("DELETE FROM daily_wellness WHERE athlete_id = ?", (athlete["id"],))
            self.db.conn.execute("DELETE FROM athletes WHERE id = ?", (athlete["id"],))
            self.db.conn.commit()
        await q.edit_message_text(
            "✅ Аккаунт сброшен!\n\nНапиши /start для регистрации заново.",
            parse_mode="Markdown"
        )

    # ==================== ДАННЫЕ ЧАСОВ ====================

    async def watch_data_menu(self, update, ctx):
        q = update.callback_query
        await q.answer()
        text = (
            f"⌚ *Импорт данных с часов*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Выбери свои часы — я покажу как выгрузить файл:"
        )
        buttons = [
            [("⌚ Garmin", "watch_garmin"), ("🍏 Apple Watch", "watch_apple")],
            [("⌚ Fitbit / Google", "watch_fitbit"), ("⌚ Xiaomi / Zepp", "watch_xiaomi")],
            [("⌚ Huawei", "watch_huawei"), ("⌚ Samsung", "watch_samsung")],
            [("📄 Любые часы (TXT/CSV)", "watch_other")],
        ]
        if self._is_admin_or_coach(q.from_user.id):
            buttons.append([("⚖️ Весы и состав тела", "bc_menu")])
        buttons.append([(f"🔙 Назад", "main_menu")])
        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def _watch_brand_instructions(self, update, ctx, brand):
        q = update.callback_query
        await q.answer()

        instructions = {
            "garmin": (
                "⌚ *Garmin*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой Garmin Connect на телефоне\n"
                "2. Нажми ⋮ → *Экспорт в CSV*\n"
                "3. Выбери период (1 день)\n"
                "4. Отправь полученный CSV-файл сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги, стресс, SpO2\n\n"
                "*Просто отправь файл!*"
            ),
            "apple": (
                "🍏 *Apple Watch*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой *Здоровье* на iPhone\n"
                "2. Профиль (🟣) → *Экспорт всех данных*\n"
                "3. Дождись генерации .zip\n"
                "4. Отправь файл сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги, SpO2\n\n"
                "*Просто отправь файл!*"
            ),
            "fitbit": (
                "⌚ *Fitbit / Google Pixel Watch*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой Fitbit App\n"
                "2. Профиль → *Настройки* → *Экспорт данных*\n"
                "3. Запроси CSV-экспорт\n"
                "4. Когда файл придёт на почту — отправь его сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги, стресс\n\n"
                "*Просто отправь файл!*"
            ),
            "xiaomi": (
                "⌚ *Xiaomi / Zepp / Amazfit*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой *Zepp Life* или *Mi Fitness*\n"
                "2. Профиль → *Настройки* → *Экспорт данных*\n"
                "3. Выбери период и формат CSV\n"
                "4. Отправь полученный файл сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги, стресс, SpO2\n\n"
                "*Просто отправь файл!*"
            ),
            "huawei": (
                "⌚ *Huawei Watch / Band*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой *Huawei Health*\n"
                "2. Я → *Настройки* → *Экспорт данных*\n"
                "3. Выбери CSV и отправь себе\n"
                "4. Перешли файл сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги, стресс, SpO2\n\n"
                "*Просто отправь файл!*"
            ),
            "samsung": (
                "⌚ *Samsung Galaxy Watch*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📤 *Как выгрузить:*\n"
                "1. Открой *Samsung Health*\n"
                "2. ⋮ → *Настройки* → *Экспорт данных*\n"
                "3. Выбери JSON-формат\n"
                "4. Отправь файл сюда\n\n"
                "✅ Поддерживается: пульс, сон, шаги\n\n"
                "*Просто отправь файл!*"
            ),
        }

        text = instructions.get(brand,
            "📄 *Любые часы*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Отправь файл с данными в формате CSV, JSON или TXT.\n\n"
            "Файл должен содержать хотя бы один из параметров:\n"
            "пульс, сон, шаги, стресс, SpO2\n\n"
            "Попробуй экспорт из приложения твоих часов."
        )

        await q.edit_message_text(
            text + "\n\n━━━━━━━━━━━━━━━━━━━━━━━━━━",
            reply_markup=self.kb([[(f"🔙 К выбору часов", "watch_data")],
                                  [(f"🔙 Главное меню", "main_menu")]]),
            parse_mode="Markdown"
        )

    async def handle_document(self, update, ctx):
        user_id = update.effective_user.id
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        # Врач/тренер может присылать CSV весов, даже если сам не спортсмен.
        if not athlete and not self._is_full_access(user_id):
            await update.message.reply_text("❌ Сначала зарегистрируйся: /start")
            return

        doc = update.message.document
        if not doc:
            return

        fname = doc.file_name.lower() if doc.file_name else ""
        supported = [".csv", ".json", ".txt", ".zip"]
        if not any(fname.endswith(ext) for ext in supported):
            await update.message.reply_text("❌ Формат не поддерживается. Отправь CSV, JSON, TXT или ZIP (Apple Health).")
            return

        # Лимит размера загружаемого файла (ZIP от Apple Health может быть до 50 МБ)
        MAX_FILE_BYTES = 50 * 1024 * 1024 if fname.endswith(".zip") else 5 * 1024 * 1024
        if getattr(doc, "file_size", None) and doc.file_size > MAX_FILE_BYTES:
            mb = round(doc.file_size / (1024 * 1024), 1)
            await update.message.reply_text(f"❌ Файл слишком большой ({mb} МБ). Максимум 5 МБ.")
            return

        await update.message.reply_text("📥 Обрабатываю файл...")

        try:
            file = await doc.get_file()
            content = await file.download_as_bytearray()
            text_content = content.decode("utf-8", errors="replace")

            # === БИОИМПЕДАНСНЫЕ ВЕСЫ: пробуем распознать CSV состава тела ===
            scale_records = parse_scale_csv_dated(text_content, fname)
            if scale_records and any(
                r.get("weight_kg") is not None or r.get("body_fat_pct") is not None
                for r in scale_records
            ):
                await self._import_scale_csv(update, ctx, scale_records, fname)
                return

            # === Умные часы: обычный путь ===
            watch_data = parse_watch(text_content, fname, raw_bytes=bytes(content))

            if not watch_data:
                await update.message.reply_text(
                    "❌ Не удалось распознать данные. Проверь формат файла.\n\n"
                    "Попробуй выбрать свои часы в меню ⌚ и отправить файл по инструкции.",
                    reply_markup=self.kb([[(f"⌚ Выбрать часы", "watch_data")],
                                          [(f"🏠 Главное меню", "main_menu")]])
                )
                return

            msg = self._format_watch_report(watch_data, athlete["age_group"])

            # Сохраняем данные часов в БД (Фаза 5) — питают рекомендации (HRV, вес, сон)
            watch_data["_source"] = fname
            self.db.save_watch_data(athlete["id"], watch_data)

            await update.message.reply_text(
                msg, reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]), parse_mode="Markdown"
            )
        except Exception as e:
            logger.error(f"Watch file error: {e}")
            await update.message.reply_text("❌ Ошибка при обработке файла.")

    async def _import_scale_csv(self, update, ctx, records, fname):
        """Импорт CSV данных биоимпедансных весов (врач/тренер — группа, спортсмен — свой)."""
        user_id = update.effective_user.id
        full_access = self._is_full_access(user_id)

        if not full_access:
            # спортсмен прислал свой файл — сохраняем на него (только одна точка замеров у него)
            athlete = self.db.get_athlete_by_telegram_id(user_id)
            if not athlete:
                await update.message.reply_text("❌ Сначала зарегистрируйся: /start")
                return
            saved = 0
            for r in records:
                data = dict(r)
                data["recorded_by"] = user_id
                # рост из анкеты
                try:
                    qd = self.db.get_questionnaire(athlete["id"])
                    if qd and qd.get("height"):
                        data.setdefault("height_cm", qd.get("height"))
                except Exception:
                    pass
                if self.db.save_body_composition(athlete["id"], r.get("record_date"), data,
                                                 source="csv", device_profile=r.get("device_profile")):
                    saved += 1
            if saved:
                # Показываем полный анализ + уведомления
                await self._analyze_and_notify_scale(update, athlete, records)
            else:
                await update.message.reply_text(
                    "⚠️ Не удалось сохранить данные. Проверь формат файла.",
                    reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
                )
            return

        # Врач/тренер: группа — раскладываем по профилям
        saved, skipped, unmapped, rows = await self._import_scale_records(
            update, ctx, records, forced_athlete_id=None)

        head = ["⚖️ *Импорт весов (CSV)*", "━━━━━━━━━━━━━━━━━━━━━━━━━━"]
        if rows:
            head.append(f"✅ Сохранено: *{saved}* записей")
            head.append("")
            head.extend(rows[:12])
            if len(rows) > 12:
                head.append(f"… и ещё {len(rows)-12}")
            head.append("")
        if skipped:
            head.append(f"⚠️ Пропущено: *{skipped}* (без привязки к спортсмену)")
        if unmapped:
            head.append("")
            head.append(f"*Не закреплённые профили:* {', '.join(unmapped)}")
            head.append("Открой ниже «Профили весов» и назначь спортсменов, затем отправь CSV снова.")

        # Уведомления спортсменам, тренерам и врачу по каждому сохранённому замеру
        if saved:
            mapping = self.db.get_scale_profiles()
            notified_athletes = set()
            for r in records:
                prof = r.get("device_profile")
                if not prof:
                    continue
                athlete_id = mapping.get(prof)
                if not athlete_id or athlete_id in notified_athletes:
                    continue
                notified_athletes.add(athlete_id)
                athlete = self.db.get_athlete_by_id(athlete_id)
                if not athlete:
                    continue
                # Уведомление спортсмену
                try:
                    athlete_text = self._build_athlete_bc_report(
                        self.db.get_latest_body_composition(athlete_id),
                        self.db.get_body_composition(athlete_id, days=90),
                        athlete
                    )
                    await self._send_admin(athlete["telegram_id"], athlete_text)
                except Exception as e:
                    logger.error(f"Athlete BC notify error: {e}")
                # Уведомление тренеру
                team = athlete.get("team")
                if team:
                    try:
                        coach_text = self._build_coach_bc_report(
                            self.db.get_latest_body_composition(athlete_id),
                            self.db.get_body_composition(athlete_id, days=90),
                            athlete
                        )
                        coaches = self.db.get_team_coaches(team)
                        for cid in coaches:
                            if cid != athlete.get("telegram_id"):
                                await self._send_admin(cid, coach_text)
                    except Exception as e:
                        logger.error(f"Coach BC notify error: {e}")

        # запрос роста, если есть куда сохранять, но нет роста (пришлёт сводку после ввода)
        if saved or (skipped and not unmapped):
            asked_height = await self._ask_height_for_records(update, ctx, records)
            if asked_height:
                return

        await update.message.reply_text(
            "\n".join(head),
            reply_markup=self.kb([
                [(f"🗂 Профили весов", "bc_profiles")],
                [(f"🏠 Главное меню", "main_menu")]
            ]),
            parse_mode="Markdown"
        )

    def _format_watch_report(self, data, age_group):
        """Красивое форматирование данных с часов."""
        lines = [
            "⌚ *📊 Данные с часов*",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━"
        ]

        norms = HR_NORMS.get(age_group, {"min": 40, "max": 70})

        fields = [
            ("💓 Пульс покоя", data.get("💓 Пульс"),
             f"Норма {age_group}: {norms['min']}-{norms['max']} уд/мин"),
            ("😴 Сон", data.get("😴 Сон"),
             "Норма: 7-9 часов"),
            ("🏃 Шаги", data.get("🏃 Шаги"),
             "Норма: 8000-12000"),
            ("😰 Стресс", data.get("😰 Стресс"),
             "0-100: <50 низкий, 50-70 средний, >70 высокий"),
            ("🫁 SpO₂", data.get("🫁 SpO2"),
             "Норма: 95-99%"),
            ("📊 HRV", data.get("📊 HRV"),
             "Выше = лучше восстановление"),
        ]

        has_data = False
        for label, value, norm in fields:
            if value is not None:
                has_data = True
                emoji = ""
                if label.startswith("💓") and isinstance(value, (int, float)):
                    if value > norms["max"] + 10:
                        emoji = "🔴"
                    elif value > norms["max"]:
                        emoji = "🟡"
                    else:
                        emoji = "🟢"
                elif label.startswith("😴"):
                    try:
                        h = float(str(value).replace("ч", ""))
                        emoji = "🔴" if h < 6 else ("🟡" if h < 7 else "🟢")
                    except:
                        pass
                elif label.startswith("🏃"):
                    if isinstance(value, int):
                        emoji = "🟡" if value < 5000 else "🟢"
                elif label.startswith("😰"):
                    if isinstance(value, int):
                        emoji = "🔴" if value > 70 else ("🟡" if value > 50 else "🟢")

                lines.append(f"\n{emoji} *{label}:* {value}")
                lines.append(f"     {norm}")

        if not has_data:
            return "❌ Нет распознанных данных."

        # Сохраняем данные в БД для аналитики
        try:
            hr_val = data.get("💓 Пульс")
            if hr_val and isinstance(hr_val, (int, float)):
                pass
        except Exception:
            pass

        return "\n".join(lines)

    # ==================== БИОИМПЕДАНСНЫЕ ВЕСЫ ====================

    async def scale_import_menu(self, update, ctx):
        """Пункт меню врача/тренера — веса и состав тела."""
        q = update.callback_query
        await q.answer()
        if not self._is_admin_or_coach(q.from_user.id):
            await q.edit_message_text("🔒 Нет доступа.", reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]))
            return
        await q.edit_message_text(
            "⚖️ *Весы и состав тела*\n\n"
            "Весы GARLYN Bodyscan Master (MovingLife).\n\n"
            "1️⃣ Взвесь группу на весах (до 50 профилей).\n"
            "2️⃣ Открой MovingLife → *Экспорт истории* → CSV.\n"
            "3️⃣ Отправь *CSV-файл сюда* — бот разложит записи по спортсменам.\n\n"
            "🔹 *Профили весов* — закрепи «Профиль N» за спортсменом, чтобы бот раскладывал автоматически.\n"
            "🔹 Замеры за один день перезаписываются (хранится последний).",
            reply_markup=self.kb([
                [(f"📈 Кто взвешился", "bc_overview")],
                [(f"🗂 Профили весов", "bc_profiles")],
                [(f"🔙 Назад", "main_menu")]
            ])
        )

    async def bc_overview(self, update, ctx):
        """Обзор: кто из спортсменов имеет данные состава тела."""
        q = update.callback_query
        await q.answer()
        if not self._is_admin_or_coach(q.from_user.id):
            return

        # Получаем всех спортсменов с данными состава тела
        athletes_with_data = self.db.get_bc_period(days=999)
        if not athletes_with_data:
            await q.edit_message_text(
                "📈 *Кто взвешился*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "Пока ни один спортсмен не взвешивался.\n"
                "Отправь CSV-файл весов для импорта.",
                reply_markup=self.kb([[(f"🔙 Назад", "bc_menu")]])
            )
            return

        # Группируем по спортсменам: последний замер + количество
        latest_by_athlete = {}
        for row in athletes_with_data:
            aid = row.get("athlete_id")
            if aid not in latest_by_athlete:
                latest_by_athlete[aid] = {
                    "name": row.get("full_name", "?"),
                    "team": row.get("team", "?"),
                    "date": row.get("record_date", "?"),
                    "weight": row.get("weight_kg"),
                    "fat": row.get("body_fat_pct"),
                    "muscle": row.get("muscle_mass_kg"),
                    "water": row.get("body_water_pct"),
                    "count": 0
                }
            latest_by_athlete[aid]["count"] += 1

        lines = [
            f"📈 *Кто взвешился* ({len(latest_by_athlete)} спортсменов)",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        ]

        # Группируем по командам
        by_team = {}
        for aid, info in latest_by_athlete.items():
            team = info["team"]
            by_team.setdefault(team, []).append((aid, info))

        for team in sorted(by_team.keys()):
            lines.append(f"*{team}*")
            for aid, info in sorted(by_team[team], key=lambda x: x[1]["name"]):
                weight = f"{info['weight']:.1f}кг" if info['weight'] else "—"
                fat = f"{info['fat']:.1f}%" if info['fat'] else "—"
                muscle = f"{info['muscle']:.1f}кг" if info['muscle'] else "—"
                lines.append(
                    f"  👤 {info['name']} | 📅 {info['date']} | "
                    f"⚖️ {weight} | жир {fat} | мышцы {muscle} | замеров: {info['count']}"
                )
            lines.append("")

        # Кнопки для просмотра деталей
        btns = []
        for aid, info in sorted(latest_by_athlete.items(), key=lambda x: x[1]["name"]):
            btns.append([(f"📊 {info['name']}", f"bc_view_{aid}")])
        btns.append([("🔙 Назад", "bc_menu")])

        # Ограничиваем количество кнопок (Telegram лимит ~100)
        if len(btns) > 50:
            btns = btns[:50]
            lines.append(f"\n_… и ещё {len(latest_by_athlete) - 50} спортсменов_")

        await q.edit_message_text(
            "\n".join(lines),
            reply_markup=self.kb(btns),
            parse_mode="Markdown"
        )

    async def bc_show_profiles(self, update, ctx, page=0):
        """Показать/закрепить сопоставление профилей весов → спортсмены (постранично)."""
        q = update.callback_query
        await q.answer()
        if not self._is_admin_or_coach(q.from_user.id):
            return
        mapping = self.db.get_scale_profiles()
        athletes = {a["id"]: a["full_name"] for a in self.db.get_all_athletes()}
        lines = ["🗂 *Профили весов → спортсмены*\n"]
        if mapping:
            for prof, aid in sorted(mapping.items()):
                name = athletes.get(aid, f"id{aid}")
                lines.append(f"{prof} → *{name}*")
        else:
            lines.append("Пока не закреплено. Выбери профиль и назначь спортсмена.")
        lines.append("\nВыбери профиль, чтобы назначить/изменить спортсмена:")

        total_pages = -(-SCALE_PROFILE_COUNT // BC_PROFILE_PAGE)
        page = max(0, min(page, total_pages - 1))
        s = page * BC_PROFILE_PAGE + 1
        e = min(page * BC_PROFILE_PAGE + BC_PROFILE_PAGE, SCALE_PROFILE_COUNT) + 1
        btns = [[(f"Профиль {n}", f"bc_map_{n}") for n in range(row, min(row + 4, e))]
                for row in range(s, e, 4)]
        nav = []
        if page > 0:
            nav.append((f"◀ {page}", f"bc_page_{page - 1}"))
        if page < total_pages - 1:
            nav.append((f"{page + 2} ▶", f"bc_page_{page + 1}"))
        if nav:
            btns.append(nav)
        btns.append([("🔙 Назад", "bc_menu")])
        await q.edit_message_text("\n".join(lines), reply_markup=self.kb(btns))

    async def bc_page(self, update, ctx):
        """Страница списка профилей весов."""
        q = update.callback_query
        page = max(0, int(q.data.replace("bc_page_", "")))
        await self.bc_show_profiles(update, ctx, page=page)

    async def bc_assign_pick(self, update, ctx):
        """После выбора профиля — список спортсменов для назначения."""
        q = update.callback_query
        await q.answer()
        if not self._is_admin_or_coach(q.from_user.id):
            return
        profile = int(q.data.replace("bc_map_", ""))
        state = self.get_state(q.from_user.id)
        state["data"]["bc_map_profile"] = profile
        athletes = self._scoped_athletes(q.from_user.id)
        if not athletes:
            await q.edit_message_text("Нет спортсменов в доступе.", reply_markup=self.kb([[(f"🔙 Назад", "bc_profiles")]]))
            return
        btns = []
        for a in athletes[:30]:
            btns.append([(f"{a['full_name']}", f"bc_assign_{profile}_{a['id']}")])
        btns.append([("🔙 Назад", "bc_profiles")])
        await q.edit_message_text(f"Профиль *{profile}* → выбери спортсмена:", reply_markup=self.kb(btns))

    async def bc_assign_save(self, update, ctx):
        """Закрепить профиль N за спортсменом."""
        q = update.callback_query
        if not self._is_admin_or_coach(q.from_user.id):
            return
        parts = q.data.split("_")
        profile = int(parts[2]); athlete_id = int(parts[3])
        self.db.set_scale_profile(f"Профиль {profile}", athlete_id, q.from_user.id)
        await q.answer("✅ Профиль закреплён", show_alert=True)
        await self.bc_show_profiles(update, ctx)

    async def bc_view_athlete(self, update, ctx):
        """Карточка состава тела спортсмена (только врач/тренер), расширенные метрики."""
        q = update.callback_query
        await q.answer()
        if not self._is_admin_or_coach(q.from_user.id):
            return
        athlete_id = int(q.data.replace("bc_view_", ""))
        history = self.db.get_body_composition(athlete_id, days=30)
        if not history:
            name = "спортсмен"
            try:
                a = self.db.get_athlete_by_id(athlete_id)
                if a:
                    name = a.get("full_name", name)
            except Exception:
                pass
            await q.edit_message_text(
                f"⚖️ *{name}* — пока нет данных составов весов.\n\nОтправь CSV из MovingLife (меню ⚖️ Весы) или взвесь профиль.",
                reply_markup=self.kb([[(f"🔙 Назад", "athlete_list")]])
            )
            return
        name = "спортсмен"
        try:
            a = self.db.get_athlete_by_id(athlete_id)
            if a:
                name = a.get("full_name", name)
        except Exception:
            pass
        lines = [f"⚖️ *Состав тела — {name}*", "━━━━━━━━━━━━━━━━━━━━━━━━━━"]
        latest = history[0]
        metrics = [
            ("Вес", latest.get("weight_kg"), "кг"),
            ("ИМТ", latest.get("bmi"), ""),
            ("Жир", latest.get("body_fat_pct"), "%"),
            ("Мышцы", latest.get("muscle_mass_kg"), "кг"),
            ("Вода", latest.get("body_water_pct"), "%"),
            ("Кости", latest.get("bone_mass_kg"), "кг"),
            ("Висц. жир", latest.get("visceral_fat_index"), ""),
            ("Подкож. жир", latest.get("subcutaneous_fat_pct"), "%"),
            ("Безжировая", latest.get("lean_mass_kg"), "кг"),
            ("Белок", latest.get("protein_pct"), "%"),
            ("BMR", latest.get("bmr_kcal"), "ккал"),
        ]
        for label, v, unit in metrics:
            if v is not None:
                num = float(v)
                lines.append(f"• {label}: *{num:.1f}* {unit}".rstrip())
        lines.append(f"\n📅 Замеров за 30 дней: {len(history)}")
        lines.append(f"(последний: {latest.get('record_date')})")
        # тренд веса
        trend = self.db.get_bc_trend(athlete_id, 30)
        wt = trend.get("weight_kg", [])
        if len(wt) >= 2:
            d = wt[-1][1] - wt[0][1]
            lines.append(f"📈 Вес 30дн: {wt[0][1]:.1f} → {wt[-1][1]:.1f} кг ({d:+.1f})")
        await q.edit_message_text(
            "\n".join(lines),
            reply_markup=self.kb([[(f"🔙 Назад", "athlete_list")]])
        )

    def _scale_summary_line(self, r):
        """Одна строка сводки по замеру состава тела."""
        parts = []
        if r.get("weight_kg") is not None:
            parts.append(f"вес {r['weight_kg']:.0f}кг")
        if r.get("body_fat_pct") is not None:
            parts.append(f"жир {r['body_fat_pct']:.1f}%")
        if r.get("muscle_mass_kg") is not None:
            parts.append(f"мышцы {r['muscle_mass_kg']:.1f}кг")
        if r.get("body_water_pct") is not None:
            parts.append(f"вода {r['body_water_pct']:.0f}%")
        date_s = str(r.get("record_date") or "?")
        return f"  📅 {date_s}: " + ", ".join(parts) if parts else f"  📅 {date_s}: —"

    async def _import_scale_records(self, update, ctx, records, forced_athlete_id=None):
        """Разложить спарсенные CSV записи весов по спортсменам и сохранить.
        Возвращает (записано, пропущено, unmapped_profiles, rows_text)."""
        user_id = update.effective_user.id
        mapping = self.db.get_scale_profiles()
        saved = 0
        skipped = 0
        unmapped = []
        rows = []
        height_cache = {}
        for r in records:
            prof = r.get("device_profile")
            # athlete_id: либо явный (forced), либо через маппинг профиля
            athlete_id = forced_athlete_id
            if athlete_id is None and prof:
                athlete_id = mapping.get(prof)
            if athlete_id is None:
                # нет привязки — копим на маппинг
                if prof and prof not in unmapped:
                    unmapped.append(prof)
                skipped += 1
                continue
            # рост из анкеты (кэш)
            if athlete_id not in height_cache:
                try:
                    qd = self.db.get_questionnaire(athlete_id)
                    height_cache[athlete_id] = qd.get("height") if qd else None
                except Exception:
                    height_cache[athlete_id] = None
            data = dict(r)
            if height_cache.get(athlete_id):
                data.setdefault("height_cm", height_cache[athlete_id])
            data["recorded_by"] = user_id
            ok = self.db.save_body_composition(athlete_id, r.get("record_date"), data,
                                               source="csv", device_profile=prof)
            if ok:
                saved += 1
                rows.append(f"👤 {athlete_id}: {self._scale_summary_line(r)}")
            else:
                skipped += 1
        return saved, skipped, unmapped, rows

    # ==================== АНАЛИЗ СОСТАВА ТЕЛА + УВЕДОМЛЕНИЯ ====================

    def _get_bc_status(self, value, norms_key, gender="male"):
        """Получить статус показателя: ✅ норма, ⚠️ погранично, 🚨 отклонение."""
        if value is None:
            return ""
        norms = BC_NORMS.get(self._current_age_group, BC_NORMS.get("Pro", {}))
        gender_norms = norms.get(gender, norms.get("male", {}))
        norm_range = gender_norms.get(norms_key)
        if not norm_range:
            return ""
        low, high = norm_range
        if low <= value <= high:
            return "✅"
        elif value < low * 0.9 or value > high * 1.1:
            return "🚨"
        else:
            return "⚠️"

    def _get_bc_status_text(self, value, norms_key, gender="male", unit=""):
        """Получить строку показателя со статусом."""
        if value is None:
            return None
        status = self._get_bc_status(value, norms_key, gender)
        return f"{value:.1f}{unit} {status}"

    def _build_athlete_bc_report(self, latest, prev, athlete):
        """Построить сообщение анализа состава тела для спортсмена."""
        if not latest:
            return "⚖️ Нет данных для анализа."

        date_s = latest.get("record_date", "?")
        gender = athlete.get("gender", "male")
        age_group = athlete.get("age_group", "Pro")
        self._current_age_group = age_group

        lines = [
            f"⚖️ *Твой анализ состава тела*",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"📅 {date_s}\n",
            f"*📊 Ключевые показатели:*"
        ]

        # 5 ключевых метрик
        weight = latest.get("weight_kg")
        fat = latest.get("body_fat_pct")
        muscle = latest.get("muscle_mass_kg")
        water = latest.get("body_water_pct")
        bmr = latest.get("bmr_kcal")

        if weight is not None:
            lines.append(f"• Вес: *{weight:.1f}* кг")
        if fat is not None:
            fat_status = self._get_bc_status(fat, "body_fat_pct", gender)
            lines.append(f"• Жир: *{fat:.1f}%* {fat_status}")
        if muscle is not None:
            lines.append(f"• Мышцы: *{muscle:.1f}* кг")
        if water is not None:
            water_status = self._get_bc_status(water, "body_water_pct", gender)
            lines.append(f"• Вода: *{water:.0f}%* {water_status}")
        if bmr is not None:
            lines.append(f"• Базовый метаболизм: *{bmr}* ккал")

        # Тренд (если есть предыдущий замер)
        if prev and len(prev) >= 2:
            # prev[-1] — текущий, prev[0] — самый старый за период
            older = prev[0]
            lines.append(f"\n*📈 Изменения за период:*")
            if weight is not None and older.get("weight_kg") is not None:
                d = weight - older["weight_kg"]
                lines.append(f"• Вес: {older['weight_kg']:.1f} → {weight:.1f} ({d:+.1f})")
            if fat is not None and older.get("body_fat_pct") is not None:
                d = fat - older["body_fat_pct"]
                emoji = "✅" if abs(d) <= 2 else ("⚠️" if abs(d) <= 5 else "🚨")
                lines.append(f"• Жир: {older['body_fat_pct']:.1f}% → {fat:.1f}% ({d:+.1f}%) {emoji}")
            if muscle is not None and older.get("muscle_mass_kg") is not None:
                d = muscle - older["muscle_mass_kg"]
                emoji = "✅" if d >= -0.5 else ("⚠️" if d >= -1.5 else "🚨")
                lines.append(f"• Мышцы: {older['muscle_mass_kg']:.1f} → {muscle:.1f} ({d:+.1f} кг) {emoji}")
            if water is not None and older.get("body_water_pct") is not None:
                d = water - older["body_water_pct"]
                lines.append(f"• Вода: {older['body_water_pct']:.0f}% → {water:.0f}% ({d:+.0f}%)")

        # Рекомендации
        recs = []
        if water is not None and water < 55:
            recs.append("⚠️ Вода ниже нормы — пей больше воды!\nПей ~2.5-3 л воды в день, особенно перед тренировкой.")
        if fat is not None:
            fat_norms = BC_NORMS.get(age_group, BC_NORMS.get("Pro", {})).get(gender, {})
            fat_high = fat_norms.get("body_fat_pct", (0, 100))[1]
            if fat > fat_high:
                recs.append(f"⚠️ Жир выше нормы ({fat_high}%). Снизь калорийность рациона.")
        if muscle is not None and prev and len(prev) >= 2:
            older = prev[0]
            if older.get("muscle_mass_kg") and muscle - older["muscle_mass_kg"] < -1.5:
                recs.append("🚨 Потеря мышц! Обратись к врачу.")

        if recs:
            lines.append(f"\n*💡 Рекомендации:*")
            lines.extend(recs)
        else:
            lines.append(f"\n💡 *Всё в порядке! Продолжай в том же духе.*")

        lines.append(f"\n⚠️ _Это не медицинское заключение._")
        return "\n".join(lines)

    def _build_admin_bc_report(self, latest, prev, athlete):
        """Построить сообщение для врача (админа) — полный анализ."""
        if not latest:
            return "⚖️ Нет данных для анализа."

        date_s = latest.get("record_date", "?")
        name = athlete.get("full_name", "?")
        team = athlete.get("team", "?")
        age_group = athlete.get("age_group", "?")
        gender = athlete.get("gender", "male")
        self._current_age_group = age_group

        lines = [
            f"⚖️ *Новый замер — {name}*",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"🏀 {team} | {age_group} | {date_s}\n",
            f"*📊 Состав тела:*"
        ]

        # Все метрики
        weight = latest.get("weight_kg")
        fat = latest.get("body_fat_pct")
        muscle = latest.get("muscle_mass_kg")
        water = latest.get("body_water_pct")
        bone = latest.get("bone_mass_kg")
        visceral = latest.get("visceral_fat_index")
        bmr = latest.get("bmr_kcal")
        amr = latest.get("amr_kcal")

        parts = []
        if weight is not None:
            parts.append(f"Вес: *{weight:.1f}* кг")
        if fat is not None:
            parts.append(f"Жир: *{fat:.1f}%*")
        if muscle is not None:
            parts.append(f"Мышцы: *{muscle:.1f}* кг")
        if parts:
            lines.append(" | ".join(parts))

        parts2 = []
        if water is not None:
            parts2.append(f"Вода: *{water:.0f}%*")
        if bone is not None:
            parts2.append(f"Кости: *{bone:.2f}* кг")
        if visceral is not None:
            parts2.append(f"Висц. жир: *{visceral:.0f}*")
        if parts2:
            lines.append(" | ".join(parts2))

        parts3 = []
        if bmr is not None:
            parts3.append(f"BMR: *{bmr}* ккал")
        if amr is not None:
            parts3.append(f"AMR: *{amr}* ккал")
        if parts3:
            lines.append(" | ".join(parts3))

        # Тренд
        if prev and len(prev) >= 2:
            older = prev[0]
            lines.append(f"\n*📈 Тренд 2 нед:*")
            trend_parts = []
            if weight is not None and older.get("weight_kg") is not None:
                d = weight - older["weight_kg"]
                trend_parts.append(f"Вес {d:+.1f}")
            if fat is not None and older.get("body_fat_pct") is not None:
                d = fat - older["body_fat_pct"]
                trend_parts.append(f"жир {d:+.1f}%")
            if muscle is not None and older.get("muscle_mass_kg") is not None:
                d = muscle - older["muscle_mass_kg"]
                trend_parts.append(f"мышцы {d:+.1f}")
            if trend_parts:
                lines.append(" | ".join(trend_parts))

        # Красные флаги
        flags = []
        if fat is not None and prev and len(prev) >= 2:
            older = prev[0]
            if older.get("body_fat_pct"):
                d = abs(fat - older["body_fat_pct"])
                if d > BC_RED_FLAGS["fat_change_pct"]:
                    flags.append(f"⚠️ Жир изменился на {d:.1f}% — возможен артефакт. Перевзвеси!")
        if muscle is not None and prev and len(prev) >= 2:
            older = prev[0]
            if older.get("muscle_mass_kg"):
                d = muscle - older["muscle_mass_kg"]
                if d < -BC_RED_FLAGS["muscle_loss_kg"]:
                    flags.append(f"🚨 Потеря мышц: {d:+.1f} кг!")
        if water is not None and water < BC_RED_FLAGS["water_critical_pct"]:
            flags.append(f"⚠️ Критическое обезвоживание: {water:.0f}%!")
        if visceral is not None and visceral > BC_RED_FLAGS["visceral_risk"]:
            flags.append(f"🔴 Высокий висцеральный жир: {visceral:.0f}")

        if flags:
            lines.append(f"\n*🚨 Красные флаги:*")
            lines.extend(flags)

        lines.append(f"\n_⚠️ GARLYN: жир занижает ~3-4% (Potter+3%)_")
        return "\n".join(lines)

    def _build_coach_bc_report(self, latest, prev, athlete):
        """Построить сообщение для тренера — краткая сводка без медицинских деталей."""
        if not latest:
            return "⚖️ Нет данных для анализа."

        name = athlete.get("full_name", "?")
        weight = latest.get("weight_kg")
        muscle = latest.get("muscle_mass_kg")

        parts = []
        if weight is not None:
            parts.append(f"⚖️ *{weight:.1f}* кг")
        if muscle is not None:
            parts.append(f"мышцы *{muscle:.1f}* кг")

        trend_parts = []
        if prev and len(prev) >= 2:
            older = prev[0]
            if weight is not None and older.get("weight_kg") is not None:
                d = weight - older["weight_kg"]
                trend_parts.append(f"{d:+.1f} кг")
            if muscle is not None and older.get("muscle_mass_kg") is not None:
                d = muscle - older["muscle_mass_kg"]
                trend_parts.append(f"мышцы {d:+.1f}")

        # Определяем статус
        status = "✅ Норма"
        if muscle is not None and prev and len(prev) >= 2:
            older = prev[0]
            if older.get("muscle_mass_kg") and muscle - older["muscle_mass_kg"] < -1.5:
                status = "🚨 Внимание!"

        lines = [
            f"📋 *Замер — {name}*",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━",
            " | ".join(parts) if parts else "—",
        ]
        if trend_parts:
            lines.append(f"📈 Тренд: {' | '.join(trend_parts)}")
        lines.append(status)
        return "\n".join(lines)

    async def _analyze_and_notify_scale(self, update, athlete, records):
        """Анализ состава тела + уведомления врачу и тренеру после импорта CSV."""
        aid = athlete["id"]

        # Получаем текущий и предыдущий замер
        latest = self.db.get_latest_body_composition(aid)
        all_history = self.db.get_body_composition(aid, days=90)

        if not latest:
            return

        # --- Текст для спортсмена ---
        athlete_text = self._build_athlete_bc_report(latest, all_history, athlete)
        try:
            await update.message.reply_text(
                athlete_text, parse_mode="Markdown",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
            )
        except Exception as e:
            logger.error(f" Athlete BC report error: {e}")

        # --- Уведомление врачу (админу) ---
        admin_text = self._build_admin_bc_report(latest, all_history, athlete)
        try:
            for admin_id in self._full_access_ids():
                if admin_id != athlete.get("telegram_id"):
                    await self._send_admin(admin_id, admin_text)
        except Exception as e:
            logger.error(f"Admin BC notify error: {e}")

        # --- Уведомление тренеру ---
        team = athlete.get("team")
        if team:
            coach_text = self._build_coach_bc_report(latest, all_history, athlete)
            try:
                coaches = self.db.get_team_coaches(team)
                full_ids = self._full_access_ids()
                for cid in coaches:
                    if cid == athlete.get("telegram_id") or cid in full_ids:
                        continue
                    await self._send_admin(cid, coach_text)
            except Exception as e:
                logger.error(f"Coach BC notify error: {e}")

    async def _ask_height_for_records(self, update, ctx, records):
        """Если у записей нет роста в анкете — просим ввести один раз."""
        # берём первый athlete_id с отсутствующим ростом
        mapping = self.db.get_scale_profiles()
        state = self.get_state(update.effective_user.id)
        for r in records:
            prof = r.get("device_profile")
            aid = mapping.get(prof) if prof else None
            if aid is None:
                continue
            qd = self.db.get_questionnaire(aid)
            if qd and qd.get("height"):
                continue
            state["step"] = "bc_height"
            state["data"]["bc_pending"] = records
            state["data"]["bc_height_athlete"] = aid
            await update.message.reply_text(
                f"📏 У спортсмена (id {aid}) нет роста в анкете.\n"
                f"Введи рост в см числом (например: 180):"
            )
            return True
        return False

    async def _bc_height_save(self, update, ctx, user_id, text):
        """Сохранить введённый рост в анкету и завершить импорт."""
        try:
            h = float(text.strip().replace(",", "."))
            if not (80 <= h <= 240):
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Введи рост числом от 80 до 240 см (например: 180):")
            return
        state = self.get_state(user_id)
        athlete_id = state.get("data", {}).get("bc_height_athlete")
        # пишем рост в анкету (обновить существующую или создать направление)
        try:
            self.db.conn.execute(
                "UPDATE questionnaires SET height = ? WHERE athlete_id = ?", (int(h), athlete_id)
            )
            self.db.conn.commit()
        except Exception as e:
            logger.error(f"set height: {e}")
        records = state.get("data", {}).get("bc_pending") or []
        self.clear_state(user_id)
        if records:
            saved, skipped, unmapped, rows = await self._import_scale_records(
                update, ctx, records, forced_athlete_id=None)
            head = [f"✅ Рост сохранён: {int(h)} см", "━━━━━━━━━━━━━━━━━━━━━━━━━━"]
            if rows:
                head.append(f"⚖️ Сохранено: *{saved}* записей состава тела")
            if skipped:
                head.append(f"⚠️ Пропущено: {skipped}")
            await update.message.reply_text(
                "\n".join(head),
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]),
                parse_mode="Markdown"
            )
        else:
            await update.message.reply_text(
                f"✅ Рост сохранён: {int(h)} см.",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
            )

    async def show_profile(self, update, ctx):
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        stats = self.db.get_athlete_stats(athlete["id"], 7)
        prev = self.db.get_athlete_stats(athlete["id"], 14)
        streak = athlete.get("survey_streak", 0)

        # Рост/вес из анкеты + расчёт ИМТ (перцентили WHO для юных, взрослые категории 19+)
        bmi_line = ""
        qd = self.db.get_questionnaire(athlete["id"])
        if qd:
            h = qd.get("height")
            w = qd.get("weight")
            if h and w:
                try:
                    h = float(h); w = float(w)
                except (TypeError, ValueError):
                    h = w = 0
                if 0 < h < 400 and 0 < w < 300:
                    bmi = w / (h / 100) ** 2
                    age_group = athlete.get("age_group", "Pro")
                    years = AGE_GROUP_YEARS.get(age_group, 25)
                    gender_key = "m" if athlete.get("gender") == "male" else "f"
                    perc = BMI_PERCENTILES.get(years)
                    if perc and gender_key in perc:
                        # Юные (до 19): перцентили WHO. <5 — скрининг; >95 — без алерта (ИМТ слаб для баскетбола).
                        p5, p95 = perc[gender_key]
                        if bmi < p5:
                            cat, emoji = "ниже 5-й перцентили — стоит обсудить с врачом", "⚠️"
                        elif bmi > p95:
                            cat, emoji = "выше 95-й перцентили (для баскетбола не показатель)", "ℹ️"
                        else:
                            cat, emoji = "в пределах нормы (5-95 перцентиль)", "✅"
                        bmi_line = (f"📏 Рост: {int(h)} см | ⚖️ Вес: {int(w)} кг\n"
                                    f"🧮 ИМТ: {bmi:.1f} ({emoji} {cat})\n")
                    else:
                        # 19+/Pro: взрослые категории WHO
                        if bmi < 18.5:
                            cat, emoji = "недостаточный вес", "⚠️"
                        elif bmi < 25:
                            cat, emoji = "норма", "✅"
                        elif bmi < 30:
                            cat, emoji = "избыточный вес", "⚠️"
                        else:
                            cat, emoji = "ожирение", "🔴"
                        bmi_line = (f"📏 Рост: {int(h)} см | ⚖️ Вес: {int(w)} кг\n"
                                    f"🧮 ИМТ: {bmi:.1f} ({emoji} {cat})\n")

        # Готовность из 3 шкал (сон+утомление+боль)
        sleep_avg = stats.get("avg_sleep", 0) or 0
        fatigue_avg = stats.get("avg_fatigue", 0) or 0
        soreness_avg = stats.get("avg_soreness", 0) or 0
        hooper = sleep_avg + fatigue_avg + soreness_avg
        if hooper >= 17:
            ready = "🟢 Отличная"
        elif hooper >= 12:
            ready = "🟡 Средняя"
        else:
            ready = "🔴 Низкая"

        text = (
            f"👤 *Профиль*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"*{athlete['full_name']}*\n"
            f"📋 {athlete['age_group']} | 🏀 {athlete.get('team', '?')}\n"
            f"🔥 Серия: {streak}д | {get_rank(streak)}\n"
            f"📊 Опросов: {athlete.get('total_surveys', 0)}\n"
            f"{bmi_line}"
            f"{self._bc_trend_block(athlete['id'])}"
            f"📊 *Готовность:* {ready} ({hooper:.0f}/21)\n"
            f"*Средние за 7 дней:*\n"
            f"😴 Сон: {score_bar(round(sleep_avg))}\n"
            f"😩 Утомление: {score_bar(round(fatigue_avg))}\n"
            f"🤕 Боль: {score_bar(round(soreness_avg))}\n"
            f"❤️ Пульс: {(stats.get('avg_hr') or 0):.0f} уд/мин\n\n"
            f"*Тренды (7д vs 14д):*\n"
            f"😴 Сон: {trend_arrow(sleep_avg, prev.get('avg_sleep'))}\n"
            f"❤️ Пульс: {trend_arrow(stats.get('avg_hr'), prev.get('avg_hr'), True)}\n"
        )

        await q.edit_message_text(text, reply_markup=self.kb([
            [(f"📈 Графики", "my_charts")],
            ([] if athlete.get("gender") != "female" else [(f"📅 Мой цикл", "my_cycle")]),
            [(f"⚖️ Обновить вес", "update_weight")],
            [(f"⚖️ Состав тела", "body_comp_menu")],
            [(f"📋 Анкета / обновить", "questionnaire")],
            [(f"🔙 Назад", "main_menu")]
        ]), parse_mode="Markdown")

    def _bc_trend_block(self, athlete_id):
        """Короткий блок тренда массы из весов (для профиля спортсмена — только вес, без психологизации)."""
        try:
            trend = self.db.get_bc_trend(athlete_id, days=30)
            w = trend.get("weight_kg", [])
            if len(w) < 2:
                return ""
            first = w[0]; last = w[-1]
            if first[1] is None or last[1] is None:
                return ""
            delta = last[1] - first[1]
            arrow = "▴" if delta > 0 else ("▾" if delta < 0 else "•")
            return (f"⚖️ *Вес по весам (30 дн):* {first[1]:.1f} → {last[1]:.1f} кг "
                    f"{arrow} ({delta:+.1f})\n\n")
        except Exception as e:
            logger.error(f"bc trend block: {e}")
            return ""

    # ==================== ОБНОВЛЕНИЕ ВЕСА ====================

    async def update_weight_start(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "update_weight"
        await q.edit_message_text(
            "⚖️ *Обновление веса*\n\nВведи *текущий вес в кг* числом (например: 72):",
            parse_mode="Markdown"
        )

    async def update_weight_save(self, update, ctx, user_id, text):
        try:
            w = float(text.strip().replace(",", "."))
            if not (30 <= w <= 200):
                raise ValueError
        except ValueError:
            await update.message.reply_text("❌ Введи вес числом от 30 до 200 (например: 72):")
            return
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if not athlete:
            await update.message.reply_text("❌ Сначала зарегистрируйся: /start")
            return
        # Обновляем вес в анкете (если есть) — ИМТ пересчитается в профиле
        qd = self.db.get_questionnaire(athlete["id"])
        if qd:
            self.db.conn.execute(
                "UPDATE questionnaires SET weight = ? WHERE athlete_id = ?",
                (int(w), athlete["id"])
            )
            self.db.conn.commit()
        self.clear_state(user_id)
        await update.message.reply_text(
            f"✅ Вес обновлён: *{int(w)} кг*",
            reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]),
            parse_mode="Markdown"
        )

    # ==================== ПРОГРЕСП СПОРТСМЕНА ====================

    async def show_my_progress(self, update, ctx):
        """Персональный дашборд: sparkline за 7 дней + личная норма."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        aid = athlete["id"]
        last7 = self.db.get_last_wellness(aid, 7)
        bl = self.db.get_individual_baseline(aid, 30)

        if not last7:
            await q.edit_message_text(
                "📊 *Мой прогресс*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "📭 Пока нет данных. Пройди опрос, и здесь появятся графики!",
                reply_markup=self.kb([[(f"📝 Пройти опрос", "do_survey"), (f"🏠 Главное меню", "main_menu")]]),
                parse_mode="Markdown"
            )
            return

        # Собираем данные за 7 дней (от старых к новым)
        dates = [r["survey_date"][5:] for r in reversed(last7)]
        sleep_vals = [r.get("sleep_score") for r in reversed(last7)]
        fatigue_vals = [r.get("fatigue_score") for r in reversed(last7)]
        soreness_vals = [r.get("muscle_soreness") for r in reversed(last7)]
        hr_vals = [r.get("resting_hr") for r in reversed(last7)]
        readiness_vals = [r.get("readiness") for r in reversed(last7)]

        # Hooper за каждый день
        hooper_vals = []
        for r in reversed(last7):
            h = sum(filter(None, [r.get("sleep_score"), r.get("fatigue_score"), r.get("muscle_soreness")]))
            hooper_vals.append(h if h else None)

        text = f"📈 *Мой прогресс — 7 дней*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"👤 {athlete['full_name']} | {athlete.get('age_group', '?')}\n\n"

        # Sparkline-линии (спарклайн из emoji-кружков)
        def _spark_line(vals, invert=False):
            clean = [v for v in vals if v is not None]
            if not clean:
                return "—"
            result = ""
            for v in vals:
                if v is None:
                    result += "—"
                    continue
                if invert:
                    if v <= 60: result += "🟢"
                    elif v <= 70: result += "🟡"
                    else: result += "🔴"
                else:
                    if v >= 6: result += "🟢"
                    elif v >= 4: result += "🟡"
                    else: result += "🔴"
            return result

        def _spark_line_hr(vals):
            result = ""
            for v in vals:
                if v is None:
                    result += "—"
                elif v <= 60: result += "🟢"
                elif v <= 70: result += "🟡"
                else: result += "🔴"
            return result

        def _trend(vals):
            clean = [v for v in vals if v is not None]
            if len(clean) < 2:
                return ""
            diff = clean[-1] - clean[0]
            if abs(diff) < 0.5:
                return "→ стабильно"
            return f"{'↑' if diff > 0 else '↓'} {abs(diff):.1f}"

        # Сон
        last_sleep = sleep_vals[-1] if sleep_vals[-1] is not None else "?"
        bl_sleep = bl.get("median", {}).get("sleep") if bl else None
        text += f"😴 *Сон:* {_spark_line(sleep_vals)} {last_sleep}/7\n"
        if bl_sleep: text += f"   Ваша норма: ~{bl_sleep:.1f} | {_trend(sleep_vals)}\n"
        text += "\n"

        # Готовность
        last_ready = readiness_vals[-1] if readiness_vals[-1] is not None else "?"
        text += f"🎯 *Готовность:* {_spark_line(readiness_vals)} {last_ready}/10\n"
        text += "\n"

        # Утомление
        last_fatigue = fatigue_vals[-1] if fatigue_vals[-1] is not None else "?"
        bl_fatigue = bl.get("median", {}).get("fatigue") if bl else None
        text += f"⚡ *Утомление:* {_spark_line(fatigue_vals)} {last_fatigue}/7\n"
        if bl_fatigue: text += f"   Ваша норма: ~{bl_fatigue:.1f}\n"
        text += "\n"

        # Боль
        last_sore = soreness_vals[-1] if soreness_vals[-1] is not None else "?"
        text += f"🤕 *Боль:* {_spark_line(soreness_vals)} {last_sore}/7\n"
        text += "\n"

        # Пульс (если есть)
        hr_clean = [v for v in hr_vals if v is not None]
        if hr_clean:
            last_hr = hr_vals[-1]
            bl_hr = bl.get("median_hr") if bl else None
            text += f"❤️ *Пульс:* {_spark_line_hr(hr_vals)} {last_hr} уд/мин\n"
            if bl_hr: text += f"   Ваша норма: ~{int(bl_hr)} уд/мин\n"
            text += "\n"

        # Hooper
        last_hooper = hooper_vals[-1] if hooper_vals[-1] is not None else "?"
        text += f"🔥 *Hooper:* {last_hooper}/21\n"

        # Боли (NRS) — если были
        pain_vals = [r.get("pain_nrs") for r in reversed(last7)]
        pain_days = [v for v in pain_vals if v is not None and v > 0]
        if pain_days:
            text += f"\n⚠️ *Боль (NRS):* {len(pain_days)}/7 дней\n"
            for i, r in enumerate(reversed(last7)):
                if r.get("pain_nrs") and r["pain_nrs"] > 0:
                    text += f"   {r['survey_date'][5:]}: {r['pain_nrs']}/10"
                    if r.get("pain_location"):
                        text += f" ({r['pain_location']})"
                    text += "\n"

        # Фаза цикла (для девушек)
        phases = [r.get("cycle_phase") for r in reversed(last7) if r.get("cycle_phase")]
        if phases:
            last_phase = phases[-1]
            text += f"\n🔄 *Фаза:* {last_phase}\n"

        # Кнопки
        buttons = [
            [(f"📊 Подробнее", "my_stats"), (f"🏠 Главное меню", "main_menu")],
        ]

        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    # ==================== ПЕРСОНАЛЬНЫЕ ЦЕЛИ ====================

    async def show_my_goals(self, update, ctx):
        """Экран целей спортсмена."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        goals = self.db.get_active_goals(athlete["id"])
        text = f"🎯 *Мои цели*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"

        if goals:
            for g in goals:
                gt = g["goal_type"]
                tv = g.get("target_value")
                label = {
                    "sleep_min": f"Сон >= {tv}",
                    "readiness_min": f"Готовность >= {tv}",
                    "pain_free": "Без боли (NRS = 0)",
                }.get(gt, gt)
                text += f"• {label}\n"
            text += "\n"
        else:
            text += "📭 Активных целей нет.\n\n"

        buttons = [
            [(f"➕ Поставить цель", "goal_add_menu")],
            [(f"🏠 Главное меню", "main_menu")],
        ]
        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def goal_add_type(self, update, ctx):
        """Выбор типа цели."""
        q = update.callback_query
        await q.answer()
        buttons = [
            [(f"😴 Сон >= 6", "goal_set_sleep_6"), (f"😴 Сон >= 7", "goal_set_sleep_7")],
            [(f"🎯 Готовность >= 7", "goal_set_readiness_7"), (f"🎯 Готовность >= 8", "goal_set_readiness_8")],
            [(f"🤕 Без боли (NRS=0)", "goal_set_painfree")],
            [(f"🔙 Назад", "my_goals")],
        ]
        await q.edit_message_text(
            "🎯 *Поставить цель*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери цель:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def goal_set_value(self, update, ctx):
        """Сохранить цель."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        d = q.data
        if d.startswith("goal_set_sleep_"):
            tv = int(d.replace("goal_set_sleep_", ""))
            self.db.add_goal(athlete["id"], "sleep_min", target_value=tv)
            label = f"Сон >= {tv}"
        elif d.startswith("goal_set_readiness_"):
            tv = int(d.replace("goal_set_readiness_", ""))
            self.db.add_goal(athlete["id"], "readiness_min", target_value=tv)
            label = f"Готовность >= {tv}"
        elif d == "goal_set_painfree":
            self.db.add_goal(athlete["id"], "pain_free")
            label = "Без боли (NRS = 0)"
        else:
            return

        await q.edit_message_text(
            f"✅ *Цель установлена:* {label}\n\nБот будет отслеживать прогресс!",
            reply_markup=self.kb([[(f"🎯 Мои цели", "my_goals"), (f"🏠 Главное меню", "main_menu")]]),
            parse_mode="Markdown"
        )

    async def goal_complete(self, update, ctx):
        """Отметить цель как достигнутую или удалить."""
        q = update.callback_query
        await q.answer()
        goal_id = int(q.data.replace("goal_done_", ""))
        self.db.update_goal_progress(goal_id, achieved=True)
        await q.edit_message_text(
            "🎉 *Цель отмечена как достигнутая!*\n\nТак держать!",
            reply_markup=self.kb([[(f"🎯 Мои цели", "my_goals"), (f"🏠 Главное меню", "main_menu")]]),
            parse_mode="Markdown"
        )

    # ==================== ГРАФИКИ ====================

    async def show_charts(self, update, ctx):
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        state = self.get_state(q.from_user.id)
        days = state.get("data", {}).get("chart_days", 7)

        # Метрики: для шкалы 1-7 (7=хорошо) invert=False (цвета по новой шкале). Пульс — отдельно.
        metrics = {
            "sleep": ("😴 Сон", False),
            "fatigue": ("😩 Утомление", False),
            "soreness": ("🤕 Боль в мышцах", False),
            "hr": ("❤️ Пульс", True),
        }

        text = f"📈 *Графики за {days} дней*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"

        for key, (label, invert) in metrics.items():
            data = self.db.get_trend_data(athlete["id"], key, days)
            vals = [v for _, v in data if v is not None]

            if len(vals) >= 2:
                avg = sum(vals) / len(vals)
                mn, mx = min(vals), max(vals)
                spark = sparkline(vals)
                spark_c = sparkline_colored(vals, invert=invert)

                text += f"{label}: `{spark}`\n"
                text += f"{spark_c} ср:{avg:.1f} мин:{mn:.0f} макс:{mx:.0f}\n\n"
            elif len(vals) == 1:
                text += f"{label}: только сегодня ({get_score_emoji(vals[0])} {vals[0]})\n\n"
            else:
                text += f"{label}: нет данных\n\n"

        # Hooper Index за период (сумма sleep+fatigue+soreness, норма до 21)
        hooper_data = self.db.get_trend_data(athlete["id"], "sleep", days)
        fatigue_hooper = self.db.get_trend_data(athlete["id"], "fatigue", days)
        soreness_hooper = self.db.get_trend_data(athlete["id"], "soreness", days)

        if hooper_data and fatigue_hooper:
            # Строим Hooper по дням
            date_map = {}
            for d, v in hooper_data:
                if v is not None:
                    date_map.setdefault(str(d), {})["sleep"] = v
            for d, v in fatigue_hooper:
                if v is not None:
                    date_map.setdefault(str(d), {})["fatigue"] = v
            for d, v in soreness_hooper:
                if v is not None:
                    date_map.setdefault(str(d), {})["soreness"] = v

            hooper_vals = []
            for d in sorted(date_map.keys()):
                m = date_map[d]
                if all(k in m for k in ["sleep", "fatigue"]):
                    h = m["sleep"] + m["fatigue"] + m.get("soreness", 0)
                    hooper_vals.append(h)

            if len(hooper_vals) >= 2:
                avg_h = sum(hooper_vals) / len(hooper_vals)
                text += f"📊 *Hooper Index:* ср {avg_h:.0f}/21\n"
                # Визуализация (чем выше - тем лучше)
                bar = ""
                for h in hooper_vals[-10:]:
                    if h >= 17:
                        bar += "🟢"
                    elif h >= 12:
                        bar += "🟡"
                    else:
                        bar += "🔴"
                text += f"{bar}\n\n"

        text += "\n*Выбери период:*"

        period_buttons = [
            [("7 дней", "chart_7"), ("14 дней", "chart_14"), ("30 дней", "chart_30")],
            [(f"🔙 Назад", "main_menu")]
        ]
        await q.edit_message_text(text, reply_markup=self.kb(period_buttons), parse_mode="Markdown")

    # ==================== КОМАНДА ====================

    async def show_team(self, update, ctx):
        q = update.callback_query
        await q.answer()
        # Состав доступен только врачу/тренеру (защита ПДн — спортсмены не видят чужие данные)
        if not self._is_full_access(q.from_user.id):
            await q.edit_message_text(
                "🔒 Список команды доступен только врачу/тренеру.",
                reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]])
            )
            return
        athletes = self.db.get_all_athletes()
        if not athletes:
            await q.edit_message_text("📭 Нет спортсменов.", reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            return

        groups = {}
        for a in athletes:
            groups.setdefault(a["age_group"], []).append(a)

        text = f"👥 *Команда ЧБК*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        for g in sorted(groups.keys()):
            text += f"*{g}* ({len(groups[g])}):\n"
            for a in groups[g]:
                s = a.get("survey_streak", 0)
                icon = "🟢" if s >= 7 else ("🟡" if s >= 3 else ("🟠" if s > 0 else "🔴"))
                text += f"  {icon} {a['full_name']} ({a.get('team', '?')}) {s}д\n"
            text += "\n"

        text += "🟢 7+д | 🟡 3-6 | 🟠 1-2 | 🔴 0"
        await q.edit_message_text(text, reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown")

    # ==================== РОЛИ И ПРАВА ====================

    def _is_full_access(self, user_id) -> bool:
        """Полный доступ (как у админа): супер-админ из ADMIN_TELEGRAM_IDS или врач из таблицы doctors."""
        return user_id in ADMIN_TELEGRAM_IDS or self.db.is_doctor(user_id)

    def _full_access_ids(self):
        """Все telegram_id для УВЕДОМЛЕНИЙ врачу: админы + врачи, тренеры — исключены.

        Тренеру НЕ приходят уведомления о каждом опросе его команды (только
        еженедельная сводка по понедельникам). Исключение действует, даже если
        тренер записан врачом или админом."""
        docs = {d["telegram_id"] for d in self.db.get_all_doctors()}
        ids = set(ADMIN_TELEGRAM_IDS) | docs
        return {i for i in ids if not self.db.is_coach(i)}

    def _is_admin_or_coach(self, user_id) -> bool:
        """Полный доступ (админ/врач) видит всё; тренер — только свои команды."""
        return self._is_full_access(user_id) or self.db.is_coach(user_id)

    def _scoped_athletes(self, user_id, athletes=None):
        """Список спортсменов в рамках прав пользователя.

        Админ/врач → только активные команды (ACTIVE_TEAMS). Тренер → только спортсмены
        его команд (активных). Обычный пользователь → пусто. Спортсмены неактивных
        команд остаются в БД, но скрыты."""
        _active = set(ACTIVE_TEAMS)
        if self._is_full_access(user_id):
            src = athletes if athletes is not None else self.db.get_all_athletes()
            return [a for a in src if a.get("team") in _active]
        teams = set(self.db.get_coach_teams(user_id)) & _active
        if not teams:
            return []
        src = athletes if athletes is not None else self.db.get_all_athletes()
        return [a for a in src if a.get("team") in teams]

    def _coach_teams(self, user_id):
        return set(self.db.get_coach_teams(user_id))

    async def coach_team_view(self, update, ctx):
        """Панель тренера: спортсмены только его команд + кнопка «Рекомендации»."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        if not self.db.is_coach(user_id):
            await q.edit_message_text("🔒 Нет доступа.", reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            return
        athletes = self._scoped_athletes(user_id)
        teams = self._coach_teams(user_id)

        # Batch: один запрос вместо 2N
        survey_map = self.db.get_today_survey_map()

        # Статус готовности команды на сегодня
        ready = unsure = not_ready = no_data = 0
        for a in athletes:
            today = survey_map.get(a["id"])
            r = today.get("readiness") if today else None
            if r is None:
                no_data += 1
            elif r >= 7:
                ready += 1
            elif r >= 5:
                unsure += 1
            else:
                not_ready += 1

        status_line = (
            f"🎯 *Готовность сегодня:* 🟢 {ready} готовы | 🟡 {unsure} под вопросом | 🔴 {not_ready} не готовы"
            + (f" | ⚪ {no_data} без опроса" if no_data else "")
        )

        text = f"👕 *Мои команды ({len(teams)}):*\n"
        text += "  " + ", ".join(sorted(teams)) + "\n\n"
        text += status_line + "\n\n"
        text += f"👥 *Спортсмены ({len(athletes)}):*\n"
        banned = set(x["id"] for x in self.db.get_banned_athletes())
        for a in athletes:
            s = a.get("survey_streak", 0)
            today = survey_map.get(a["id"])
            r = today.get("readiness") if today else None
            if r is None:
                ricon = "⚪"
            elif r >= 7:
                ricon = "🟢"
            elif r >= 5:
                ricon = "🟡"
            else:
                ricon = "🔴"
            icon = "🟢" if s >= 7 else ("🟡" if s >= 3 else ("🟠" if s > 0 else "🔴"))
            lock = "🔒 " if a["id"] in banned else ""
            text += f"{lock}{ricon} {a['full_name']} | {a.get('age_group', '?')} | 🔥{s}д\n"

        btns = []
        for a in athletes:
            btns.append([(f"🏥 Рекомендации: {a['full_name']}", f"recs_{a['id']}")])
        btns.append([(f"📊 Сводка за неделю", "coach_week_summary")])
        btns.append([(f"📊 Мой отчёт (Excel)", "report_export_menu")])
        btns.append([(f"⚖️ Отчёт по весам", "report_body_comp")])
        btns.append([(f"🔙 Главное меню", "main_menu")])
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def coach_athlete_recs(self, update, ctx):
        """Рекомендации для конкретного спортсмена (только своей команды / админ)."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        athlete_id = int(q.data.replace("recs_", ""))
        if not self._is_admin_or_coach(user_id):
            await q.edit_message_text("🔒 Нет доступа.", reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            return
        # Проверяем, что спортсмен в рамках прав
        scoped = {a["id"] for a in self._scoped_athletes(user_id)}
        if athlete_id not in scoped:
            await q.edit_message_text("🔒 Нет доступа к этому спортсмену.", reply_markup=self.kb([[(f"🔙 Главное меню", "main_menu")]]))
            return
        a = next((x for x in self.db.get_all_athletes() if x["id"] == athlete_id), None)
        if not a:
            await q.edit_message_text("❌ Спортсмен не найден.", reply_markup=self.kb([[(f"🔙 Моя команда", "coach_team_view")]]))
            return
        hist = self.db.get_last_wellness(athlete_id, 1)
        head = (
            f"🏥 *Рекомендации врача*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👤 *{a['full_name']}* | {a.get('age_group', '?')} | 🏀 {a.get('team', '?')}\n"
            f"📅 Последний опрос: {hist[0].get('survey_date') if hist else 'нет'}\n\n"
        )
        if not hist:
            text = head + "📭 Опросов пока нет."
        else:
            recs = self._doctor_recs(hist[0], a.get("age_group", "Pro"), athlete_id=athlete_id)
            text = head + recs + "\n\n⚠️ *Это не медицинское заключение.* При сомнениях — обратись к врачу лично."
        btns = [[(f"🔙 Моя команда", "coach_team_view")], [(f"🏠 Главное меню", "main_menu")]]
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    def _coach_summary_text(self, teams):
        """Текст недельной сводки для тренера по его командам."""
        if not teams:
            return ""
        lines = ["📊 *Недельная сводка по командам*\n━━━━━━━━━━━━━━━━━━━━━━━━━━"]
        for team in sorted(teams):
            s = self.db.get_team_week_summary(team, 7)
            if not s["athletes"]:
                lines.append(f"\n👕 *{team}*\n— спортсменов нет")
                continue
            total = s["athletes"] * 7
            fill = f"{s['surveys']} из {total}"
            if total:
                fill += f" ({s['surveys'] * 100 // total}%)"
            avg_r = s["avg_readiness"]
            lines.append(
                f"\n👕 *{team}*\n"
                f"👥 Спортсменов: {s['athletes']} (отвечали {s['active']})\n"
                f"📝 Опросов за неделю: {fill}\n"
                f"⚡ Средняя готовность: {avg_r:.1f}/10" if avg_r else
                f"\n👕 *{team}*\n"
                f"👥 Спортсменов: {s['athletes']} (отвечали {s['active']})\n"
                f"📝 Опросов за неделю: {fill}\n"
                f"⚡ Средняя готовность: —"
            )
            flags = []
            if s["pain_hi"]:
                flags.append(f"🤕 Боль (NRS≥5): {s['pain_hi']}")
            if s["analgesics"]:
                flags.append(f"💊 Обезболивающие: {s['analgesics']}")
            if s["illness"]:
                flags.append(f"🤒 Болезнь: {s['illness']}")
            lines.append("🚨 *Требуют внимания:*\n  • " + "\n  • ".join(flags) if flags
                         else "✅ За неделю всё спокойно")
        return "\n".join(lines)

    async def coach_week_summary(self, update, ctx):
        """Сводка за неделю по командам тренера (по кнопке)."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        if not self.db.is_coach(user_id):
            await q.edit_message_text("🔒 Нет доступа.")
            return
        text = self._coach_summary_text(self._coach_teams(user_id)) or "Нет команд."
        btns = [[(f"🔙 Моя команда", "coach_team_view")], [(f"🏠 Главное меню", "main_menu")]]
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def _send_weekly_coach_summary(self, context):
        """Понедельничная сводка по командам — каждому тренеру."""
        if TESTING:
            logger.info("TESTING MODE — сводки тренерам не отправляются")
            return
        bot = context.bot
        for coach in self.db.get_all_coaches():
            text = self._coach_summary_text(coach["teams"])
            if not text:
                continue
            try:
                await bot.send_message(
                    chat_id=coach["telegram_id"], text=text, parse_mode="Markdown"
                )
            except Exception as e:
                logger.error(f"Coach summary error for {coach['telegram_id']}: {e}")

    # ==================== УПРАВЛЕНИЕ ТРЕНЕРАМИ (ТОЛЬКО АДМИН) ====================
    async def show_coach_admin(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        coaches = self.db.get_all_coaches()
        text = "🏅 *Тренеры и команды*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        if not coaches:
            text += "Тренеров пока нет.\n"
        for c in coaches:
            text += f"👤 {' / '.join(c['names'])} ({c['telegram_id']})\n"
            text += f"   🏀 {', '.join(c['teams']) or '— без команд'}\n\n"
        btns = [[(f"➕ Добавить тренера", "coach_add")]]
        for c in coaches:
            btns.append([(f"✏️ {c['telegram_id']} — изменить", f"coach_edit_{c['telegram_id']}"),
                         (f"🗑 Удалить", f"coach_del_{c['telegram_id']}")])
        btns.append([(f"🔙 Управление", "admin_manage")])
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def coach_add_ask(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        state = self.get_state(q.from_user.id)
        state["step"] = "admin_coach_add"
        await q.edit_message_text(
            "➕ *Добавить тренера*\n\nПришли *Telegram ID* пользователя, которого нужно сделать тренером.",
            reply_markup=self.kb([[(f"🔙 Тренеры", "coach_menu")]])
        )

    async def coach_edit(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        tg_id = int(q.data.replace("coach_edit_", ""))
        state = self.get_state(q.from_user.id)
        state["data"]["coach_edit_id"] = tg_id
        state["data"]["coach_edit_teams"] = self.db.get_coach_teams(tg_id)
        await self._render_coach_teams(q, tg_id, state["data"]["coach_edit_teams"])

    async def coach_toggle(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        _, team, tg_id = q.data.rsplit("_", 2)
        tg_id = int(tg_id)
        state = self.get_state(q.from_user.id)
        teams = list(state["data"].get("coach_edit_teams", []))
        if team in teams:
            teams.remove(team)
        else:
            teams.append(team)
        state["data"]["coach_edit_teams"] = teams
        await self._render_coach_teams(q, tg_id, teams)

    def _coach_picker_parts(self, tg_id, teams):
        teams = list(teams)
        btns = []
        for t in TEAMS:
            mark = "✅" if t in teams else "⬜"
            btns.append([(f"{mark} {t}", f"coach_toggle_{t}_{tg_id}")])
        btns.append([(f"💾 Сохранить", f"coach_save_{tg_id}")])
        btns.append([(f"🔙 Тренеры", "coach_menu")])
        lines = "\n".join(f"  {'✅' if t in teams else '⬜'} {t}" for t in TEAMS)
        text = (
            f"✏️ *Команды тренера (id {tg_id})*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Отметь/сними команды:\n{lines}\n\n"
            f"*Выбрано:* {', '.join(teams) or '—'}"
        )
        return text, btns

    async def _render_coach_teams(self, q, tg_id, teams):
        text, btns = self._coach_picker_parts(tg_id, teams)
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def _send_coach_team_picker(self, update, tg_id, teams):
        text, btns = self._coach_picker_parts(tg_id, teams)
        await update.message.reply_text(text, reply_markup=self.kb(btns))

    async def coach_save(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        tg_id = int(q.data.replace("coach_save_", ""))
        state = self.get_state(q.from_user.id)
        teams = state["data"].get("coach_edit_teams", [])
        self.db.set_coach_teams(tg_id, teams)
        await q.edit_message_text(
            f"✅ Тренер (id {tg_id}) обновлён. Команды: {', '.join(teams) or '—'}",
            reply_markup=self.kb([[(f"🔙 Тренеры", "coach_menu")]])
        )

    async def coach_delete(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        tg_id = int(q.data.replace("coach_del_", ""))
        self.db.remove_coach(tg_id)
        await q.edit_message_text(f"🗑 Тренер (id {tg_id}) удалён.",
                                  reply_markup=self.kb([[(f"🔙 Тренеры", "coach_menu")]]))
        await self.show_coach_admin(update, ctx)

    # ==================== УПРАВЛЕНИЕ ВРАЧАМИ (полный доступ, как у админа) ====================
    async def show_doctor_admin(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_full_access(uid):
            return
        docs = self.db.get_all_doctors()
        admins = sorted(ADMIN_TELEGRAM_IDS)
        text = "🩺 *Врачи и админы*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        text += "*(полный доступ, как у админа)*\n\n"
        text += "👑 *Супер-админы (из .env):*\n"
        for a in admins:
            text += f"  • id {a}\n"
        text += "\n🩺 *Врачи:*\n"
        if not docs:
            text += "  Врачей пока нет.\n"
        for d in docs:
            text += f"  • {' / '.join(d['names'])} ({d['telegram_id']})\n"
        btns = [[(f"➕ Добавить врача", "doctor_add")]]
        for d in docs:
            btns.append([(f"🗑 Удалить {d['telegram_id']}", f"doctor_del_{d['telegram_id']}")])
        btns.append([(f"🔙 Управление", "admin_manage")])
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def doctor_add_ask(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_full_access(uid):
            return
        state = self.get_state(uid)
        state["step"] = "admin_doctor_add"
        await q.edit_message_text(
            "🩺 *Добавить врача*\n\nПришли *Telegram ID* пользователя, которому нужен полный доступ (как у админа).\n\n"
            "⭐ Совет: супер-админ получает id в чате бота через 'мой ID' или у техподдержки.",
            reply_markup=self.kb([[(f"🔙 Врачи", "doctor_menu")]])
        )

    async def doctor_delete(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        tg_id = int(q.data.replace("doctor_del_", ""))
        self.db.remove_doctor(tg_id)
        await q.edit_message_text(f"🗑 Врач (id {tg_id}) удалён.",
                                  reply_markup=self.kb([[(f"🔙 Врачи", "doctor_menu")]]))
        await self.show_doctor_admin(update, ctx)

    # ==================== ДОСТИЖЕНИЯ ====================

    async def show_achievements(self, update, ctx):
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        streak = athlete.get("survey_streak", 0)
        total = athlete.get("total_surveys", 0)

        checks = [
            (1, "🌱 Первая запись"), (3, "🔥 Огненная серия"),
            (7, "⭐ Неделя без пропусков"), (14, "🏆 Две недели!"),
            (30, "💎 Месяц идеала!"),
        ]

        unlocked, locked = [], []
        for th, name in checks:
            (unlocked if streak >= th else locked).append(f"{name} ({th}д)" if streak < th else name)

        text = f"🏆 *Достижения*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        text += f"👤 {athlete['full_name']}\n🔥 {streak}д | {get_rank(streak)}\n📊 {total} опросов\n\n"

        if unlocked:
            text += f"*Открытые ({len(unlocked)}):*\n" + "".join(f"  ✅ {a}\n" for a in unlocked) + "\n"
        if locked:
            text += f"*Заблокированные:*\n" + "".join(f"  🔒 {a}\n" for a in locked[:5])

        await q.edit_message_text(text, reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown")

    # ==================== ОПРОСНИК ====================

    async def start_survey(self, update, ctx):
        q = update.callback_query
        await q.answer()
        # Тренер — админ команды, опрос не проходит
        if self.db.is_coach(q.from_user.id):
            await q.edit_message_text(
                "👕 Тренер не проходит опрос — панель тренера в главном меню.",
                reply_markup=self.kb([[(f"👕 Моя команда", "coach_team_view")],
                                      [(f"🏠 Главное меню", "main_menu")]])
            )
            return
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        if self.db.is_athlete_banned(athlete["id"]):
            await q.edit_message_text("🔒 Ты заблокирован. Обратись к администратору.", reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]))
            return

        if self.db.has_survey_today(athlete["id"]):
            await q.edit_message_text("✅ Уже прошёл опрос сегодня!", reply_markup=self.kb([[(f"🏠 Главная", "main_menu")]]))
            return

        state = self.get_state(q.from_user.id)
        state["step"] = "survey_sleep"
        state["data"] = {"athlete": athlete}

        await q.edit_message_text(
            f"📝 *Опрос* (1/8)\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"😴 *Качество сна*\n\nКак спал?\n\n"
            f"1 — 😫 Ужасно\n4 — 😐 Нормально\n7 — 😍 Отлично\n\nВыбери:",
            reply_markup=self.kb(self.score_buttons("sleep")), parse_mode="Markdown"
        )

    async def survey_callback(self, update, ctx):
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        state = self.get_state(user_id)
        step = state.get("step")
        data = state.get("data", {})
        athlete = data.get("athlete")

        # Если athlete нет в state — пробуем дозагрузить из БД
        if not athlete:
            athlete_db = self.db.get_athlete_by_telegram_id(user_id)
            if athlete_db:
                data["athlete"] = athlete_db
                athlete = athlete_db
            else:
                await q.edit_message_text(
                    "❌ *Ошибка*. Пожалуйста, начните заново.",
                    reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
                )
                return

        # Если step не survey — значит опрос устарел / рестарт
        if not step or not step.startswith("survey_"):
            await q.edit_message_text(
                "❌ *Опрос устарел*. Начните заново.",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
            )
            return

        parts = q.data.split("_")
        if len(parts) < 2:
            await q.edit_message_text(
                "❌ Ошибка данных. Попробуйте ещё раз.",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
            )
            return

        # Зона боли: painloc_<зона> — значение это название зоны (кириллица), не число
        if parts[0] == "painloc":
            data["pain_location"] = parts[1]
            state["step"] = "survey_pain_game"
            await q.edit_message_text(
                "🏀 *Болит ли на игре/тренировке?*",
                reply_markup=self.kb([[(f"✅ Да", "paingame_1"), (f"❌ Нет", "paingame_0")]]),
                parse_mode="Markdown"
            )
            return

        # srv_<поле>_<значение> — поле в parts[1]; двухчастные (paingame_/illness_/analg_) — поле в parts[0]
        field = parts[1] if parts[0] == "srv" else parts[0]

        try:
            value = int(parts[-1])
        except ValueError:
            await q.edit_message_text(
                "❌ Ошибка данных. Попробуйте ещё раз.",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
            )
            return

        field_map = {"sleep": "sleep_score", "fatigue": "fatigue_score",
                     "soreness": "muscle_soreness", "readiness": "readiness",
                     "painnrs": "pain_nrs"}
        if field in field_map:
            data[field_map[field]] = value

        # Новая последовательность опроса (урезанные шкалы + новые пункты):
        # сон → часы сна → readiness → утомление → боль → NRS → локация → на игре → болезнь → обезболивающие
        if field in ("sleep", "readiness", "fatigue", "soreness"):
            nxt = {
                "sleep": "survey_sleep_hours",
                "readiness": "survey_fatigue",
                "fatigue": "survey_soreness",
                "soreness": "survey_pain_nrs",
            }[field]
            state["step"] = nxt
            labels = {
                "sleep_hours": "😴 *Сколько часов реально спал?*\n\nНапиши числом (например: 7.5):",
                "fatigue": "😩 *Утомление*\n\nНасколько бодр?\n\n1 💀 — Упадок\n4 🦦 — Умеренно\n7 ⚡ — Бодрый\n\nВыбери:",
                "soreness": "🤕 *Мышечная боль*\n\nБолят мышцы?\n\n1 🤕 — Сильно\n4 😣 — Чувствуется\n7 ✨ — Не болят\n\nВыбери:",
            }
            if nxt == "survey_sleep_hours":
                await q.edit_message_text(
                    f"📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n{labels['sleep_hours']}",
                    parse_mode="Markdown"
                )
            elif nxt == "survey_pain_nrs":
                # NRS 0-10 (0 = боли нет) — свои кнопки
                await q.edit_message_text(
                    "📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    "🤕 *Оцени боль по шкале 0-10*\n\n0 — боли нет\n5 — умеренная\n10 — сильнейшая\n\nВыбери:",
                    reply_markup=self.kb([
                        [("0 — нет", "srv_painnrs_0"), ("1", "srv_painnrs_1"), ("2", "srv_painnrs_2")],
                        [("3", "srv_painnrs_3"), ("4", "srv_painnrs_4"), ("5", "srv_painnrs_5")],
                        [("6", "srv_painnrs_6"), ("7", "srv_painnrs_7"), ("8", "srv_painnrs_8")],
                        [("9", "srv_painnrs_9"), ("10", "srv_painnrs_10")],
                    ]), parse_mode="Markdown"
                )
            else:
                await q.edit_message_text(
                    f"📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n{labels[nxt.replace('survey_', '')]}\n\nВыбери:",
                    reply_markup=self.kb(self.score_buttons(nxt.replace("survey_", ""))),
                    parse_mode="Markdown"
                )
            return

        if field == "painnrs":
            # NRS 0-10: 0 = боли нет → пропускаем локацию
            if value == 0:
                data["pain_location"] = ""
                data["pain_on_game"] = 0
                state["step"] = "survey_illness"
                await self._ask_illness(update, ctx, state)
            else:
                state["step"] = "survey_pain_location"
                await q.edit_message_text(
                    f"📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n🤕 *Где болит?* (боль {value}/10)\n\nВыбери зону:",
                    reply_markup=self.kb([
                        [("🦶 Голеностоп", "painloc_голеностоп"), ("🦵 Колено", "painloc_колено")],
                        [("🍗 Бедро", "painloc_бедро"), ("🔙 Поясница", "painloc_поясница")],
                        [("💪 Плечо", "painloc_плечо"), ("🖐 Кисть", "painloc_кисть")],
                        [("🎗 Шея", "painloc_шея"), ("❓ Другое", "painloc_другое")],
                    ]), parse_mode="Markdown"
                )
            return

        if field == "paingame":
            data["pain_on_game"] = value
            state["step"] = "survey_illness"
            await self._ask_illness(update, ctx, state)
            return

        if field == "illness":
            data["illness_flag"] = {
                0: "", 1: "температура", 2: "насморк", 3: "голова", 4: "другое"
            }.get(value, "")
            state["step"] = "survey_analgesics"
            await q.edit_message_text(
                "💊 *Принимал ли обезболивающие?*\n\n(это важно — они могут скрывать травму)",
                reply_markup=self.kb([[(f"✅ Да", "analg_1"), (f"❌ Нет", "analg_0")]]),
                parse_mode="Markdown"
            )
            return

        if field == "analg":
            data["analgesics"] = value
            # Пульс запрашиваем ВСЕМ (в т.ч. младшим U14-U16, Simple-протокол)
            state["step"] = "survey_hr"
            await q.edit_message_text(
                f"📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"❤️ *Пульс покоя*\n\nИзмеряй утром, как только проснулся, ещё лежа в кровати и не вставая.\nПосчитай удары за 15 секунд и умножь на 4.\n\nВведи число:",
                parse_mode="Markdown"
            )
            return

        await q.edit_message_text("❌ Ошибка данных. Попробуйте ещё раз.",
                                  reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]))

    async def _ask_illness(self, update, ctx, state):
        """Флаг болезни (температура/насморк/голова) — шаг опроса."""
        q = update.callback_query
        await q.edit_message_text(
            "🤒 *Признаки болезни?*\n\nТемпература, насморк, головная боль?",
            reply_markup=self.kb([
                [("✅ Всё нормально", "illness_0")],
                [("🌡 Температура", "illness_1")],
                [("🤧 Насморк", "illness_2")],
                [("🤕 Голова болит", "illness_3")],
            ]), parse_mode="Markdown"
        )

    async def handle_text(self, update, ctx):
        user_id = update.effective_user.id
        if not self.check_rate_limit(user_id):
            await update.message.reply_text("⚠️ Подожди минуту.")
            return

        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if athlete and self.db.is_athlete_banned(athlete["id"]):
            await update.message.reply_text("🔒 Ты заблокирован. Обратись к администратору.")
            self.clear_state(user_id)
            return

        state = self.get_state(user_id)
        step = state.get("step")
        text = update.message.text.strip()

        # Регистрация: имя
        if step == "reg_name":
            name = re.sub(r'[^\w\s\-]', '', text.strip())
            if len(name.split()) < 2 or len(name) < 3:
                await update.message.reply_text("❌ Введи *Имя и Фамилию*:", parse_mode="Markdown")
                return

            name = " ".join(name.split()).title()
            _tg_fn = update.effective_user.first_name
            _reg_fn = _tg_fn if self._looks_like_name(_tg_fn) else name.split()[0]
            ok = self.db.register_athlete(
                telegram_id=user_id, username=update.effective_user.username,
                full_name=name, age_group=state["data"]["age_group"],
                team=state["data"].get("team", "Не указана"),
                first_name=_reg_fn
            )
            if not ok:
                # Если пользователь уже существует — пробуем войти
                existing = self.db.get_athlete_by_telegram_id(user_id)
                if existing:
                    await update.message.reply_text(
                        f"✅ *Добро пожаловать снова!*\n\n👤 {existing['full_name']}\n📋 {existing['age_group']}\n🏀 {existing.get('team', '?')}\n\nВы уже зарегистрированы 🏀",
                        reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]), parse_mode="Markdown"
                    )
                else:
                    await update.message.reply_text("❌ Ошибка регистрации. Попробуйте позже.", parse_mode="Markdown")
                return
            # Сохраняем пол, если указан
            gender = state["data"].get("gender")
            if gender:
                athlete = self.db.get_athlete_by_telegram_id(user_id)
                if athlete:
                    self.db.update_athlete_gender(athlete["id"], gender)
            self.clear_state(user_id)
            gender_text = "♂ Мужской" if gender == "male" else "♀ Женский" if gender == "female" else ""
            gender_line = f"\n👤 Пол: {gender_text}" if gender_text else ""
            # Если анкета ещё не заполнена — предложить
            athlete_obj = self.db.get_athlete_by_telegram_id(user_id)
            has_q = self.db.has_questionnaire(athlete_obj["id"]) if athlete_obj else False
            if not has_q:
                await update.message.reply_text(
                    f"✅ *Регистрация завершена!*\n\n👤 {name}\n📋 {state['data']['age_group']}\n🏀 {state['data'].get('team', '?')}{gender_line}\n\n📋 *Теперь заполни анкету — это важно для врача!*",
                    reply_markup=self.kb([[(f"📋 Пройти анкету", "questionnaire")]]), parse_mode="Markdown"
                )
            else:
                await update.message.reply_text(
                    f"✅ *Регистрация завершена!*\n\n👤 {name}\n📋 {state['data']['age_group']}\n🏀 {state['data'].get('team', '?')}{gender_line}\n\nДобро пожаловать! 🏀",
                    reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]), parse_mode="Markdown"
                )
            return

        # Админ: ввод Telegram ID нового тренера
        if step == "admin_coach_add":
            if user_id not in ADMIN_TELEGRAM_IDS:
                self.clear_state(user_id)
                return
            raw = text.strip().lstrip("@")
            if not raw.isdigit():
                await update.message.reply_text(
                    "❌ Это не похоже на числовой Telegram ID. Пришли число (или @username не поддерживается — нужен ID).",
                    reply_markup=self.kb([[(f"🔙 Тренеры", "coach_menu")]])
                )
                return
            tg_id = int(raw)
            state["data"]["coach_edit_id"] = tg_id
            state["data"]["coach_edit_teams"] = self.db.get_coach_teams(tg_id)
            state["step"] = None  # дальше работаем кнопками (coach_toggle_*)
            await self._send_coach_team_picker(update, tg_id, state["data"]["coach_edit_teams"])
            return

        # Спортсмен: ввод Telegram ID (добавление админом)
        if step == "admin_add_athlete_tg_id":
            if user_id not in ADMIN_TELEGRAM_IDS:
                self.clear_state(user_id)
                return
            raw = text.strip().lstrip("@")
            if not raw.isdigit():
                await update.message.reply_text(
                    "❌ Это не похоже на числовой Telegram ID. Пришли число.",
                    reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]])
                )
                return
            state["data"]["add_tg_id"] = raw
            # Переход к выбору команды
            buttons = [[(team, f"admin_add_team_{team}")] for team in TEAMS]
            buttons.append([("🔙 Управление", "admin_manage")])
            await update.message.reply_text(
                f"✅ Telegram ID: *{raw}*\n\nВыбери *команду:*",
                reply_markup=self.kb(buttons), parse_mode="Markdown"
            )
            return

        # Спортсмен: ввод ФИО (добавление админом)
        if step == "admin_add_athlete_name":
            if user_id not in ADMIN_TELEGRAM_IDS:
                self.clear_state(user_id)
                return
            name = re.sub(r'[^\w\s\-]', '', text.strip())
            if len(name.split()) < 2 or len(name) < 3:
                await update.message.reply_text(
                    "❌ Введи *Фамилию Имя* ( minimum 2 слова):",
                    parse_mode="Markdown"
                )
                return
            name = " ".join(name.split()).title()
            state["data"]["add_name"] = name
            await self.admin_add_athlete_save(update, user_id, state)
            return

        # Врач: ввод Telegram ID нового врача
        if step == "admin_doctor_add":
            if user_id not in ADMIN_TELEGRAM_IDS:
                self.clear_state(user_id)
                return
            raw = text.strip().lstrip("@")
            if not raw.isdigit():
                await update.message.reply_text(
                    "❌ Это не похоже на числовой Telegram ID. Пришли число.",
                    reply_markup=self.kb([[(f"🔙 Врачи", "doctor_menu")]])
                )
                return
            tg_id = int(raw)
            self.db.add_doctor(tg_id)
            state["step"] = None
            await update.message.reply_text(
                f"✅ Врач id *{tg_id}* добавлен — теперь у него полный доступ, как у админа.",
                reply_markup=self.kb([[(f"🔙 Врачи", "doctor_menu")]]),
                parse_mode="Markdown"
            )
            return

        # Врач/тренер: текст ответа спортсмену по жалобе
        if step == "admin_reply":
            await self._admin_reply_send(update, ctx, user_id, state, text)
            return

        # Обновление веса (своё, из профиля)
        if step == "update_weight":
            await self.update_weight_save(update, ctx, user_id, text)
            return

        # Ввод роста при импорте весов (нет роста в анкете)
        if step == "bc_height":
            await self._bc_height_save(update, ctx, user_id, text)
            return

        # Пульс
        if step == "survey_hr":
            try:
                hr = int(text)
                if not (30 <= hr <= 220):
                    raise ValueError
            except ValueError:
                await update.message.reply_text("❌ Введи число от 30 до 220:")
                return
            state["data"]["resting_hr"] = hr
            athlete = state.get("data", {}).get("athlete")
            is_simple = athlete and athlete.get("age_group") in SIMPLE_PROTOCOLS
            if is_simple:
                # Младшие (U14-U16): после пульса опрос завершается (без тренировки/sRPE)
                await self._route_after_survey(update, ctx, user_id, state)
                return
            state["step"] = "survey_training"
            await update.message.reply_text(
                "💪 *Была ли тренировка вчера?*",
                reply_markup=self.kb([[(f"✅ Да", "train_yes"), (f"❌ Нет", "train_no")]]),
                parse_mode="Markdown"
            )
            return

        # Фактические часы сна (число, напр. 7.5)
        if step == "survey_sleep_hours":
            try:
                hours = float(text.strip().replace(",", "."))
                if not (0 < hours <= 24):
                    raise ValueError
            except ValueError:
                await update.message.reply_text("❌ Введи часы числом (например: 7.5):")
                return
            state["data"]["sleep_hours"] = hours
            state["step"] = "survey_readiness"
            await update.message.reply_text(
                f"📝 *Опрос*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"💪 *Готовность к тренировке?*\n\n"
                f"1 — совсем не готов\n5 — средне\n10 — полностью готов\n\nВыбери:",
                reply_markup=self.kb(self.score_buttons("readiness", mn=1, mx=10)),
                parse_mode="Markdown"
            )
            return

        # SRPE
        if step == "survey_srpe":
            try:
                srpe = int(text)
                if not (1 <= srpe <= 10):
                    raise ValueError
            except ValueError:
                await update.message.reply_text("❌ Введи число от 1 до 10:")
                return
            state["data"]["sRPE_score"] = srpe
            await self._route_after_survey(update, ctx, user_id, state)
            return

        # Длина цикла (ручной ввод числа)
        if step == "survey_cycle_length":
            if not text.strip().isdigit():
                await update.message.reply_text("❌ Введи число дней цикла (21-35):")
                return
            length = int(text.strip())
            if not (21 <= length <= 35):
                await update.message.reply_text("❌ Норма цикла 21-35 дней. Введи число от 21 до 35:")
                return
            saved = await self._apply_cycle_length(user_id, length)
            if not saved:
                await update.message.reply_text(
                    "❌ *Опрос прерван*. Начните заново.",
                    reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
                )
                return
            await self.ask_cycle_day(update, ctx)
            return

        # День цикла (ввод числа)
        if step == "survey_cycle":
            athlete = state.get("data", {}).get("athlete")
            if not athlete:
                await update.message.reply_text(
                    "❌ *Опрос прерван* (возможно бот перезагружался).\n\nХотите начать заново?",
                    reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")],
                                          [(f"🏠 Главное меню", "main_menu")]]),
                    parse_mode="Markdown"
                )
                return
            try:
                cd = int(text)
                cl = athlete.get("cycle_length_default", 28)
                if not (1 <= cd <= cl):
                    raise ValueError
            except ValueError:
                await update.message.reply_text(f"❌ Введи число от 1 до {athlete.get('cycle_length_default', 28)}:")
                return

            from cycle_medicine import get_cycle_phase
            phase_key, phase_info = get_cycle_phase(cd, athlete.get("cycle_length_default", 28))
            state["data"]["cycle_day"] = cd
            state["data"]["cycle_length"] = athlete.get("cycle_length_default", 28)
            state["data"]["cycle_phase"] = phase_info.name_ru if phase_info else ""
            await self._ask_complaints(update, ctx)
            return

        # Жалобы (ввод текста)
        if step == "complaint_text_input":
            state["data"]["complaints"] = sanitize_text(text, keep_nl=True)
            await self._finish_survey(update, user_id, state)
            return

        # Анкета: текстовые ответы
        if step == "q_age":
            try:
                age = int(text)
                if age < 10 or age > 99: raise ValueError
                state["q_data"]["age"] = str(age)
                state["step"] = "q_phone"
                await update.message.reply_text("📋 *Блок 1: Общие данные*\n\nВведи номер телефона:", parse_mode="Markdown")
            except:
                await update.message.reply_text("❌ Введи число от 10 до 99")
            return

        if step == "q_phone":
            phone = text.strip()
            if len(phone) < 5:
                await update.message.reply_text("❌ Введи номер телефона (например: +7 900 123-45-67):")
                return
            state["q_data"]["phone"] = phone
            state["step"] = "q_birth_date"
            await update.message.reply_text("📋 *Блок 1: Общие данные*\n\nВведи дату рождения в формате *ДД.ММ.ГГГГ* (например: 15.03.2008):", parse_mode="Markdown")
            return

        if step == "q_birth_date":
            import re as _re
            if not _re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text.strip()):
                await update.message.reply_text("❌ Введи дату в формате *ДД.ММ.ГГГГ* (например: 15.03.2008):", parse_mode="Markdown")
                return
            state["q_data"]["birth_date"] = text.strip()
            state["step"] = "q_gender"
            await update.message.reply_text("📋 *Блок 1: Общие данные*\n\nВаш пол?", reply_markup=self.kb([[("М", "q_gender_М"), ("Ж", "q_gender_Ж")]]), parse_mode="Markdown")
            return

        if step == "q_height_weight":
            if text == "-":
                state["q_data"]["height"] = ""
                state["q_data"]["weight"] = ""
            else:
                parts = text.split()
                try:
                    h = float(parts[0]); w = float(parts[1])
                    if not (0 < h < 400 and 0 < w < 300):
                        raise ValueError
                except (ValueError, IndexError):
                    await update.message.reply_text("❌ Введи рост и вес числами через пробел (например: 185 82) или «-»")
                    return
                state["q_data"]["height"] = parts[0]
                state["q_data"]["weight"] = parts[1]
            state["step"] = "q_trauma_12m"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nБыли травмы за последние 12 месяцев?", reply_markup=self.kb([[("Да", "q_trauma_Да"), ("Нет", "q_trauma_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_trauma_detail":
            state["q_data"]["trauma_12m_detail"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_zones"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nЗоны, которые беспокоили за 3 месяца:", reply_markup=self.kb([
                [("Голеностопы", "q_zones_Голеностопы"), ("Колени", "q_zones_Колени")],
                [("Бёдра", "q_zones_Бёдра"), ("Поясница", "q_zones_Поясница")],
                [("Кисти", "q_zones_Кисти"), ("Плечи", "q_zones_Плечи")],
                [("Шея", "q_zones_Шея"), ("Ничего", "q_zones_Ничего")],
            ]), parse_mode="Markdown")
            return

        if step == "q_pain_detail":
            state["q_data"]["pain_now_detail"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_chronic"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nЕсть рецидивирующая (повторяющаяся) травма?", reply_markup=self.kb([[("Да", "q_chronic_Да"), ("Нет", "q_chronic_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_chronic_detail":
            state["q_data"]["chronic_detail"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_surgery"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nБыли операции, из-за спорта?", reply_markup=self.kb([[("Да", "q_surgery_Да"), ("Нет", "q_surgery_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_surgery_detail":
            state["q_data"]["surgery_detail"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_surgery_date"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nКогда была(и) операция(и) (год или дата)?\nНапример: 2023 или 12.05.2023", parse_mode="Markdown")
            return

        if step == "q_surgery_date":
            state["q_data"]["surgery_date"] = text.strip()
            await self._q_continue_meds(update, ctx, state)
            return

        if step == "q_meds_detail":
            state["q_data"]["meds_detail"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_allergies"
            await update.message.reply_text("📋 *Блок 2: Травмы*\n\nЕсть аллергии?", reply_markup=self.kb([[("Да", "q_allergies_Да"), ("Нет", "q_allergies_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_allergies_detail":
            state["q_data"]["allergies_detail"] = sanitize_text(text, keep_nl=True)
            await self._q_save_then_block3(update, ctx)
            return

        if step == "q_supp_other_input":
            # пользователь ввёл свой вариант спортпита — добавить к выбранным
            cur = state["q_data"].get("supplements", [])
            if isinstance(cur, str):
                cur = [cur] if cur and cur != "Ничего" else []
            cur = list(cur) if isinstance(cur, list) else []
            if text.strip():
                cur.append(text.strip())
            state["q_data"]["supplements"] = cur
            # автoсохранение
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_supplements"
            await self._q_show_supp_buttons(update, ctx, cur)
            return

        if step == "q_goal":
            state["q_data"]["goal"] = sanitize_text(text, keep_nl=True)
            state["step"] = "q_wish"
            await update.message.reply_text("📋 *Блок 6: Цели*\n\nЕсть что-то, с чем нужна помощь?\n(можно пропустить — напиши «-»)", parse_mode="Markdown")
            return

        if step == "q_wish":
            state["q_data"]["wish"] = (sanitize_text(text, keep_nl=True)) if text != "-" else ""
            # Сохраняем
            athlete = self.db.get_athlete_by_telegram_id(user_id)
            if athlete:
                state["q_data"]["athlete_id"] = athlete["id"]
                await self._db_run(self.db.save_questionnaire, athlete["id"], state["q_data"])
                self.db.complete_questionnaire(athlete["id"])
            self.clear_state(user_id)
            await self._finish_questionnaire_offer(update, ctx)
            return

        # Консультация — ввод жалоб текстом
        if step == "consult_text_input":
            state["data"]["consult_complaints"] = sanitize_text(text, keep_nl=True)
            await self._show_calendar(update, ctx, "consult")
            return

        # Своё время напоминания
        if step == "set_reminder_custom":
            try:
                parts = text.split(":")
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
                if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                    raise ValueError
                global REMINDER_HOUR, REMINDER_MINUTE, REMINDER_TZ
                REMINDER_HOUR = hour
                REMINDER_MINUTE = minute
                REMINDER_TZ = "Asia/Yekaterinburg"
                self.db.set_setting("reminder_hour", str(hour))
                self.db.set_setting("reminder_tz", REMINDER_TZ)
                self._schedule_reminder_job(hour, minute)
                self.clear_state(user_id)
                await update.message.reply_text(
                    f"✅ Время напоминаний: *{hour:02d}:{minute:02d}* по Челябинску (сохранено)",
                    parse_mode="Markdown",
                    reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]])
                )
            except:
                await update.message.reply_text("❌ Неверный формат. Напиши время как ЧЧ:ММ (например: 14:30)")
                return
            return

        if step:
            await update.message.reply_text("Используй кнопки.", reply_markup=self.kb([[(f"🏠 Меню", "main_menu")]]))
        else:
            # Проверяем, зарегистрирован ли пользователь — может быть потеря состояния
            athlete = self.db.get_athlete_by_telegram_id(user_id)
            if athlete:
                await update.message.reply_text(
                    "❌ *Опрос прерван* (возможно бот перезагружался).\n\nХотите начать заново?",
                    reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")],
                                          [(f"🏠 Главное меню", "main_menu")]]),
                    parse_mode="Markdown"
                )
            else:
                await update.message.reply_text("Напиши /start для входа.")

        # Сохраняем сессию в БД (переживает рестарт) — асинхронно
        try:
            await self.persist_user_state(user_id)
        except Exception:
            pass

    async def callback_handler(self, update, ctx):
        q = update.callback_query
        try:
            await q.answer()
        except Exception:
            pass
        d = q.data
        logger.info(f"Callback: {d} from {q.from_user.id}")

        # Само-синхронизация имени из Telegram (только если похоже на имя, не ник)
        if q.from_user.first_name and self._looks_like_name(q.from_user.first_name):
            try:
                _ath = self.db.get_athlete_by_telegram_id(q.from_user.id)
                if _ath and _ath.get("first_name") != q.from_user.first_name:
                    self.db.set_athlete_first_name(_ath["id"], q.from_user.first_name)
            except Exception:
                pass

        # Глобальный try/except — любая ошибка не должна вешать бота
        try:
            if d == "main_menu":
                await self.show_main_menu(update, ctx)
            elif d == "consent_accept":
                # регистрируем согласие и переходим к выбору возрастной группы
                self.db.record_consent(q.from_user.id)
                await self._ask_age_group(update, ctx)
            elif d == "consent_decline":
                await q.edit_message_text("❌ Без согласия на обработку данных использовать бот нельзя.\nЕсли передумаешь — напиши /start.", reply_markup=self.kb([[(f"📄 Дать согласие", "consent_accept")]]), parse_mode="Markdown")
            elif d.startswith("reg_gender_"):
                await self.reg_gender_callback(update, ctx)
            elif d.startswith("reg_"):
                await self.reg_callback(update, ctx)
            elif d.startswith("team_"):
                await self.team_callback(update, ctx)
            elif d == "my_progress":
                await self.show_my_progress(update, ctx)
            elif d == "my_goals":
                await self.show_my_goals(update, ctx)
            elif d.startswith("goal_add_"):
                await self.goal_add_type(update, ctx)
            elif d.startswith("goal_set_"):
                await self.goal_set_value(update, ctx)
            elif d.startswith("goal_done_"):
                await self.goal_complete(update, ctx)
            elif d == "my_stats":
                await self.show_profile(update, ctx)
            elif d == "my_profile":
                await self.show_profile(update, ctx)
            elif d in ("my_charts",):
                await self.show_charts(update, ctx)
            elif d.startswith("chart_"):
                days = int(d.replace("chart_", ""))
                state = self.get_state(q.from_user.id)
                state["data"]["chart_days"] = days
                await self.show_charts(update, ctx)
            elif d == "team_list":
                await self.show_team(update, ctx)
            elif d == "achievements":
                await self.show_achievements(update, ctx)
            elif d == "cancel_survey":
                self.clear_state(q.from_user.id)
                await self.show_main_menu(update, ctx)
            elif d == "do_survey":
                await self.start_survey(update, ctx)
            elif d.startswith("srv_"):
                await self.survey_callback(update, ctx)
            elif d.startswith(("painloc_", "paingame_", "illness_", "analg_")):
                await self.survey_callback(update, ctx)
            elif d == "train_yes":
                await self._training_yes(update, ctx)
            elif d == "train_no":
                await self._training_no(update, ctx)
            elif d == "help_menu":
                await self.show_help(update, ctx)
            elif d == "questionnaire_list":
                await self.show_questionnaire_list(update, ctx)
            elif d == "questionnaire":
                await self.start_questionnaire(update, ctx)
            elif d == "q_restart":
                await self.questionnaire_restart(update, ctx)
            elif d.startswith("q_"):
                await self.handle_questionnaire_answer(update, ctx)
            elif d == "start_reg":
                await self.cmd_start(update, ctx)
            elif d == "gender_mandatory_male":
                await self._gender_mandatory_chosen(update, ctx, "male")
            elif d == "gender_mandatory_female":
                await self._gender_mandatory_chosen(update, ctx, "female")
            elif d == "view_today":
                await self.show_profile(update, ctx)
            elif d == "admin_report":
                await self.show_admin_report(update, ctx)
            elif d == "watch_data":
                await self.watch_data_menu(update, ctx)
            elif d.startswith("watch_"):
                brand = d.replace("watch_", "")
                await self._watch_brand_instructions(update, ctx, brand)
            elif d == "bc_menu":
                await self.scale_import_menu(update, ctx)
            elif d == "bc_overview":
                await self.bc_overview(update, ctx)
            elif d == "bc_profiles":
                await self.bc_show_profiles(update, ctx)
            elif d.startswith("bc_page_"):
                await self.bc_page(update, ctx)
            elif d.startswith("bc_map_"):
                await self.bc_assign_pick(update, ctx)
            elif d.startswith("bc_assign_"):
                await self.bc_assign_save(update, ctx)
            elif d.startswith("bc_view_"):
                await self.bc_view_athlete(update, ctx)
            elif d == "body_comp_menu":
                await self.body_comp_menu(update, ctx)
            elif d == "body_comp_history":
                await self.body_comp_history(update, ctx)
            elif d == "reset_account":
                await self.reset_account(update, ctx)
            elif d == "confirm_reset":
                await self.confirm_reset(update, ctx)
            elif d == "export_csv":
                await self.export_csv(update, ctx)
            elif d == "athlete_list":
                await self.athlete_list(update, ctx)
            elif d.startswith("qview_"):
                await self.show_questionnaire_detail(update, ctx)
            elif d.startswith("athlete_page_"):
                await self.athlete_list(update, ctx, int(d.split("_")[-1]))
            elif d == "ban_menu":
                await self.show_ban_menu(update, ctx)
            elif d.startswith("ban_page_"):
                await self.show_ban_menu(update, ctx, int(d.split("_")[-1]))
            elif d.startswith("ban_"):
                a_id = int(d.split("_")[1])
                self.db.ban_athlete(a_id, q.from_user.id)
                await q.answer("✅ Заблокирован", show_alert=True)
                await self.show_ban_menu(update, ctx)
            elif d.startswith("unban_"):
                a_id = int(d.split("_")[1])
                self.db.unban_athlete(a_id)
                await q.answer("✅ Разблокирован", show_alert=True)
                await self.show_ban_menu(update, ctx)
            elif d == "admin_manage":
                await self.show_admin_manage(update, ctx)
            elif d == "admin_add_athlete":
                await self.admin_add_athlete_start(update, ctx)
            elif d.startswith("admin_add_team_"):
                await self.admin_add_athlete_team(update, ctx)
            elif d.startswith("admin_add_age_"):
                await self.admin_add_athlete_age(update, ctx)
            elif d == "daily_report":
                await self.daily_report(update, ctx)
            elif d == "delete_athlete":
                await self.delete_athlete_menu(update, ctx)
            elif d.startswith("del_") and not d.startswith("delconfirm_"):
                await self.delete_athlete_confirm(update, ctx)
            elif d.startswith("delconfirm_"):
                await self.delete_athlete_final(update, ctx)
            elif d == "reminder_settings":
                await self.reminder_settings(update, ctx)
            elif d.startswith("set_reminder_"):
                await self.set_reminder_time(update, ctx)
            elif d == "send_reminder_now":
                await self.send_reminder_now(update, ctx)
            elif d == "send_q_reminder":
                await self.send_questionnaire_reminder(update, ctx)
            # === НОВЫЕ ОБРАБОТЧИКИ ===
            elif d == "report_export_menu":
                await self.report_export_menu(update, ctx)
            elif d == "export_q":
                await self.export_questionnaires_xlsx(update, ctx)
            elif d.startswith("report_team_"):
                await self.report_choose_period(update, ctx)
            elif d.startswith("report_period_"):
                await self.generate_xlsx_report(update, ctx)
            elif d == "report_watches":
                await self.report_watches_choose_period(update, ctx)
            elif d.startswith("report_watch_team_"):
                await self.report_watches_choose_period(update, ctx)
            elif d.startswith("report_watch_period_"):
                await self.generate_watch_xlsx_report(update, ctx)
            elif d == "report_body_comp":
                await self.report_body_comp_choose_period(update, ctx)
            elif d.startswith("bc_team_"):
                # Тренер выбрал команду для отчёта по весам → показать выбор периода
                team = d.replace("bc_team_", "")
                state = self.get_state(uid)
                state["data"]["bc_report_team"] = team
                buttons = [
                    [("📅 Последний замер", "report_bc_period_1")],
                    [("📅 2 недели", "report_bc_period_14")],
                    [("📅 Месяц", "report_bc_period_30")],
                    [("📅 3 месяца", "report_bc_period_90")],
                    [("📅 Всё время", "report_bc_period_999")],
                    [("🔙 Назад", "report_body_comp")]
                ]
                await q.edit_message_text(
                    f"⚖️ *Отчёт по весам — {team}*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    "Замеры проводятся раз в 2 недели. Выбери период:",
                    reply_markup=self.kb(buttons), parse_mode="Markdown"
                )
            elif d.startswith("report_bc_period_"):
                await self.generate_bc_xlsx_report(update, ctx)
            elif d == "set_gender":
                await self.set_gender_menu(update, ctx)
            # === ТРЕНЕР (роль coach) ===
            elif d == "coach_team_view":
                await self.coach_team_view(update, ctx)
            elif d.startswith("recs_"):
                await self.coach_athlete_recs(update, ctx)
            elif d == "coach_week_summary":
                await self.coach_week_summary(update, ctx)
            elif d == "coach_menu":
                await self.show_coach_admin(update, ctx)
            elif d == "coach_add":
                await self.coach_add_ask(update, ctx)
            elif d.startswith("coach_edit_"):
                await self.coach_edit(update, ctx)
            elif d.startswith("coach_toggle_"):
                await self.coach_toggle(update, ctx)
            elif d.startswith("coach_save_"):
                await self.coach_save(update, ctx)
            elif d.startswith("coach_del_"):
                await self.coach_delete(update, ctx)
            elif d == "doctor_menu":
                await self.show_doctor_admin(update, ctx)
            elif d == "doctor_add":
                await self.doctor_add_ask(update, ctx)
            elif d.startswith("doctor_del_"):
                await self.doctor_delete(update, ctx)
            elif d.startswith("reply_athlete_"):
                await self._admin_reply_start(update, ctx)
            elif d == "admin_complaints":
                await self.admin_complaints_menu(update, ctx)
            elif d.startswith("admin_complaints_page_"):
                await self.admin_complaints_page(update, ctx)
            elif d == "pdf_report_menu":
                await self.pdf_report_menu(update, ctx)
            elif d.startswith("pdf_"):
                await self.pdf_report_generate(update, ctx)
            # === ПОЛ (ПЕРВИЧНЫЙ ВОПРОС В ОПРОСЕ) ===
            # Должен обрабатываться ДО общего префикса gender_ (иначе save_gender поглотит кнопку)
            elif d == "gender_first_male":
                await self._gender_first_chosen(update, ctx, "male")
            elif d == "gender_first_female":
                await self._gender_first_chosen(update, ctx, "female")
            elif d.startswith("gender_"):
                await self.save_gender(update, ctx)
            elif d == "cycle_len_custom":
                await self._ask_cycle_length_custom(update, ctx)
            elif d == "cycle_custom":
                await self._ask_cycle_day_custom(update, ctx)
            elif d.startswith("cycle_len_"):
                await self._cycle_length_chosen(update, ctx)
            elif d.startswith("cycle_"):
                await self.set_cycle_day(update, ctx)
            elif d == "my_cycle":
                await self.show_cycle_info(update, ctx)
            elif d == "update_weight":
                await self.update_weight_start(update, ctx)
            # === ЖАЛОБЫ И КОНСУЛЬТАЦИЯ ===
            elif d == "complaint_text":
                await self.complaint_text(update, ctx)
            elif d == "complaint_none":
                await self.complaint_none(update, ctx)
            elif d == "consultation_start":
                await self.consultation_start(update, ctx)
            elif d == "consult_text":
                await self.consultation_text(update, ctx)
            elif d == "consult_no":
                await self.consultation_no_complaints(update, ctx)
            elif d.startswith("consult_") and d not in ["consultation_start", "consult_no", "consult_text"]:
                await self.consultation_date(update, ctx)

        # Конец try/except для callback_handler
        except Exception as e:
            if "Message is not modified" in str(e):
                # Двойной клик по кнопке: сообщение уже в нужном виде — молча игнорируем
                try:
                    await q.answer()
                except Exception:
                    pass
            else:
                logger.error(f"Callback handler error: {e} | data={d}", exc_info=True)
                try:
                    await q.edit_message_text(
                        "❌ Произошла ошибка. Попробуйте ещё раз.",
                        reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
                    )
                except Exception:
                    pass

        # Сохраняем сессию в БД (переживает рестарт) — асинхронно, без блокировки
        try:
            await self.persist_user_state(user_id)
        except Exception:
            pass

    async def _training_yes(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        if state.get("step") != "survey_training":
            await q.edit_message_text(
                "❌ *Опрос прерван*. Начните заново.",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
            )
            return
        state["data"]["had_training"] = 1
        state["step"] = "survey_srpe"
        await q.edit_message_text(
            "⚡ *Оцените нагрузку (1-10):*\n\n1 — легко\n5 — умеренно\n10 — максимум\n\nВведи число:",
            parse_mode="Markdown"
        )

    async def _training_no(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        if state.get("step") != "survey_training":
            await q.edit_message_text(
                "❌ *Опрос прерван*. Начните заново.",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
            )
            return
        state["data"]["had_training"] = 0
        state["data"]["sRPE_score"] = None
        await self._route_after_survey(update, ctx, q.from_user.id, state)

    async def _route_after_survey(self, update, ctx, user_id, state):
        """Маршрутизация после опроса: пол → цикл → жалобы."""
        athlete = state.get("data", {}).get("athlete")
        if not athlete:
            await self._finish_survey(update, user_id, state)
            return

        gender = athlete.get("gender")
        age_group = athlete.get("age_group", "")
        cycle_length = None

        # Подгружаем свежие данные из БД (на случай если пол/цикл только что сохранили)
        fresh = self.db.get_athlete_by_telegram_id(user_id)
        if fresh:
            cycle_length = fresh.get("cycle_length_default")
        if not cycle_length:
            cycle_length = athlete.get("cycle_length_default")

        # Если пол не указан → спросить пол
        if gender is None:
            state["step"] = "ask_gender_first"
            q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
            text = "👤 *Укажите ваш пол*\n\nЭто нужно для дополнительных вопросов в опросе."
            buttons = self.kb([
                [("♂ Мужской", "gender_first_male")],
                [("♀ Женский", "gender_first_female")],
            ])
            if q:
                await q.edit_message_text(text, reply_markup=buttons, parse_mode="Markdown")
            else:
                await update.message.reply_text(text, reply_markup=buttons, parse_mode="Markdown")
            return

        # Если девушка → спросить длину цикла (если не сохранена) или день цикла
        if gender == "female":
            if not cycle_length:
                await self._ask_cycle_length(update, ctx)
            else:
                await self.ask_cycle_day(update, ctx)
        else:
            await self._ask_complaints(update, ctx)

    async def _gender_first_chosen(self, update, ctx, gender):
        """Сохранить пол, выбранный во время опроса, и продолжить."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        state = self.get_state(user_id)
        athlete = state.get("data", {}).get("athlete")

        if not athlete:
            await q.edit_message_text(
                "❌ *Опрос прерван*. Начните заново.",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
            )
            return

        # Сохраняем пол в БД
        if athlete:
            self.db.update_athlete_gender(athlete["id"], gender)
            athlete["gender"] = gender

        # Если female → спросить длину цикла, иначе → жалобы
        if gender == "female":
            await self._ask_cycle_length(update, ctx)
        else:
            await self._ask_complaints(update, ctx)

    async def _finish_survey(self, source, user_id, state):
        data = state["data"]
        athlete = data.get("athlete")
        if not athlete:
            return

        recs = self._doctor_recs(data, athlete["age_group"], athlete_id=athlete["id"])

        self.db.save_survey(athlete_id=athlete["id"], data={
            "sleep": data.get("sleep_score"), "stress": data.get("stress_score"),
            "fatigue": data.get("fatigue_score"), "soreness": data.get("muscle_soreness"),
            "mood": data.get("mood_score"), "hr": data.get("resting_hr"),
            "hrv": data.get("hrv_ms"), "training": data.get("had_training", 0),
            "srpe": data.get("sRPE_score"),
            "cycle_day": data.get("cycle_day"), "cycle_length": data.get("cycle_length"),
            "cycle_phase": data.get("cycle_phase"),
            "complaints": data.get("complaints"),
            "protocol": "simple" if athlete["age_group"] in SIMPLE_PROTOCOLS else "full",
            "auto_recommendation": recs if recs != "✅ Все показатели в норме!" else "",
            "sleep_hours": data.get("sleep_hours"),
            "readiness": data.get("readiness"),
            "pain_nrs": data.get("pain_nrs"),
            "pain_location": data.get("pain_location"),
            "pain_on_game": data.get("pain_on_game", 0),
            "illness_flag": data.get("illness_flag"),
            "analgesics": data.get("analgesics", 0),
        })

        # Персонализация цикла: если отмечен день 1 (начало месячных) — пересчитываем
        # фактическую длину цикла по истории отметок и обновляем cycle_length_default.
        cycle_update_note = ""
        if data.get("cycle_day") == 1 and athlete.get("gender") == "female":
            try:
                new_len = self.db.update_cycle_length_from_history(athlete["id"])
                if new_len and new_len != athlete.get("cycle_length_default"):
                    cycle_update_note = f"\n\n📅 Цикл персонализирован: ~*{new_len} дн.* (по твоим отметкам)"
                    athlete["cycle_length_default"] = new_len
            except Exception as e:
                logger.warning(f"cycle length update: {e}")

        # Готовность из 3 шкал (сон+утомление+боль; стресс/настроение убраны из опроса)
        hooper = sum(filter(None, [data.get("sleep_score"), data.get("fatigue_score"),
                                    data.get("muscle_soreness")]))
        hooper_max = 21

        athlete = self.db.get_athlete_by_telegram_id(user_id)
        streak = athlete.get("survey_streak", 0) if athlete else 0

        motivation = ""
        if streak >= 30: motivation = MOTIVATIONAL_MESSAGES["streak_30"]
        elif streak >= 14: motivation = MOTIVATIONAL_MESSAGES["streak_14"]
        elif streak >= 7: motivation = MOTIVATIONAL_MESSAGES["streak_7"]
        elif streak >= 3: motivation = MOTIVATIONAL_MESSAGES["streak_3"]

        text = (
            f"✅ *Опрос завершён!*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        )
        if hooper >= 17:
            text += f"🟢 *Уровень готовности: Отличный* ({hooper}/21)\n\n"
        elif hooper >= 12:
            text += f"🟡 *Уровень готовности: Средний* ({hooper}/21)\n\n"
        else:
            text += f"🔴 *Уровень готовности: Низкий* ({hooper}/21)\n\n"
        text += (
            f"😴 Сон: {get_score_emoji(data.get('sleep_score'))} {score_bar(data.get('sleep_score', 0))}\n"
            f"😩 Утомление: {get_score_emoji(data.get('fatigue_score'))} {score_bar(data.get('fatigue_score', 0))}\n"
        )
        if data.get("resting_hr"):
            text += f"❤️ Пульс: {data['resting_hr']} уд/мин\n"
        text += f"\n🔥 Серия: {streak} дней | {get_rank(streak)}\n"
        if motivation:
            text += f"\n{motivation}\n"
        if cycle_update_note:
            text += cycle_update_note + "\n"
        if recs:
            text += f"\n*🏥 Рекомендации врача:*\n{recs}\n"
            text += f"\n⚠️ *Это не медицинское заключение.* При сомнениях обратись к врачу лично."

        buttons = [[(f"🏠 Главное меню", "main_menu")]]

        try:
            if hasattr(source, "callback_query") and source.callback_query:
                await source.callback_query.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
            elif hasattr(source, "message") and source.message:
                await source.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
            elif hasattr(source, "edit_message_text"):
                await source.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Finish survey error: {e}")

        # Отправляем рекомендации врачу (админу)
        try:
            for admin_id in self._full_access_ids():
                if admin_id != user_id:  # Не дублируем тому кто прошел
                    athlete_name = athlete.get("full_name", "?")
                    team = athlete.get("team", "?")
                    admin_text = (
                        f"📋 *Новый опрос — {athlete_name}*\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"🏀 Команда: {team} | {athlete.get('age_group', '?')}\n\n"
                        f"😴 Сон: {get_score_emoji(data.get('sleep_score'))} {score_bar(data.get('sleep_score', 0))}\n"
                        f"😩 Утомление: {get_score_emoji(data.get('fatigue_score'))} {score_bar(data.get('fatigue_score', 0))}\n"
                    )
                    if data.get("resting_hr"):
                        admin_text += f"❤️ Пульс: {data['resting_hr']} уд/мин\n"
                    complaints = data.get("complaints", "")
                    if complaints:
                        admin_text += f"💬 Жалобы: {complaints}\n"
                    admin_text += f"\n🔥 Серия: {streak}д | Готовность: {hooper}/21"
                    if recs and recs != "✅ Все показатели в норме!":
                        admin_text += f"\n⚠️ *Нужно внимание:*\n{recs}\n"
                    else:
                        admin_text += f"\n✅ Все в норме\n"

                    # Кнопка-ссылка на диалог со спортсменом — открывает личный чат Telegram
                    need_reply = bool(complaints) or (recs and recs != "✅ Все показатели в норме!")
                    admin_buttons = ([[InlineKeyboardButton(f"💬 Написать {athlete_name}", url=f"tg://user?id={athlete['telegram_id']}")]]
                                     if need_reply and athlete.get("telegram_id") else None)

                    await self._send_admin(admin_id, admin_text, buttons=admin_buttons)
        except Exception as e:
            logger.error(f"Admin notify error: {e}")

        # Уведомление тренеру команды спортсмена (в дополнение к врачу)
        try:
            _team = athlete.get("team")
            if _team:
                _coaches = self.db.get_team_coaches(_team)
                _full_ids = self._full_access_ids()
                for _cid in _coaches:
                    if _cid == user_id or _cid in _full_ids:
                        continue
                    _ctext = (
                        f"📋 *Опрос пройден — {athlete.get('full_name', '?')}*\n"
                        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                        f"🏀 {_team} | {athlete.get('age_group', '?')}\n"
                    )
                    if hooper >= 17:
                        _ctext += f"🟢 Готовность: отличная ({hooper}/21)\n"
                    elif hooper >= 12:
                        _ctext += f"🟡 Готовность: средняя ({hooper}/21)\n"
                    else:
                        _ctext += f"🔴 Готовность: низкая ({hooper}/21)\n"
                    _nrs = data.get("pain_nrs")
                    if _nrs is not None and _nrs > 0:
                        _ctext += f"🤕 Боль: {_nrs}/10" + (f" ({data.get('pain_location', '')})" if data.get("pain_location") else "") + "\n"
                    if data.get("resting_hr"):
                        _ctext += f"❤️ Пульс: {data['resting_hr']} уд/мин\n"
                    if data.get("complaints"):
                        _ctext += f"💬 Жалобы: {data['complaints']}\n"
                    await self._send_admin(_cid, _ctext, buttons=None)
        except Exception as e:
            logger.error(f"Coach notify error: {e}")

        # 🚨 Алерт врачу при красных флагах опроса (боль NRS≥5, обезболивающие, болезнь)
        try:
            _flags = []
            _nrs = data.get("pain_nrs")
            if _nrs is not None and _nrs >= 5:
                _flags.append(f"🤕 Боль {_nrs}/10" + (f" ({data.get('pain_location', '')})" if data.get("pain_location") else ""))
            if data.get("analgesics"):
                _flags.append("💊 Обезболивающие (маскируют травму)")
            if data.get("illness_flag"):
                _flags.append(f"🤒 Болезнь: {data['illness_flag']}")
            if data.get("pain_on_game"):
                _flags.append("🏀 Болит на игре/тренировке")
            if _flags:
                _alert = (
                    f"🚨 *ТРЕБУЕТ ВНИМАНИЯ — {athlete.get('full_name', '?')}*\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"🏀 {athlete.get('team', '?')} | {athlete.get('age_group', '?')}\n\n"
                    + "\n".join(f"• {f}" for f in _flags)
                )
                _alert_btn = [[InlineKeyboardButton(f"💬 Написать {athlete.get('full_name', '?')}", url=f"tg://user?id={athlete['telegram_id']}")]] if athlete.get("telegram_id") else None
                for _aid in self._full_access_ids():
                    if _aid != user_id:
                        await self._send_admin(_aid, _alert, buttons=_alert_btn)
        except Exception as e:
            logger.error(f"Red-flag alert error: {e}")

        # Проверка прогресса целей
        try:
            goal_msgs = self.db.check_goals_progress(athlete["id"], data)
            if goal_msgs:
                for gm in goal_msgs:
                    try:
                        if hasattr(source, "callback_query") and source.callback_query:
                            await source.callback_query.message.reply_text(gm, parse_mode="Markdown")
                        elif hasattr(source, "message") and source.message:
                            await source.message.reply_text(gm, parse_mode="Markdown")
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"Goal check: {e}")

        self.clear_state(user_id)

    async def _send_admin(self, admin_id, text, buttons=None):
        """Отправить сообщение админу."""
        if TESTING:
            return
        try:
            bot = self._bot
            reply_markup = self.kb(buttons) if buttons else None
            await bot.send_message(
                chat_id=admin_id, text=text, parse_mode="Markdown",
                reply_markup=reply_markup
            )
        except Exception as e:
            logger.error(f"Send to admin {admin_id}: {e}")

    async def _admin_reply_start(self, update, ctx):
        """Врач/тренер отвечает спортсмену по жалобе — ждём текст сообщения."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        try:
            athlete_id = int(q.data.replace("reply_athlete_", ""))
        except ValueError:
            return
        athlete = self.db.get_athlete_by_id(athlete_id)
        if not athlete:
            await q.edit_message_text("❌ Спортсмен не найден.")
            return
        state = self.get_state(q.from_user.id)
        state["step"] = "admin_reply"
        state["data"]["reply_athlete_id"] = athlete_id
        await q.edit_message_text(
            f"📩 *Написать {athlete['full_name']}*\n\n"
            f"Напиши текст — спортсмен получит его в Telegram:\n"
            f"(для отмены отправь /cancel)",
            parse_mode="Markdown"
        )

    async def _admin_reply_send(self, update, ctx, user_id, state, text):
        """Отправка текста врача спортсмену."""
        athlete_id = state.get("data", {}).get("reply_athlete_id")
        self.clear_state(user_id)
        if not athlete_id:
            return
        athlete = self.db.get_athlete_by_id(athlete_id)
        if not athlete or not athlete.get("telegram_id"):
            await update.message.reply_text("❌ Спортсмен не найден.")
            return
        msg = sanitize_text(text, keep_nl=True)
        try:
            await ctx.bot.send_message(
                chat_id=athlete["telegram_id"],
                text=f"📩 *Сообщение от врача:*\n\n{msg}",
                parse_mode="Markdown"
            )
            await update.message.reply_text(
                f"✅ Отправлено: *{athlete['full_name']}*",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]),
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.error(f"Admin reply send error: {e}")
            await update.message.reply_text(
                "❌ Не удалось отправить (спортсмен мог заблокировать бота).",
                reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]])
            )

    def _doctor_recs(self, data, age_group, athlete_id=None):
        recs = []
        # Личная норма (30 дн.) для подсказок в тексте рекомендаций
        bl = None
        if athlete_id:
            try:
                bl = self.db.get_individual_baseline(athlete_id, 30)
            except Exception:
                pass

        def _personal_note(metric):
            if not bl:
                return ""
            med = (bl.get("median") or {}).get(metric)
            return f" (ваша норма: ~{med:.1f})" if med is not None else ""

        def _esc(s):
            # Пользовательский текст не должен ломать Markdown (непарные * _ [ `)
            return str(s).replace("\\", "\\\\").replace("*", "\\*").replace("_", "\\_").replace("[", "\\[").replace("`", "\\`")

        sleep = data.get("sleep_score")
        if sleep is not None:
            if sleep <= 2:
                recs.append("😴 *Критический дефицит сна*\n• Норма: 7-9ч, оценка 5-7\n• Ложиться до 22:30\n• Исключить кофеин после 16:00\n• При сохранении >3 дней → врач")
            elif sleep <= 4:
                recs.append(f"😴 *Недостаток сна{_personal_note('sleep')}*\n• Норма: 7-9ч, оценка 5-7\n• Ложиться до 23:00\n• Убрать гаджеты за 1ч до сна")
        stress = data.get("stress_score")
        if stress is not None:
            if stress <= 2:
                recs.append("🧘 *Высокий стресс*\n• Дыхание 4-7-8: 5 циклов\n• Прогулка 30мин\n• Снизить нагрузку на 20%")
            elif stress <= 3:
                recs.append(f"🧘 *Умеренный стресс{_personal_note('stress')}*\n• Дыхательные практики 5мин\n• Ограничить соцсети за 2ч до сна")
        fatigue = data.get("fatigue_score")
        if fatigue is not None:
            if fatigue <= 2:
                recs.append("⚡ *Критическое утомление*\n• Отдых 1-2 дня\n• Сон не менее 9ч\n• Белок 1.6-2.2г/кг")
            elif fatigue <= 4:
                recs.append(f"⚡ *Повышенное утомление{_personal_note('fatigue')}*\n• Легкая тренировка (50%)\n• Сон +1ч\n• Вода 2-3л")
        hr = data.get("resting_hr")
        hr_norms = HR_NORMS.get(age_group, {"min": 40, "max": 70})
        if hr:
            personal_hr = ""
            if bl and bl.get("median_hr") is not None:
                personal_hr = f"\n• Ваша личная норма: ~{int(bl['median_hr'])} уд/мин"
            if hr > hr_norms["max"] + 15:
                recs.append(f"❤️ *Пульс критически высокий ({hr})*\n• Норма {age_group}: {hr_norms['min']}-{hr_norms['max']}{personal_hr}\n• Исключить интенсивные тренировки\n• При пульсе >80 >3 дней → ЭКГ")
            elif hr > hr_norms["max"] + 5:
                recs.append(f"❤️ *Пульс выше нормы ({hr})*\n• Норма: {hr_norms['min']}-{hr_norms['max']}{personal_hr}\n• Снизить нагрузку на 30%")
        # HRV-рекомендации (только для Full-протокола, где HRV собирается)
        hrv = data.get("hrv_ms") or data.get("hrv")
        hrv_norms = {"U16": {"min": 40, "max": 60, "crit": 30}, "U17": {"min": 35, "max": 55, "crit": 25},
                     "U18": {"min": 35, "max": 55, "crit": 25}, "U19": {"min": 35, "max": 50, "crit": 25},
                     "U21": {"min": 35, "max": 50, "crit": 25}, "Pro": {"min": 30, "max": 50, "crit": 20}}
        if hrv and age_group in hrv_norms:
            hn = hrv_norms[age_group]
            if hrv < hn["crit"]:
                recs.append(f"📉 *HRV критически низкий ({hrv}мс)*\n• Норма {age_group}: {hn['min']}-{hn['max']}мс\n• Признак перетренированности или болезни\n• Отдых 1-2 дня, лёгкая тренировка")
            elif hrv < hn["min"]:
                recs.append(f"📉 *HRV снижен ({hrv}мс)*\n• Норма: {hn['min']}-{hn['max']}мс\n• Снизить интенсивность на 30%\n• Качественный сон 8ч+")
        soreness = data.get("muscle_soreness")
        if soreness is not None and soreness <= 2:
            recs.append("🤕 *Выраженная мышечная боль*\n• Активное восстановление (плавание)\n• Массаж через 48ч\n• При боли >3 дней → врач")
        mood = data.get("mood_score")
        if mood is not None and mood <= 2:
            recs.append("😊 *Сниженное настроение*\n• Прогулка 30мин/день\n• Общение с близкими\n• При сохранении >2 недель → психолог")
        cycle_phase = data.get("cycle_phase")
        if cycle_phase:
            phase_lower = cycle_phase.lower()
            if "менструа" in phase_lower or "фоллику" in phase_lower:
                recs.append("🔄 *Фаза: Менструальная/Фолликулярная (1-13 день)*\n• Нагрузка: легкая-средняя\n• Добавки: Железо, Витамин C, Магний\n• Восстановление: растяжка, йога\n• Питание: больше белка, зелени")
            elif "овуля" in phase_lower:
                recs.append("🔄 *Фаза: Овуляторная (14-16 день)*\n• Нагрузка: пиковая — можно максимум\n• Добавки: Омега-3, Цинк\n• Восстановление: качественная заминка\n• Питание: сложные углеводы")
            elif "люте" in phase_lower or "желт" in phase_lower:
                recs.append("🔄 *Фаза: Лютеиновая (17-28 день)*\n• Нагрузка: сниженная\n• Добавки: Магний, Витамин B6, Омега-3\n• Восстановление: сон +1ч, массаж\n• Питание: магний, калий, уменьшить соль")
            else:
                recs.append(f"🔄 *Фаза цикла:* {cycle_phase}\n• Учитывай фазу при планировании нагрузок")

        # ============ ОБЪЕКТИВНЫЕ ФЛАГИ (Фаза 4, по научной критике) ============
        # RHR-флаг: пульс ≥ медиана(7д) + max(8, 10%) — флаг инфекции/перегрузки (2 дня подряд — в _individual_recs)
        if athlete_id and hr:
            try:
                if self.db.rhr_flag(athlete_id, hr, days=7):
                    recs.append(f"💓 *Пульс заметно выше личной нормы ({hr})*\n• Возможный признак начала болезни или перегрузки\n• Проверь температуру, при подтверждении — отдых")
            except Exception as e:
                logger.warning(f"rhr flag: {e}")

        # EWMA ACWR (uncoupled): только при ≥28 дней sRPE. Пороги не жёсткие (калибровка на своих данных).
        if athlete_id:
            try:
                acwr = self.db.ewma_acwr(athlete_id)
                if acwr:
                    ratio = acwr["acwr"]
                    if ratio >= 1.5:
                        recs.append(f"📊 *Нагрузка резко выросла (ACWR {ratio})*\n• Острая нагрузка {acwr['acute']} vs хроническая {acwr['chronic']}\n• Обсуди с тренером снижение на 2-3 дня")
                    elif ratio >= 1.3:
                        recs.append(f"📊 *Нагрузка на границе (ACWR {ratio})*\n• Острая/хроническая: {acwr['acute']}/{acwr['chronic']}\n• Контролируй восстановление (сон, HRV)")
                    elif ratio < 0.8:
                        recs.append(f"📊 *Нагрузка низкая (ACWR {ratio})*\n• Если возврат после паузы — наращивай постепенно")
            except Exception as e:
                logger.warning(f"acwr: {e}")

        # ============ АНКЕТА → РЕКОМЕНДАЦИИ (правила «если-то») ============
        # Данные анкеты (датированы) влияют на фокус рекомендаций. Устаревшая анкета (>120 дней) помечается.
        if athlete_id:
            try:
                qd = self.db.get_questionnaire(athlete_id)
                if qd:
                    q_stale = False
                    try:
                        from datetime import datetime as _dt
                        if qd.get("completed_at"):
                            q_dt = _dt.fromisoformat(str(qd["completed_at"]).replace("Z", "+00:00"))
                            if (datetime.now() - q_dt.replace(tzinfo=None)).days > 120:
                                q_stale = True
                    except Exception:
                        pass

                    # 1. Зоны травм за 3 мес / боль в зоне: если сегодня боль (≤3/7) — фокус
                    zones = str(qd.get("zones") or "")
                    pain_now = str(qd.get("pain_now_detail") or qd.get("pain_now") or "")
                    if zones and zones != "Ничего":
                        sore = data.get("muscle_soreness")
                        pain_nrs = data.get("pain_nrs")
                        if (sore is not None and sore <= 3) or (pain_nrs is not None and pain_nrs >= 4):
                            recs.append(f"🩹 *Проблемная зона из анкеты:* {_esc(zones)}\n• Боль в этой зоне — снизь нагрузку, покажи врачу")
                    if pain_now and pain_now not in ("Нет", "нет", "-"):
                        recs.append(f"🩹 *В анкете отмечена боль:* {_esc(pain_now[:120])}\n• Отслеживай динамику, при усилении — осмотр")

                    # 2. Операция в анамнезе → боли в оперированной конечности вес ×2
                    surgery = str(qd.get("surgery_detail") or "")
                    if surgery and surgery not in ("Нет", "нет", "-"):
                        pain_nrs = data.get("pain_nrs")
                        if pain_nrs is not None and pain_nrs >= 3:
                            recs.append(f"🏥 *Операция в анамнезе ({_esc(surgery[:60])}) + боль сегодня ({pain_nrs}/10)*\n• Обязательно показать врачу, даже если боль слабая")

                    # 3. Страх рецидива высокий → ступенчатая экспозиция
                    # Показываем НЕ каждый день: только если анкета свежая (≤4 мес) и сегодня
                    # есть сопутствующий сигнал осторожности с нагрузкой (боль/крепатура/недосып/
                    # повышенное утомление/«болит на игре»/низкая готовность).
                    reinjury = str(qd.get("reinjury_fear") or "")
                    if reinjury == "Да, постоянно" and not q_stale:
                        _pain_today = (data.get("pain_nrs") is not None and data.get("pain_nrs") >= 3)
                        _on_game = bool(data.get("pain_on_game"))
                        _not_ready = (data.get("readiness") is not None and data.get("readiness") < 5)
                        _sore = (data.get("muscle_soreness") is not None and data.get("muscle_soreness") <= 3)
                        _tired = (data.get("fatigue_score") is not None and data.get("fatigue_score") <= 4)
                        _sleep = (data.get("sleep_score") is not None and data.get("sleep_score") <= 4)
                        if _pain_today or _on_game or _not_ready or _sore or _tired or _sleep:
                            recs.append("🧠 *Страх рецидива высокий*\n• Ступенчатое возвращение к нагрузке (начни с лёгкого)\n• Обсуди с врачом программу возврата")

                    # 4. Лекарства, влияющие на пульс → предупреждение (нормы RHR/HRV не применять)
                    meds = str(qd.get("meds_detail") or "")
                    if meds and any(k in meds.lower() for k in ("стимулят", "аддерал", "ритолин", "концерт", "страттера", "бета-блокат", "пропранолол", "атенолол", "метопролол")):
                        recs.append("💊 *Лекарства, влияющие на пульс*\n• Нормы пульса/HRV к тебе не применяются — судим по самочувствию")

                    # 5. Вода/диета низкие → гидратация
                    try:
                        water = float(qd.get("water") or 0)
                        if 0 < water < 1.5:
                            recs.append("💧 *Мало воды (анкета)*\n• Цель: 2-3л в день, пей до/во время/после тренировки")
                    except (TypeError, ValueError):
                        pass

                    # 6. Сезон → целевой диапазон нагрузки
                    season = str(qd.get("season") or "")
                    if "соревн" in season.lower() or "сезон" in season.lower():
                        pass  # целевой диапазон используется в Фазе 4 (прокси нагрузки)

                    # 7. Свежесть анкеты
                    if q_stale:
                        recs.append("📋 *Анкета устарела (>4 мес)* — обнови её (меню «Обновить анкету»), чтобы рекомендации были точными")
            except Exception as e:
                logger.warning(f"questionnaire recs: {e}")

        # ============ ТРЕНД-АНАЛИЗ (3 дня подряд) ============
        if athlete_id:
            try:
                hist3 = self.db.get_last_wellness(athlete_id, 3)
                if len(hist3) >= 3:
                    for metric, label, invert in [
                        ("sleep_score", "Сон", True),
                        ("readiness", "Готовность", True),
                        ("resting_hr", "Пульс", False),
                        ("fatigue_score", "Утомление", True),
                    ]:
                        vals = [h.get(metric) for h in hist3 if h.get(metric) is not None]
                        if len(vals) == 3:
                            if invert:
                                if vals[0] < vals[1] < vals[2]:
                                    recs.append(f"📉 *{label} снижается 3 дня* ({vals[2]}→{vals[1]}→{vals[0]})")
                                elif vals[0] > vals[1] > vals[2]:
                                    recs.append(f"📈 *{label} растёт 3 дня* ({vals[2]}→{vals[1]}→{vals[0]}) ✅")
                            else:
                                if vals[0] > vals[1] > vals[2]:
                                    recs.append(f"📈 *{label} растёт 3 дня* ({vals[2]}→{vals[1]}→{vals[0]})")
                                elif vals[0] < vals[1] < vals[2]:
                                    recs.append(f"📉 *{label} снижается 3 дня* ({vals[2]}→{vals[1]}→{vals[0]})")
            except Exception as e:
                logger.warning(f"trend analysis in recs: {e}")

        # Индивидуальные коридоры (личная норма 30 дн.) — если у спортсмена есть достаточная история
        if athlete_id:
            try:
                bl = self.db.get_individual_baseline(athlete_id, 30)
                if bl:
                    # для правила «2 отклонения подряд» берём предыдущий опрос (вчера)
                    hist = self.db.get_last_wellness(athlete_id, 2)
                    prev = hist[1] if len(hist) >= 2 else None
                    ind = self._individual_recs(data, bl, prev=prev)
                    if ind:
                        pad = "\n\n".join(ind)
                        recs.append("📊 *Отклонение от личной нормы (30 дн., 2 дня подряд):*\n" + pad)
            except Exception as e:
                logger.warning(f"individual baseline recs: {e}")

        return "\n\n".join(recs) if recs else "✅ Все показатели в норме!"

    def _individual_recs(self, data, bl, prev=None):
        """Алерты по личной норме (30 дн.), только при 2 ОТКЛОНЕНИЯХ ПОДРЯД (сегодня + вчера).
        Пороги: медиана ± max(1.5σ, 1.0) — устойчиво к выбросам.
        ВНИМАНИЕ: шкала опроса единая «7=хорошо, 1=плохо» для всех пяти показателей,
        поэтому для stress/fatigue/soreness плохо = балл НИЖЕ нормы (как и sleep/mood)."""
        median = bl.get("median") or {}
        std = bl.get("std") or {}

        def dev(value, kind):
            """True, если `value` выходит за личный порог (медиана - порог)."""
            if value is None:
                return False
            m = median.get(kind)
            if m is None:
                return False
            s = std.get(kind, 0) or 0
            threshold = max(1.0, 1.5 * s) if s > 0 else 1.5
            return value < m - threshold

        def dev_hr(value):
            """Пульс: выше медианы + max(1.5σ, +10%)."""
            if value is None or bl.get("median_hr") is None:
                return False
            s = bl.get("std_hr", 0) or 0
            thr = bl["median_hr"] + max(1.5 * s, bl["median_hr"] * 0.10)
            return value > thr

        out = []
        # если нет данных за «вчера» — не сигналим по индивидуальному (только общие коридоры)
        if prev is None:
            return out

        # пульс
        cur_hr = data.get("resting_hr")
        prev_hr = prev.get("resting_hr")
        if dev_hr(cur_hr) and dev_hr(prev_hr):
            out.append(f"❤️ *Пульс выше личной нормы 2 дня подряд ({cur_hr} vs ~{int(bl['median_hr'])}/30д)*\n• Обратить внимание на восстановление")

        checks = [
            ("sleep_score", "sleep", "😴 *Сон ниже личной нормы 2 дня подряд*"),
            ("mood_score", "mood", "😊 *Настроение ниже личной нормы 2 дня подряд*"),
            ("stress_score", "stress", "🧘 *Стресс выше личной нормы 2 дня подряд*"),
            ("fatigue_score", "fatigue", "⚡ *Утомление выше личной нормы 2 дня подряд*"),
            ("muscle_soreness", "soreness", "🤕 *Боль выше личной нормы 2 дня подряд*"),
        ]
        for key, kind, label in checks:
            cur = data.get(key); prv = prev.get(key)
            if dev(cur, kind) and dev(prv, kind):
                out.append(f"{label} ({cur} vs ~{median.get(kind):.1f}/30д)")
        return out
    


    # ==================== УДАЛЕНИЕ СПОРТСМЕНА (ТОЛЬКО АДМИН) ====================

    async def delete_athlete_menu(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        athletes = self.db.get_all_athletes()
        if not athletes:
            await q.edit_message_text("📭 Нет спортсменов.", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        buttons = []
        last_team = None
        for a in athletes:
            team = a.get("team") or "Без команды"
            if team != last_team:
                last_team = team
                buttons.append([(f"🏀 {team}", "hdr")])
            buttons.append([(f"❌ {a['full_name']}", f"del_{a['id']}")])
        buttons.append([(f"🔙 Назад", "main_menu")])

        await q.edit_message_text(
            "🗑 *Удаление спортсмена*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери кого удалить:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def delete_athlete_confirm(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        athlete_id = int(q.data.replace("del_", ""))
        athlete = self.db.get_athlete_by_id(athlete_id) if hasattr(self.db, 'get_athlete_by_id') else None

        # Fallback
        if not athlete:
            row = self.db.conn.execute("SELECT * FROM athletes WHERE id = ?", (athlete_id,)).fetchone()
            athlete = dict(row) if row else None

        if not athlete:
            await q.edit_message_text("❌ Спортсмен не найден.", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        await q.edit_message_text(
            f"⚠️ *Удалить {athlete['full_name']}?*\n\nВсе данные будут удалены.",
            reply_markup=self.kb([
                [("✅ Да, удалить", f"delconfirm_{athlete_id}")],
                [("❌ Нет", "admin_manage")]
            ]), parse_mode="Markdown"
        )

    async def delete_athlete_final(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        athlete_id = int(q.data.replace("delconfirm_", ""))
        self.db.conn.execute("DELETE FROM daily_wellness WHERE athlete_id = ?", (athlete_id,))
        self.db.conn.execute("DELETE FROM alerts WHERE athlete_id = ?", (athlete_id,))
        self.db.conn.execute("DELETE FROM athletes WHERE id = ?", (athlete_id,))
        self.db.conn.commit()

        await q.edit_message_text(
            "✅ Спортсмен удалён.",
            reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown"
        )

    # ==================== НАПОМИНАНИЯ ====================

    async def reminder_settings(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        text = (
            f"⏰ *Настройка напоминаний*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Текущее время: *{REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d}*\n\n"
            f"Выбери время или напиши своё (например: 14:30):"
        )

        buttons = [
            [("🌅 08:00", "set_reminder_8"), ("🌞 12:00", "set_reminder_12")],
            [("🌇 18:00", "set_reminder_18"), ("🌙 20:00", "set_reminder_20")],
            [("✏️ Своё время", "set_reminder_custom")],
            [("📨 Разослать напоминание сейчас", "send_reminder_now")],
            [("📋 Предложить заполнить анкету", "send_q_reminder")],
            [("🔕 Выключить", "set_reminder_off")],
            [("🔙 Назад", "main_menu")]
        ]

        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    def _schedule_reminder_job(self, hour, minute=0):
        """Пере)планировать ежедневную рассылку напоминаний. Старые задания снимаем во избежанie дублей."""
        if not self.job_queue or hour is None:
            return
        import pytz
        from datetime import datetime as dt
        local_tz = pytz.timezone(REMINDER_TZ)
        now_local = dt.now(local_tz)
        target_local = now_local.replace(hour=hour, minute=minute, second=0, microsecond=0)
        target_utc = target_local.astimezone(pytz.UTC)
        for old in self.job_queue.get_jobs_by_name("daily_reminder"):
            old.schedule_removal()
        self.job_queue.run_daily(
            self._send_daily_reminder,
            time=target_utc.time(),
            days=tuple(range(7)),
            name="daily_reminder"
        )

    def _unschedule_reminder_job(self):
        """Полностью убрать ежедневную рассылку напоминаний."""
        if not self.job_queue:
            return
        for old in self.job_queue.get_jobs_by_name("daily_reminder"):
            old.schedule_removal()

    async def set_reminder_time(self, update, ctx):
        global REMINDER_HOUR, REMINDER_MINUTE
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        data = q.data.replace("set_reminder_", "")
        if data == "off":
            REMINDER_HOUR = None
            self._unschedule_reminder_job()
            self.db.set_setting("reminder_hour", "")
            await q.edit_message_text("🔕 Напоминания выключены.", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        if data == "custom":
            state = self.get_state(q.from_user.id)
            state["step"] = "set_reminder_custom"
            await q.edit_message_text("✏️ *Напиши время в формате ЧЧ:ММ*\n\nНапример: 09:30 или 14:15", parse_mode="Markdown", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        if data == "send_now":
            await q.edit_message_text("📨 *Отправляю напоминания спортсменам...*", parse_mode="Markdown")
            try:
                await self._send_daily_reminder(ctx)
                await q.edit_message_text("✅ *Напоминания отправлены!*", parse_mode="Markdown", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            except Exception as e:
                logger.error(f"Send now error: {e}")
                await q.edit_message_text(f"❌ Ошибка: {e}", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        hour = int(data)
        REMINDER_HOUR = hour
        REMINDER_MINUTE = 0
        REMINDER_TZ = "Asia/Yekaterinburg"
        self.db.set_setting("reminder_hour", str(hour))
        self.db.set_setting("reminder_tz", REMINDER_TZ)
        self._schedule_reminder_job(hour, 0)

        await q.edit_message_text(
            f"⏰ Напоминание настроено на *{hour:02d}:00* по Челябинску\n\n"
            f"Спортсменам будет приходить уведомление в это время ежедневно.",
            reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown"
        )

    async def send_reminder_now(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        await q.edit_message_text("📨 *Отправляю напоминания спортсменам...*", parse_mode="Markdown")
        try:
            await self._send_daily_reminder(ctx)
            await q.edit_message_text("✅ *Напоминания отправлены!*", parse_mode="Markdown", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
        except Exception as e:
            logger.error(f"Send now error: {e}")
            await q.edit_message_text(f"❌ Ошибка: {e}", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))

    async def send_questionnaire_reminder(self, update, ctx):
        """Рассылка тем спортсменам, у кого не заполнена анкета здоровья."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        await q.edit_message_text("📋 *Отправляю предложение заполнить анкету...*", parse_mode="Markdown")
        try:
            bot = ctx.bot
            athletes = self.db.get_all_athletes()
            target = []
            for a in athletes:
                if self.db.is_coach(a["telegram_id"]):
                    continue
                if self.db.has_questionnaire(a["id"]):
                    continue
                target.append(a)
            sent = 0
            failed = 0
            for a in target:
                try:
                    await bot.send_message(
                        chat_id=a["telegram_id"],
                        text=(
                            "📋 *ЧБК — Анкета здоровья*\n"
                            "━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                            f"👋 Привет, {self._first_name(a)}!\n\n"
                            "Мы видим, что ты ещё не заполнил(а) *анкету спортсмена*.\n"
                            "Она нужна врачу: рост, вес, амплуа, травмы, противопоказания.\n\n"
                            "⏱️ Это займёт пару минут.\n\n"
                            "👉 Нажми /start → выбери *«📋 Заполнить анкету»*"
                        ),
                        parse_mode="Markdown"
                    )
                    sent += 1
                except Exception as e:
                    failed += 1
                    logger.error(f"Q-reminder error for {a['full_name']} (tg={a['telegram_id']}): {e}")
            logger.info(f"Questionnaire remind sent={sent} failed={failed} of {len(target)}")
            await q.edit_message_text(
                f"✅ *Предложение отправлено:* {sent}\n"
                f"⏭️ Пропущено (уже заполнили): {len(athletes) - len(target)}\n"
                f"❌ Не доставилось: {failed}\n\n"
                f"Без анкеты осталось: {len(target)}",
                parse_mode="Markdown", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]])
            )
        except Exception as e:
            logger.error(f"Send questionnaire reminder error: {e}")
            await q.edit_message_text(f"❌ Ошибка: {e}", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))

    # ==================== ОТЧЕТ ПО ДНЯМ ====================

    async def daily_report(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        athletes = self.db.get_all_athletes()
        today = date.today()

        passed = []
        not_passed = []
        for a in athletes:
            if self.db.has_survey_today(a["id"]):
                passed.append(a)
            else:
                not_passed.append(a)

        text = (
            f"📅 *Отчёт за {today.strftime('%d.%m.%Y')}*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"✅ *Прошли опрос ({len(passed)}):*\n"
        )
        for a in passed:
            text += f"  • {a['full_name']} ({a['team']})\n"

        if not_passed:
            text += f"\n❌ *Не прошли ({len(not_passed)}):*\n"
            for a in not_passed:
                text += f"  • {a['full_name']} ({a['team']})\n"

        text += f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"Итого: {len(passed)}/{len(athletes)} прошли"

        buttons = [[(f"📊 Экспорт CSV", "export_csv")], [(f"🔙 Назад", "main_menu")]]
        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    # ==================== ОТЧЕТ ДЛЯ ВРАЧА ====================

    async def show_admin_manage(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return

        text = (
            f"⚙️ *Управление ботом*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Выбери действие:"
        )
        buttons = [
            [(f"👥 Список спортсменов", "athlete_list")],
            [(f"➕ Добавить спортсмена", "admin_add_athlete")],
            [(f"🔒 Блокировка", "ban_menu")],
            [(f"📋 Анкеты", "questionnaire_list")],
            [(f"📅 Отчёт за сегодня", "daily_report")],
            [(f"🚨 Жалобы", "admin_complaints")],
            [(f"🗑 Удалить спортсмена", "delete_athlete")],
            [(f"⏰ Напоминания", "reminder_settings")],
            [(f"🏅 Тренеры", "coach_menu")],
            [(f"🩺 Врачи", "doctor_menu")],
            [(f"📄 PDF-отчёт", "pdf_report_menu")],
            [(f"📊 Экспорт CSV", "export_csv")],
            [(f"🔙 Назад", "main_menu")]
        ]
        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    # ==================== ДОБАВЛЕНИЕ СПОРТСМЕНА АДМИНОМ ====================

    async def admin_add_athlete_start(self, update, ctx):
        """Начать процесс добавления спортсмена — запрос Telegram ID."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        state = self.get_state(q.from_user.id)
        state["step"] = "admin_add_athlete_tg_id"
        state["data"] = {}
        await q.edit_message_text(
            "➕ *Добавить спортсмена*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Пришли *Telegram ID* спортсмена (число):",
            reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]]),
            parse_mode="Markdown"
        )

    async def admin_add_athlete_team(self, update, ctx):
        """Выбор команды для нового спортсмена."""
        q = update.callback_query
        await q.answer()
        team = q.data.replace("admin_add_team_", "")
        state = self.get_state(q.from_user.id)
        state["data"]["add_team"] = team
        buttons = [[(desc, f"admin_add_age_{key}")] for key, desc in AGE_GROUPS.items()]
        buttons.append([("🔙 Назад", "admin_add_athlete")])
        await q.edit_message_text(
            f"➕ *Команда:* {team}\n\nВыбери *возрастную группу:*",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def admin_add_athlete_age(self, update, ctx):
        """Выбор возрастной группы — запрос ФИО."""
        q = update.callback_query
        await q.answer()
        age = q.data.replace("admin_add_age_", "")
        state = self.get_state(q.from_user.id)
        state["data"]["add_age_group"] = age
        state["step"] = "admin_add_athlete_name"
        await q.edit_message_text(
            f"➕ *Команда:* {state['data']['add_team']}\n"
            f"*Группа:* {AGE_GROUPS.get(age, age)}\n\n"
            f"Введи *Фамилию Имя* спортсмена:",
            parse_mode="Markdown"
        )

    async def admin_add_athlete_save(self, update, user_id, state):
        """Сохранить нового спортсмена в БД."""
        data = state.get("data", {})
        tg_id = data.get("add_tg_id")
        team = data.get("add_team")
        age_group = data.get("add_age_group")
        name = data.get("add_name")

        if not all([tg_id, team, age_group, name]):
            await update.message.reply_text("❌ Ошибка: не все данные заполнены.",
                reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]]))
            return

        # Проверяем, нет ли уже такого telegram_id
        existing = self.db.get_athlete_by_telegram_id(int(tg_id))
        if existing:
            await update.message.reply_text(
                f"⚠️ Спортсмен с ID {tg_id} уже зарегистрирован:\n"
                f"👤 {existing['full_name']} | {existing['team']}",
                reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]])
            )
            return

        ok = self.db.register_athlete(
            telegram_id=int(tg_id), username=None,
            full_name=name, age_group=age_group, team=team
        )
        if ok:
            await update.message.reply_text(
                f"✅ *Спортсмен добавлен!*\n\n"
                f"👤 {name}\n"
                f"🏀 Команда: {team}\n"
                f"📋 Группа: {AGE_GROUPS.get(age_group, age_group)}\n"
                f"🆔 Telegram ID: {tg_id}\n\n"
                f"Попроси спортсмена написать /start в боте.",
                reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]]),
                parse_mode="Markdown"
            )
        else:
            await update.message.reply_text(
                "❌ Ошибка сохранения. Возможно, этот ID уже есть в базе.",
                reply_markup=self.kb([[(f"🔙 Управление", "admin_manage")]])
            )
        self.clear_state(user_id)

    async def admin_complaints_menu(self, update, ctx):
        """Панель жалоб: список спортсменов с жалобами за период."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        buttons = [
            [("📅 Сегодня", "admin_complaints_page_1")],
            [("📆 Неделя", "admin_complaints_page_7")],
            [("📊 Месяц", "admin_complaints_page_30")],
            [("🔙 Управление", "admin_manage")],
        ]
        await q.edit_message_text(
            "🚨 *Жалобы спортсменов*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери период:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def admin_complaints_page(self, update, ctx):
        """Показать жалобы за N дней."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        days = int(q.data.replace("admin_complaints_page_", ""))
        complaints = self.db.get_athletes_with_complaints(days)
        if not complaints:
            period = {1: "сегодня", 7: "неделю", 30: "месяц"}.get(days, f"{days} дн.")
            await q.edit_message_text(
                f"✅ Жалоб за {period} нет!",
                reply_markup=self.kb([[(f"🔙 Назад", "admin_complaints"), (f"🏠 Главное меню", "main_menu")]])
            )
            return
        text = f"🚨 *Жалобы за {days} дн. ({len(complaints)}):*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        btns = []
        for c in complaints[:15]:  # лимит чтобы не переполнить
            date_str = c.get("survey_date", "?")
            text += f"👤 *{c['full_name']}* ({c.get('team', '?')})\n"
            text += f"   📅 {date_str} | 💬 {c.get('complaints', '?')}\n\n"
            if c.get("telegram_id"):
                btns.append([InlineKeyboardButton(f"💬 Написать: {c['full_name']}", url=f"tg://user?id={c['telegram_id']}")])
        if len(complaints) > 15:
            text += f"… и ещё {len(complaints) - 15}\n"
        btns.append([("🔙 Назад", "admin_complaints"), (f"🏠 Главное меню", "main_menu")])
        await q.edit_message_text(text, reply_markup=self.kb(btns), parse_mode="Markdown")

    async def pdf_report_menu(self, update, ctx):
        """Меню PDF-отчётов: выбрать спортсмена."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        athletes = self._scoped_athletes(q.from_user.id)
        if not athletes:
            await q.edit_message_text("📭 Нет спортсменов.", reply_markup=self.kb([[(f"🔙 Назад", "admin_manage")]]))
            return
        buttons = []
        for a in athletes[:20]:
            buttons.append([(f"📄 {a['full_name']}", f"pdf_{a['id']}")])
        buttons.append([("🔙 Назад", "admin_manage")])
        await q.edit_message_text(
            "📄 *PDF-отчёт по спортсмену*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери спортсмена:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def pdf_report_generate(self, update, ctx):
        """Сгенерировать и отправить PDF-отчёт (текстовый) по спортсмену."""
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            return
        try:
            athlete_id = int(q.data.replace("pdf_", ""))
        except ValueError:
            return
        try:
            from reports import ReportGenerator
            gen = ReportGenerator()
            gen.db = self.db  # используем ту же БД (внешний файл создаёт свою)
            path = gen.generate_athlete_report(athlete_id, days=7)
            if not path:
                await q.edit_message_text("❌ Нет данных за 7 дней.", reply_markup=self.kb([[(f"🔙 PDF-отчёт", "pdf_report_menu")]]))
                return
            athlete = self.db.get_athlete_by_id(athlete_id)
            name = athlete["full_name"] if athlete else "спортсмен"
            with open(path, "rb") as f:
                await ctx.bot.send_document(
                    chat_id=q.from_user.id,
                    document=f,
                    filename=f"отчёт_{name}_{date.today()}.txt",
                    caption=f"📄 Отчёт: {name} (7 дней)"
                )
            await q.edit_message_text(
                "✅ Отчёт отправлен.",
                reply_markup=self.kb([[(f"🔙 PDF-отчёт", "pdf_report_menu"), (f"🏠 Главное меню", "main_menu")]])
            )
        except Exception as e:
            logger.error(f"PDF report error: {e}")
            await q.edit_message_text(f"❌ Ошибка: {e}", reply_markup=self.kb([[(f"🔙 Назад", "admin_manage")]]))

    async def show_admin_report(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id):
            await q.edit_message_text("❌ Нет доступа.")
            return

        data = self.db.get_athletes_week_stats()
        teams = {}
        for a in data:
            teams.setdefault(a.get("team", "?"), []).append(a)

        text = f"📋 *ОТЧЕТ ДЛЯ ВРАЧА*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n📅 {date.today().strftime('%d.%m.%Y')}\n\n"

        total_ok, total_warn, total_crit = 0, 0, 0

        for tn in sorted(teams.keys()):
            ta = teams[tn]
            text += f"*🏀 {tn}* ({len(ta)}):\n"
            for a in ta:
                name = a["full_name"]
                streak = a.get("survey_streak", 0)
                days = a.get("days_count", 0)
                if days > 0:
                    sleep = a.get("avg_sleep")
                    stress = a.get("avg_stress")
                    fatigue = a.get("avg_fatigue")
                    hr = a.get("avg_hr")

                    issues = []
                    # Шкалы 1-7, 7=хорошо (сон/утомление: низкий балл = плохо; стресс: НИЗКИЙ балл = «бесит всё» = плохо)
                    if sleep and sleep < 3: issues.append("сон🔴"); total_crit += 1
                    elif sleep and sleep < 5: issues.append("сон🟡"); total_warn += 1
                    if stress and stress < 3: issues.append("стресс🔴"); total_crit += 1
                    elif stress and stress < 5: issues.append("стресс🟡"); total_warn += 1
                    if fatigue and fatigue < 3: issues.append("утом.🔴"); total_crit += 1
                    elif fatigue and fatigue < 5: issues.append("утом.🟡"); total_warn += 1
                    if hr and hr > 70: issues.append("пульс🔴"); total_crit += 1
                    elif hr and hr > 60: issues.append("пульс🟡"); total_warn += 1

                    status = "🟢" if not issues else ("🔴" if any("🔴" in i for i in issues) else "🟡")
                    if not issues: total_ok += 1

                    ss = f"{sleep:.1f}" if sleep else "—"
                    sts = f"{stress:.1f}" if stress else "—"
                    fs = f"{fatigue:.1f}" if fatigue else "—"
                    hs = f"{hr:.0f}" if hr else "—"

                    text += f"  {status} *{name}* ({a['age_group']})\n"
                    text += f"     😴{ss} 😰{sts} 😩{fs} ❤️{hs} | 🔥{streak}д\n"
                    if issues:
                        text += f"     ⚠️ {', '.join(issues)}\n"
                else:
                    text += f"  ⚪ *{name}* — нет данных\n"
            text += "\n"

        text += f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        text += f"👥 Всего: {len(data)} | 🟢 {total_ok} | 🟡 {total_warn} | 🔴 {total_crit}\n"

        await q.edit_message_text(text, reply_markup=self.kb([
            [(f"📊 Экспорт CSV", "export_csv")],
            [(f"🔙 Назад", "main_menu")]
        ]), parse_mode="Markdown")

    async def export_csv(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        athletes = self._scoped_athletes(uid)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ФИО", "Группа", "Команда", "Серия", "Опросов", "Сон", "Утомление", "Пульс", "Активность"])

        for a in athletes:
            stats = self.db.get_athlete_stats(a["id"], 7)
            writer.writerow([
                a["full_name"], a["age_group"], a["team"],
                a.get("survey_streak", 0), a.get("total_surveys", 0),
                round(stats.get("avg_sleep", 0) or 0, 1),
                round(stats.get("avg_fatigue", 0) or 0, 1),
                round(stats.get("avg_hr", 0) or 0, 1),
                a.get("last_active", "—")
            ])

        output.seek(0)
        await q.message.reply_document(
            document=output.getvalue().encode("utf-8-sig"),
            filename=f"report_{date.today().strftime('%Y%m%d')}.csv",
            caption="📊 Отчет по спортсменам"
        )

    async def export_questionnaires_xlsx(self, update, ctx):
        """Экспорт анкет спортсменов в Excel (админ — все, тренер — свои команды)."""
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        athletes = self._scoped_athletes(uid)
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Анкеты"

        hf = Font(bold=True, size=11, color="FFFFFF")
        hfl = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # поля анкеты (ключ -> заголовок)
        fields = [
            ("age", "Возраст"), ("gender", "Пол"), ("phone", "Телефон"), ("birth_date", "Дата рождения"), ("position", "Позиция"),
            ("level", "Уровень"), ("experience", "Стаж"), ("height", "Рост"),
            ("weight", "Вес"), ("trauma_12m", "Травмы 12мес"),
            ("trauma_12m_detail", "Травмы детали"), ("zones", "Проблемные зоны"),
            ("pain_now", "Боль сейчас"), ("pain_now_detail", "Боль детали"),
            ("chronic", "Рецидивы"), ("chronic_detail", "Рецидивы детали"),
            ("surgery", "Операции"), ("surgery_detail", "Операции детали"), ("surgery_date", "Операции даты"),
            ("meds", "Лекарства/БАДы"), ("meds_detail", "БАДы детали"),
            ("allergies", "Аллергии"), ("allergies_detail", "Аллергии детали"),
            ("train_count", "Трен/нед"), ("train_duration", "Длит. трен"),
            ("season", "Сезон"), ("form_score", "Форма"), ("sleep_score", "Сон (анкета)"),
            ("warmup", "Разминка"), ("recovery", "Восстановление"),
            ("water", "Вода л"), ("diet", "Питание"), ("pre_meal", "За сколько ест"),
            ("supplements", "Спортпит"), ("motivation", "Мотивация"),
            ("stress", "Внешний стресс"), ("match_state", "Состояние перед матчем"),
            ("reinjury_fear", "Страх травмы"), ("goal", "Цель"), ("wish", "Пожелания"),
        ]

        # ---- Определяем блоки анкеты (для читабельной двухуровневой шапки) ----
        # блок: (название, цвет, список ключей)
        blocks = [
            ("ОБЩИЕ ДАННЫЕ", "4472C4", ["age","gender","phone","birth_date","position","level","experience","height","weight"]),
            ("ТРАВМЫ / ЗДОРОВЬЕ", "70AD47", ["trauma_12m","trauma_12m_detail","zones","pain_now","pain_now_detail",
                                                "chronic","chronic_detail","surgery","surgery_detail","surgery_date",
                                                "meds","meds_detail","allergies","allergies_detail"]),
            ("ТРЕНИРОВКИ", "ED7D31", ["train_count","train_duration","season","form_score","sleep_score","warmup","recovery"]),
            ("ПИТАНИЕ", "FFC000", ["water","diet","pre_meal","supplements"]),
            ("ПСИХОЛОГИЯ", "A9A9A9", ["motivation","stress","match_state","reinjury_fear","goal","wish"]),
        ]
        # индекс начала каждой колонки-поля: базовые (ФИО,Команда,Группа,Статус) = 1-4, поля с 5
        key_col = {}
        ci = 5
        for bl in blocks:
            for k in bl[2]:
                key_col[k] = ci
                ci += 1
        total_cols = ci - 1  # 4 базовых + все поля

        def _col_letter(n):
            return get_column_letter(n)

        ws.merge_cells(f"A1:{_col_letter(total_cols)}1")
        c = ws.cell(row=1, column=1, value="🏀 ЧБК — АНКЕТЫ СПОРТСМЕНОВ")
        c.font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells(f"A2:{_col_letter(total_cols)}2")
        c = ws.cell(row=2, column=1, value=f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Всего: {len(athletes)}")
        c.font = Font(bold=True, size=10, color="808080")

        # ---- Двухуровневая шапка: строка 4 = блоки, строка 5 = поля ----
        r_hdr = 4
        # базовые колонки 1-4
        base_block_cols = [("СПОРТСМЕН", "4472C4", 4)]
        # блоки-колонки
        block_ranges = []
        cur = 5
        for bname, bcolor, keys in blocks:
            n_keys = len(keys)
            block_ranges.append((bname, bcolor, cur, cur + n_keys - 1))
            cur += n_keys
        # мержим блоки в строке 4
        for bname, bcolor, c0, c1 in block_ranges:
            ws.merge_cells(start_row=r_hdr, start_column=c0, end_row=r_hdr, end_column=c1)
        # базовые 4 колонки объединяем в "Спортсмен"
        ws.merge_cells(start_row=r_hdr, start_column=1, end_row=r_hdr, end_column=4)
        # заполняем заголовки блоков
        from openpyxl.styles import PatternFill as _PF2
        def _block_cell(col, text, color):
            cell = ws.cell(row=r_hdr, column=col, value=text)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = hdr_align
            cell.border = thin_border
            return cell
        _block_cell(1, "СПОРТСМЕН", "4472C4")
        for bname, bcolor, c0, c1 in block_ranges:
            _block_cell(c0, bname, bcolor)
        ws.row_dimensions[r_hdr].height = 20

        # строка 5 = названия полей
        r2 = 5
        base_headers = ["ФИО", "Команда", "Группа", "Статус"]
        for ci2, h in enumerate(base_headers, 1):
            c = ws.cell(row=r2, column=ci2, value=h)
            c.font = hf
            c.fill = hfl
            c.alignment = hdr_align
            c.border = thin_border
        for key, h in fields:
            ci2 = key_col[key]
            c = ws.cell(row=r2, column=ci2, value=h)
            # подкрасить заголовок поля цветом блока
            for bname, bcolor, c0, c1 in block_ranges:
                if c0 <= ci2 <= c1:
                    c.fill = PatternFill(start_color=bcolor, end_color=bcolor, fill_type="solid")
            c.font = hf
            c.alignment = hdr_align
            c.border = thin_border
        ws.freeze_panes = f"B6"
        ws.auto_filter.ref = f"A{r2}:{_col_letter(total_cols)}{r2}"

        r = 6
        zebra = False
        for a in athletes:
            qd = self.db.get_questionnaire(a["id"]) or {}
            is_done = bool(qd.get("completed_at"))
            status = "✅ Заполнена" if is_done else "⬜ Не заполнена"
            row_vals = [a["full_name"], a.get("team", ""), a.get("age_group", ""), status]
            for key, _ in fields:
                v = qd.get(key, "")
                if isinstance(v, float):
                    v = round(v, 1)
                row_vals.append(v if v is not None else "")
            row_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid") if zebra else None
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = thin_border
                if row_fill:
                    c.fill = row_fill
                if ci == 4:
                    c.font = Font(bold=True, size=10)
                    c.alignment = Alignment(horizontal="center")
            r += 1
            zebra = not zebra

        # ширины
        from openpyxl.utils import get_column_letter
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
        for ci in range(5, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        # ---- Лист «Легенда»: расшифровка значений для читателя ----
        ws_l = wb.create_sheet("Легенда")
        ws_l.column_dimensions["A"].width = 34
        ws_l.column_dimensions["B"].width = 70
        lr = 1
        c = ws_l.cell(lr, 1, "ЧТО ЗНАЧИТ КАЖДОЕ ПОЛЕ В АНКЕТЕ")
        c.font = Font(bold=True, size=13, color="1F4E79"); lr += 1
        c = ws_l.cell(lr, 1, "Цвет поля в шапке = блок анкеты. Краткая расшифровка ниже.")
        c.font = Font(size=9, italic=True, color="808080"); lr += 2
        legend_rows = [
            ("СПОРТСМЕН", "ФИО, команда, возрастная группа, заполнена ли анкета (✅/⬜)."),
            ("Возраст / Дата рождения", "Полные годы и дата рождения для мед. карты."),
            ("Телефон", "Номер для связи (спортсмен видит у себя)."),
            ("Пол", "М / Ж. Для девушек в ежедневном опросе добавляются вопросы цикла."),
            ("Позиция / Уровень / Стаж", "PG/SG/SF/PF/C/Универсал; уровень игры; стаж в баскетболе."),
            ("Рост / Вес", "см / кг (например 185/82)."),
            ("Травмы 12 мес", "Да/Нет + детали: какие травмы и как лечил."),
            ("Проблемные зоны", "Какие зоны беспокоили (голеностоп, колени, спина…)."),
            ("Боль сейчас", "Есть ли боль прямо сейчас + где и как давно."),
            ("Рецидивы", "Повторяющаяся травма: какая, как часто обостряется."),
            ("Операции", "Да/Нет + что за операции + дата/год."),
            ("Лекарства / БАДы", "Принимает ли лекарства/БАД/физиотерапию + какие."),
            ("Аллергии", "Да/Нет + на что аллергия (если Да)."),
            ("Тренировки/нед", "Сколько тренировок в неделю (0-10)."),
            ("Длит. тренировки", "В минутах (30-180)."),
            ("Сезон", "Сезон / Предсезонка / Межсезонье / Пауза."),
            ("Форма (1-10)", "10 = отличная форма, 1 = плохая."),
            ("Сон в анкете (1-10)", "10 = отличный сон, 1 = плохой (базовое качество сна)."),
            ("Разминка / Восстановление", "Практикует ли разминку/заминку и восстановительные дни."),
            ("Вода (л) / Питание / За сколько ест", "Пьёт воды в день; есть ли план питания; время до тренировки."),
            ("Спортпит", "Протеин/Креатин/Хондропротекторы/Изотоники/Кофеин/Другое (через запятую)."),
            ("Мотивация (1-10)", "10 = максимальная мотивация."),
            ("Внешний стресс / Состояние перед матчем", "Мешают ли внешние факторы; тревога перед играми."),
            ("Страх травмы", "Боится ли повторной травмы, насколько сильно."),
            ("Цель / Пожелания", "Главная цель и особая нужда в помощи (для врача)."),
        ]
        for col1, col2 in legend_rows:
            c = ws_l.cell(lr, 1, col1); c.font = Font(bold=True, size=10)
            c = ws_l.cell(lr, 2, col2); c.font = Font(size=10)
            lr += 1
        lr += 1
        c = ws_l.cell(lr, 1, "ЦВЕТА БЛОКОВ (в шапке таблицы Анкеты)")
        c.font = Font(bold=True, size=11); lr += 1
        for bname, bcolor, *_ in blocks:
            cc = ws_l.cell(lr, 1, bname)
            cc.fill = PatternFill(start_color=bcolor, end_color=bcolor, fill_type="solid")
            cc.font = Font(bold=True, size=10, color="FFFFFF")
            lr += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        await q.message.reply_document(
            document=output,
            filename=f"ankety_{date.today().strftime('%Y%m%d')}.xlsx",
            caption="📋 Анкеты спортсменов (справка для врача)"
        )

    # ==================== ЭКСПОРТ В EXCEL ====================

    async def report_export_menu(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return
        is_admin = self._is_full_access(uid)
        if is_admin:
            teams = set(ACTIVE_TEAMS) & set(a["team"] for a in self.db.get_all_athletes() if a.get("team"))
        else:
            teams = self._coach_teams(uid) & set(ACTIVE_TEAMS)

        buttons = []
        if is_admin:
            buttons.append([(f"👥 Все команды", "report_team_all")])
        for t in sorted(teams):
            buttons.append([(f"🏀 {t}", f"report_team_{t}")])
        if is_admin:
            buttons.append([(f"📋 Экспорт анкет", "export_q")])
            buttons.append([(f"⌚ Отчёт по часам", "report_watches")])
        buttons.append([(f"⚖️ Отчёт по весам", "report_body_comp")])
        buttons.append([(f"🔙 Назад", "main_menu")])

        await q.edit_message_text(
            "📊 *Экспорт в Excel*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери команду:",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def report_choose_period(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        team = q.data.replace("report_team_", "")
        # Тренер может выбирать только свои команды
        if not self._is_full_access(uid):
            if team != "all" and team not in self._coach_teams(uid):
                await q.edit_message_text("🔒 Нет доступа к этой команде.",
                                           reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
                return
        state = self.get_state(uid)
        state["data"]["report_team"] = team

        buttons = [
            [("📅 Сегодня", f"report_period_1")],
            [("📅 Последние 7 дней", f"report_period_7")],
            [("📅 Последние 30 дней", f"report_period_30")],
            [("📅 Последние 90 дней", f"report_period_90")],
            [("🔙 Назад", "report_export_menu")]
        ]

        display_team = "Все команды" if team == "all" else team
        await q.edit_message_text(
            f"📊 *Выбери период*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nКоманда: *{display_team}*\n\nЗа какой период выгрузить данные?",
            reply_markup=self.kb(buttons), parse_mode="Markdown"
        )

    async def generate_xlsx_report(self, update, ctx):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        days = int(q.data.replace("report_period_", ""))
        state = self.get_state(uid)
        team = state.get("data", {}).get("report_team", "all")
        # Тренеру запрещён «все команды»
        if not self._is_full_access(uid):
            if team == "all" or team not in self._coach_teams(uid):
                await q.edit_message_text("🔒 Нет доступа к этой команде.",
                                           reply_markup=self.kb([[(f"🔙 Главное меню", "main_menu")]]))
                return
        team_filter = None if team == "all" else team

        period_names = {1: "последний замер", 14: "2 недели", 30: "месяц", 90: "3 месяца", 999: "всё время"}

        wb = Workbook()

        # ---- Стили ----
        hf = Font(bold=True, size=11, color="FFFFFF")
        hfl = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        good_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        crit_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        # Лёгкая зебра для блоков дней (чередование) + жирная граница стыка дней
        zebra_fill = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
        zebra_hdr_fill = PatternFill(start_color="8EAADB", end_color="8EAADB", fill_type="solid")
        day_side = Side(style="medium", color="1F4E79")
        day_right = Border(left=thin_border.left, right=day_side, top=thin_border.top, bottom=thin_border.bottom)
        day_left = Border(left=day_side, right=thin_border.right, top=thin_border.top, bottom=thin_border.bottom)
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=10, color="808080")
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap = Alignment(wrap_text=True, vertical="top")

        def cell(ws, row, col, value, font=None, fill=None, align=None, border=thin_border):
            c = ws.cell(row=row, column=col, value=value)
            if font: c.font = font
            if fill is not None: c.fill = fill
            if align: c.alignment = align
            if border: c.border = border
            return c

        # ---- Получаем данные по дням ----
        # Сводка по командам (средние для шапки) и детальные опросы
        summary_by_team = {}
        athletes = self.db.get_team_stats_period(team_filter, days)
        if team_filter is None:
            # «Все команды» = только активные (остальные в БД сохраняются, но скрыты)
            _active = set(ACTIVE_TEAMS)
            athletes = [a for a in athletes if a.get("team") in _active]
        for a in athletes:
            t = a.get("team", "?")
            summary_by_team.setdefault(t, []).append(a)

        # Детальные опросы по дням
        wellness_rows = self.db.get_wellness_by_period(days, team_filter)
        if team_filter is None:
            wellness_rows = [r for r in wellness_rows if r.get("team") in set(ACTIVE_TEAMS)]

        # Группируем детальные опросы по командам
        rows_by_team = {}
        for row in wellness_rows:
            t = row.get("team", "?")
            rows_by_team.setdefault(t, []).append(row)

        # Команды для листов
        all_teams = sorted(set(list(summary_by_team.keys()) + list(rows_by_team.keys())))
        if not all_teams:
            await q.edit_message_text("❌ Нет данных за выбранный период.", reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        # ---- ЛИСТ СВОДКА (первый) ----
        ws0 = wb.active
        ws0.title = "Сводка"
        ws0.merge_cells("A1:H1")
        cell(ws0, 1, 1, "📋 ЧБК — СВОДНЫЙ ОТЧЁТ", font=title_font, border=None)
        ws0.merge_cells("A2:H2")
        cell(ws0, 2, 1, f"Период: {period_names[days]} | Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", font=subtitle_font, border=None)

        r = 4
        sum_headers = ["Команда", "Спортсменов", "С опросами", "Опросов", "Участие %", "Сон ср.", "Утомл. ср."]
        for ci, h in enumerate(sum_headers, 1):
            cell(ws0, r, ci, h, font=hf, fill=hfl, align=hdr_align)
        r += 1

        for t in all_teams:
            ta = summary_by_team.get(t, [])
            rows = rows_by_team.get(t, [])
            total = len(ta)
            with_data = sum(1 for a in ta if a.get("days_count", 0) > 0)
            if total > 0 and with_data > 0:
                # средние из сводки
                avg_s = sum((a.get("avg_sleep") or 0) for a in ta if a.get("avg_sleep")) / sum(1 for a in ta if a.get("avg_sleep"))
                avg_f = sum((a.get("avg_fatigue") or 0) for a in ta if a.get("avg_fatigue")) / sum(1 for a in ta if a.get("avg_fatigue"))
                particip = round(with_data / total * 100) if total else 0
            else:
                avg_s = avg_f = 0
                particip = 0
            cell(ws0, r, 1, t, font=bold_font)
            cell(ws0, r, 2, total, align=center)
            cell(ws0, r, 3, with_data, align=center)
            cell(ws0, r, 4, len(rows), align=center)
            cell(ws0, r, 5, f"{particip}%", align=center)
            cell(ws0, r, 6, f"{avg_s:.1f}", align=center, fill=self._xl_score_fill(avg_s))
            cell(ws0, r, 7, f"{avg_f:.1f}", align=center, fill=self._xl_score_fill(avg_f))
            r += 1

        # Итоговая строка по всем командам
        total_ath = sum(len(summary_by_team.get(t, [])) for t in all_teams)
        total_rows = sum(len(rows_by_team.get(t, [])) for t in all_teams)
        cell(ws0, r, 1, "ИТОГО", font=Font(bold=True, size=11))
        cell(ws0, r, 2, total_ath, font=bold_font, align=center)
        cell(ws0, r, 3, "-", align=center)
        cell(ws0, r, 4, total_rows, font=bold_font, align=center)
        for ci in range(5, 8):
            cell(ws0, r, ci, "-", align=center)
        r += 1

        # эмблема клуба в правом верхнем углу листа «Сводка»
        self._add_logo_to_ws(ws0, anchor="J3", size=90)

        for ci in range(1, 9):
            ws0.column_dimensions[chr(64+ci)].width = 18

        # ---- ЛИСТЫ ПО КОМАНДАМ: детальные опросы по дням ----
        for t in all_teams:
            ws = wb.create_sheet(title=t[:31])
            ta = summary_by_team.get(t, [])
            rows = rows_by_team.get(t, [])

            r = 1
            ws.merge_cells("A1:L1")
            cell(ws, r, 1, f"🏀 ЧБК — Опросы по дням: {t}", font=title_font, border=None)
            r += 1
            ws.merge_cells(f"A{r}:L{r}")
            cell(ws, r, 1, f"Период: {period_names[days]} | Дата отчёта: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Всего спортсменов: {len(ta)}", font=subtitle_font, border=None)
            r += 1

            # ===== БЛОК: МАТРИЦА «ИГРОК-СТРОКА, ДЕНЬ-СТОЛБЕЦ, ПОДСТОЛБЦЫ=ПОКАЗАТЕЛИ» =====
            from openpyxl.utils import get_column_letter
            # Ключи row: full_name, team, age_group, athlete_id, survey_date,
            # sleep_score, stress_score, fatigue_score, muscle_soreness, mood_score,
            # resting_hr, hrv_ms, had_training, sRPE_score, cycle_day, cycle_phase, complaints
            metrics = ["Сон", "Утомл", "Крепатура", "Боль", "Пульс", "sRPE", "Hooper"]
            namedays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

            def _m_hindex(r):
                v = [r.get("sleep_score"), r.get("fatigue_score"),
                     r.get("muscle_soreness")]
                v = [x for x in v if x is not None]
                return sum(v) if v else None

            # собираем по игроку: name -> {date: row}; даты — общий упорядоченный набор
            athlete_days = {}
            all_dates = []
            for row in rows:
                name = row.get("full_name", "?")
                d = row.get("survey_date")
                athlete_days.setdefault(name, {})[str(d)] = row
                if str(d) not in all_dates:
                    all_dates.append(str(d))
            all_dates.sort()

            # заголовок команды уже есть на r-2 (merge A:L). Дальше легенда + шапка.
            # ЛЕГЕНДА (с цветными ●)
            n_total_cols = 1 + len(all_dates) * len(metrics) + 1  # Игрок | дни*метрики | Ср.
            legend_top = n_total_cols if n_total_cols > 13 else 13
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=legend_top)
            cell(ws, r, 1, "ЛЕГЕНДА   (шкалы: Сон/Утомл/Крепатура 1-7, Боль NRS 0-10, sRPE 1-10, Hooper 3-21)", font=Font(bold=True, size=11, color="1F4E79"), border=None)
            r += 1

            circ_colors = ["1E7B1E", "BF8F00", "C00000"]
            # Строка 3: шкала 1-7
            cell(ws, r, 1, "Шкала 1-7:", font=Font(size=9, bold=True, color="404040"), border=None)
            seq1 = [(circ_colors[0], "6-7 норма"), (circ_colors[1], "4-5 вни.мание"), (circ_colors[2], "<4 критично")]
            sc = 3
            for col_c, lab in seq1:
                cc = cell(ws, r, sc, "●", font=Font(bold=True, size=11, color=col_c), border=None)
                cell(ws, r, sc + 1, lab, font=normal_font, border=None)
                sc += 3
            r += 1
            # Строка: Пульс + Hooper + пропуск
            cell(ws, r, 1, "Пульс:", font=Font(size=9, bold=True, color="404040"), border=None)
            seq2 = [(circ_colors[2], ">70"), (circ_colors[1], "61-70"), (circ_colors[0], "<=60")]
            sc = 3
            for col_c, lab in seq2:
                cell(ws, r, sc, "●", font=Font(bold=True, size=11, color=col_c), border=None)
                cell(ws, r, sc + 1, lab, font=normal_font, border=None)
                sc += 3
            cell(ws, r, 12, "Hooper:", font=Font(size=9, bold=True, color="404040"), border=None)
            seq3 = [(circ_colors[0], ">=17"), (circ_colors[1], "12-16"), (circ_colors[2], "<12")]
            sc = 14
            for col_c, lab in seq3:
                cell(ws, r, sc, "●", font=Font(bold=True, size=11, color=col_c), border=None)
                cell(ws, r, sc + 1, lab, font=normal_font, border=None)
                sc += 3
            cell(ws, r, 23, "Пропуск:", font=Font(size=9, bold=True, color="404040"), border=None)
            cc = cell(ws, r, 25, "", fill=gray_fill)
            cell(ws, r, 26, "не заполнен · «Ср.»=Hooper", font=normal_font, border=None)
            r += 1

            # Двухуровневая шапка: строка дат (SR_DATE), строка метрик (SR_METR)
            SR_DATE = r
            SR_METR = r + 1
            # Игрок — объединён вертикально по двум строкам шапки
            ws.merge_cells(start_row=SR_DATE, start_column=1, end_row=SR_METR, end_column=1)
            cell(ws, SR_DATE, 1, "Игрок", font=hf, fill=hfl, align=hdr_align)
            col = 2
            for di, dt in enumerate(all_dates):
                try:
                    from datetime import datetime as _ddt
                    wd = namedays[_ddt.strptime(dt, "%Y-%m-%d").weekday()]
                    datelabel = dt[5:] + "\n" + wd
                except Exception:
                    datelabel = dt[5:]
                # лёгкая зебра: чётные блоки дней — светло-голубой фон у ячейки даты (и её блока)
                hdr_fill = zebra_hdr_fill if di % 2 == 1 else hfl
                ws.merge_cells(start_row=SR_DATE, start_column=col, end_row=SR_DATE, end_column=col + len(metrics) - 1)
                cell(ws, SR_DATE, col, datelabel, font=hf, fill=hdr_fill, align=hdr_align)
                # жирная граница справа у последней колонки блока дня (визуальный стык дней)
                last_col = col + len(metrics) - 1
                for k, m in enumerate(metrics):
                    bd = day_right if (col + k) == last_col else thin_border
                    cell(ws, SR_METR, col + k, m, font=Font(bold=True, size=8, color="FFFFFF"), fill=hfl, align=hdr_align, border=bd)
                    if (col + k) == last_col:
                        # жирный правый борт и у ячейки даты
                        dc = ws.cell(row=SR_DATE, column=last_col)
                        dc.border = Border(left=thin_border.left, right=day_side, top=thin_border.top, bottom=thin_border.bottom)
                col += len(metrics)
            # колонка «Ср.» объединена по двум строкам
            ws.merge_cells(start_row=SR_DATE, start_column=col, end_row=SR_METR, end_column=col)
            cell(ws, SR_DATE, col, "Ср.", font=hf, fill=hfl, align=hdr_align)
            SRCOL = col
            r = SR_METR + 1

            ws.freeze_panes = f"B{r}"

            if not athlete_days:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SRCOL)
                cell(ws, r, 1, "Нет опросов за данный период.", font=Font(italic=True, color="808080"), border=None)
                r += 1
            else:
                for name in sorted(athlete_days.keys()):
                    daysmap = athlete_days[name]
                    # по умолчанию числовые значения; пустой день = серый
                    col = 2
                    ho_list = []
                    for dt in all_dates:
                        row = daysmap.get(dt)
                        if row is None:
                            for k in range(len(metrics)):
                                cc = cell(ws, r, col + k, "", border=thin_border)
                                cc.fill = gray_fill
                                if k == len(metrics) - 1:
                                    cc.border = day_right
                            col += len(metrics)
                            continue
                        for k, m in enumerate(metrics):
                            if m == "Сон": v = row.get("sleep_score")
                            elif m == "Утомл": v = row.get("fatigue_score")
                            elif m == "Крепатура": v = row.get("muscle_soreness")
                            elif m == "Боль": v = row.get("pain_nrs")
                            elif m == "Пульс": v = row.get("resting_hr")
                            elif m == "sRPE": v = row.get("sRPE_score")
                            elif m == "Hooper": v = _m_hindex(row)
                            v = v if v is not None else ""
                            cc = cell(ws, r, col + k, v, align=center)
                            if isinstance(v, (int, float)):
                                if m == "Пульс":
                                    cc.fill = self._xl_hr_fill(v)
                                elif m == "Боль":
                                    cc.fill = self._xl_pain_fill(v)
                                elif m == "sRPE":
                                    cc.fill = self._xl_srpe_fill(v)
                                elif m == "Hooper":
                                    cc.fill = self._xl_hooper_fill(v)
                                else:
                                    cc.fill = self._xl_score_fill(v)
                                if m == "Hooper":
                                    ho_list.append(v)
                        # жирная граница справа у последней колонки блока дня (стык дней)
                        last_col = col + len(metrics) - 1
                        ws.cell(row=r, column=last_col).border = day_right
                        col += len(metrics)
                    # средний Hooper в колонке «Ср.»
                    if ho_list:
                        av = round(sum(ho_list) / len(ho_list), 1)
                        cc = cell(ws, r, SRCOL, av, align=center, font=Font(bold=True, size=10))
                        cc.fill = self._xl_hooper_fill(av)
                    else:
                        cell(ws, r, SRCOL, "", border=thin_border)
                    # имя игрока слева
                    cell(ws, r, 1, name, font=bold_font)
                    r += 1

            # ширина колонок
            ws.column_dimensions["A"].width = 22
            first_big = min(len(all_dates) * len(metrics) + 2, 60)
            for i in range(2, first_big):
                cl = get_column_letter(i)
                if ws.column_dimensions[cl].width < 5.2:
                    ws.column_dimensions[cl].width = 5.2
            ws.column_dimensions[get_column_letter(SRCOL)].width = 8

            # эмблема клуба в правом верхнем углу листа команды (после последней колонки матрицы)
            self._add_logo_to_ws(ws, anchor=f"{get_column_letter(SRCOL + 2)}{SR_DATE}", size=80)

        # ---- Лист «Легенда»: расшифровка показателей (для тренера) ----
        ws_leg = wb.create_sheet("Легенда")
        ws_leg.column_dimensions["A"].width = 28
        ws_leg.column_dimensions["B"].width = 95
        _lr = 1
        _c = ws_leg.cell(_lr, 1, "ЧТО ЗНАЧИТ КАЖДЫЙ ПОКАЗАТЕЛЬ")
        _c.font = Font(bold=True, size=13, color="1F4E79"); _lr += 1
        _c = ws_leg.cell(_lr, 1, "Расшифровка показателей в отчёте по дням (что это и как читать).")
        _c.font = Font(size=9, italic=True, color="808080"); _lr += 2
        legend_rows = [
            ("ХУПЕР (Hooper)", "Индекс готовности = Сон (1-7) + Утомление (1-7) + Крепатура (1-7). Итог 3-21. Чем выше — тем лучше готов к нагрузке. 17-21 — отличная готовность, 12-16 — средняя, 3-11 — низкая (риск перегрузки/болезни)."),
            ("sRPE", "Субъективная оценка нагрузки на тренировке, 1-10 (RPE — насколько тяжело было). 1-4 — лёгкая, 5-7 — средняя, 8-10 — тяжёлая. Высокий sRPE несколько дней подряд = пора снизить нагрузку."),
            ("ПУЛЬС ПОКОЯ (RHR)", "Ударов в минуту, измеряется утром сразу после пробуждения, лёжа. Чем ниже — тем тренированнее сердце. Рост пульса покоя на 5-10 уд/мин к личной норме = возможная перегрузка или начало болезни. >70 — внимание, >80 — риск."),
            ("HRV (вариабельность)", "Вариабельность сердечного ритма, мс. Отражает баланс нервной системы и уровень восстановления. Чем выше — тем лучше организм восстановился. Источник — умные часы/биоимпеданс (в ежедневном опросе не измеряется)."),
            ("КРЕПАТУРА", "Мышечная болезненность (DOMS) после нагрузки, 1-7. 7 = мышцы не болят, 1 = сильная боль. Низкий балл = мышцы не восстановились."),
            ("БОЛЬ NRS", "Оценка боли по шкале 0-10: 0 = боли нет, 10 = невыносимая. 1-4 — лёгкая, 5 и выше — красный флаг: показать врачу."),
            ("СОН / УТОМЛЕНИЕ", "Шкалы 1-7, где 7 = отличный сон / бодрость, 1 = плохой сон / упадок сил. Входят в расчёт Хупера."),
        ]
        for _col1, _col2 in legend_rows:
            _c = ws_leg.cell(_lr, 1, _col1); _c.font = Font(bold=True, size=10)
            _c = ws_leg.cell(_lr, 2, _col2); _c.font = Font(size=10); _c.alignment = Alignment(wrap_text=True, vertical="top")
            _lr += 1
        _lr += 1
        _c = ws_leg.cell(_lr, 1, "ЦВЕТА"); _c.font = Font(bold=True, size=11); _lr += 1
        for _col, _lab in [("E2EFDA", "Зелёный — норма"), ("FFF2CC", "Жёлтый — внимание"),
                           ("F8D7DA", "Красный — критично / нужен осмотр"), ("F2F2F2", "Серый — нет данных (опрос не пройден)")]:
            _c = ws_leg.cell(_lr, 1, "●")
            _c.fill = PatternFill(start_color=_col, end_color=_col, fill_type="solid")
            _c.font = Font(bold=True, size=10)
            _c = ws_leg.cell(_lr, 2, _lab); _c.font = Font(size=10)
            _lr += 1

        # Убираем move_worksheet - Сводка уже первый лист
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        await q.message.reply_document(
            document=output,
            filename=f"report_{date.today().strftime('%Y%m%d')}.xlsx",
            caption=f"📊 Excel-отчёт за {period_names.get(days, f'{days} дн.')} — детализация по дням"
        )

    # ==================== МЕНЮ СОСТАВА ТЕЛА (СПОРТСМЕН) ====================

    async def body_comp_menu(self, update, ctx):
        """Меню состава тела для спортсмена — последний замер и навигация."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return
        latest = self.db.get_latest_body_composition(athlete["id"])
        if not latest:
            await q.edit_message_text(
                "⚖️ *Состав тела*\n\nУ тебя пока нет данных с весов.\n"
                "Попроси тренера/врача взвесить тебя на биоимпедансных весах.",
                reply_markup=self.kb([[(f"🔙 Назад", "view_today")]]),
                parse_mode="Markdown"
            )
            return
        lines = [
            f"⚖️ *Состав тела*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n",
            f"📅 Последний замер: *{latest.get('record_date')}*\n"
        ]
        metrics = [
            ("Вес", latest.get("weight_kg"), "кг"),
            ("ИМТ", latest.get("bmi"), ""),
            ("Жир", latest.get("body_fat_pct"), "%"),
            ("Мышцы", latest.get("muscle_mass_kg"), "кг"),
            ("Вода", latest.get("body_water_pct"), "%"),
            ("Кости", latest.get("bone_mass_kg"), "кг"),
            ("Висц. жир", latest.get("visceral_fat_index"), ""),
            ("BMR", latest.get("bmr_kcal"), "ккал"),
        ]
        for label, v, unit in metrics:
            if v is not None:
                lines.append(f"• {label}: *{float(v):.1f}* {unit}".rstrip())
        await q.edit_message_text(
            "\n".join(lines),
            reply_markup=self.kb([
                [(f"📜 История (10 замеров)", "body_comp_history")],
                [(f"📈 Графики", "my_charts")],
                [(f"🔙 Назад", "view_today")]
            ]),
            parse_mode="Markdown"
        )

    async def body_comp_history(self, update, ctx):
        """История последних 10 замеров состава тела (для спортсмена)."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return
        history = self.db.get_body_composition(athlete["id"], days=365)
        if not history:
            await q.edit_message_text(
                "📜 *История состава тела*\n\nПока нет данных.",
                reply_markup=self.kb([[(f"🔙 Назад", "body_comp_menu")]]),
                parse_mode="Markdown"
            )
            return
        recent = history[:10]
        lines = [f"📜 *История состава тела* (последние {len(recent)} замеров)\n"]
        for r in recent:
            date_s = r.get("record_date", "?")
            parts = []
            if r.get("weight_kg") is not None:
                parts.append(f"вес {r['weight_kg']:.1f}кг")
            if r.get("body_fat_pct") is not None:
                parts.append(f"жир {r['body_fat_pct']:.1f}%")
            if r.get("muscle_mass_kg") is not None:
                parts.append(f"мышцы {r['muscle_mass_kg']:.1f}кг")
            if r.get("body_water_pct") is not None:
                parts.append(f"вода {r['body_water_pct']:.0f}%")
            if r.get("bone_mass_kg") is not None:
                parts.append(f"кости {r['bone_mass_kg']:.2f}кг")
            lines.append(f"📅 {date_s}: " + (", ".join(parts) if parts else "—"))
        await q.edit_message_text(
            "\n".join(lines),
            reply_markup=self.kb([[(f"🔙 Назад", "body_comp_menu")]]),
            parse_mode="Markdown"
        )

    # ==================== ОТЧЁТ ПО СОСТАВУ ТЕЛА (EXCEL) ====================

    async def report_body_comp_choose_period(self, update, ctx):
        """Выбор команды (для тренера) и периода для Excel-отчёта по составу тела."""
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        is_admin = self._is_full_access(uid)
        state = self.get_state(uid)

        if is_admin:
            # Админ: сразу выбор периода (все команды)
            state["data"]["bc_report_team"] = "all"
            buttons = [
                [("📅 Последний замер", "report_bc_period_1")],
                [("📅 2 недели", "report_bc_period_14")],
                [("📅 Месяц", "report_bc_period_30")],
                [("📅 3 месяца", "report_bc_period_90")],
                [("📅 Всё время", "report_bc_period_999")],
                [("🔙 Назад", "report_export_menu")]
            ]
            await q.edit_message_text(
                "⚖️ *Отчёт по весам — выбор периода*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "Замеры проводятся раз в 2 недели. Выбери период:",
                reply_markup=self.kb(buttons), parse_mode="Markdown"
            )
        else:
            # Тренер: если одна команда — автовыбор, иначе выбор
            teams = self._coach_teams(uid) & set(ACTIVE_TEAMS)
            if len(teams) == 1:
                team = list(teams)[0]
                state["data"]["bc_report_team"] = team
                buttons = [
                    [("📅 Последний замер", "report_bc_period_1")],
                    [("📅 2 недели", "report_bc_period_14")],
                    [("📅 Месяц", "report_bc_period_30")],
                    [("📅 3 месяца", "report_bc_period_90")],
                    [("📅 Всё время", "report_bc_period_999")],
                    [("🔙 Назад", "report_export_menu")]
                ]
                await q.edit_message_text(
                    f"⚖️ *Отчёт по весам — {team}*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    "Замеры проводятся раз в 2 недели. Выбери период:",
                    reply_markup=self.kb(buttons), parse_mode="Markdown"
                )
            else:
                buttons = []
                for t in sorted(teams):
                    buttons.append([(f"🏀 {t}", f"bc_team_{t}")])
                buttons.append([("🔙 Назад", "report_export_menu")])
                await q.edit_message_text(
                    "⚖️ *Отчёт по весам — выбор команды*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    "Выбери команду:",
                    reply_markup=self.kb(buttons), parse_mode="Markdown"
                )

    async def generate_bc_xlsx_report(self, update, ctx):
        """Генерация Excel-отчёта по данным биоимпедансных весов (состав тела)."""
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        days = int(q.data.replace("report_bc_period_", ""))
        state = self.get_state(uid)
        # Если был выбор команды — используем (report_body_comp可以选择所有或 конкретную)
        team = state.get("data", {}).get("bc_report_team", "all")
        if not self._is_full_access(uid):
            if team == "all" or team not in self._coach_teams(uid):
                await q.edit_message_text("🔒 Нет доступа к этой команде.",
                                           reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
                return
        team_filter = None if team == "all" else team

        period_names = {1: "последний замер", 14: "2 недели", 30: "месяц", 90: "3 месяца", 999: "всё время"}

        wb = Workbook()

        # ---- Стили (те же что в generate_xlsx_report) ----
        hf = Font(bold=True, size=11, color="FFFFFF")
        hfl = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=10, color="808080")
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        def cell(ws, row, col, value, font=None, fill=None, align=None, border=thin_border):
            c = ws.cell(row=row, column=col, value=value)
            if font: c.font = font
            if fill is not None: c.fill = fill
            if align: c.alignment = align
            if border: c.border = border
            return c

        # ---- Получаем данные ----
        rows = self.db.get_bc_period(days, team_filter)
        if team_filter is None:
            _active = set(ACTIVE_TEAMS)
            rows = [r for r in rows if r.get("team") in _active]

        if not rows:
            await q.edit_message_text("❌ Нет данных состава тела за выбранный период.",
                                       reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
            return

        # Группируем по командам
        by_team = {}
        for r in rows:
            t = r.get("team", "?")
            by_team.setdefault(t, []).append(r)

        all_teams = sorted(by_team.keys())

        # ---- Лист «Сводка» ----
        ws0 = wb.active
        ws0.title = "Сводка"
        ws0.merge_cells("A1:E1")
        cell(ws0, 1, 1, "⚖️ ЧБК — ОТЧЁТ ПО СОСТАВУ ТЕЛА", font=title_font, border=None)
        ws0.merge_cells("A2:E2")
        cell(ws0, 2, 1,
             f"Период: {period_names.get(days, f'{days} дн.')} | Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
             font=subtitle_font, border=None)

        r = 4
        sum_headers = ["Команда", "Спортсменов", "Ср. вес (кг)", "Ср. жир (%)", "Ср. мышцы (кг)"]
        for ci, h in enumerate(sum_headers, 1):
            cell(ws0, r, ci, h, font=hf, fill=hfl, align=hdr_align)
        r += 1

        for t in all_teams:
            team_rows = by_team[t]
            # Последний замер каждого спортсмена — для средних
            latest_by_athlete = {}
            for row in team_rows:
                aid = row.get("athlete_id")
                if aid not in latest_by_athlete:
                    latest_by_athlete[aid] = row
            n_athletes = len(latest_by_athlete)
            weights = [v["weight_kg"] for v in latest_by_athlete.values()
                       if v.get("weight_kg") is not None]
            fats = [v["body_fat_pct"] for v in latest_by_athlete.values()
                    if v.get("body_fat_pct") is not None]
            muscles = [v["muscle_mass_kg"] for v in latest_by_athlete.values()
                       if v.get("muscle_mass_kg") is not None]
            avg_w = sum(weights) / len(weights) if weights else 0
            avg_f = sum(fats) / len(fats) if fats else 0
            avg_m = sum(muscles) / len(muscles) if muscles else 0

            cell(ws0, r, 1, t, font=bold_font)
            cell(ws0, r, 2, n_athletes, align=center)
            cell(ws0, r, 3, f"{avg_w:.1f}" if weights else "—", align=center)
            cell(ws0, r, 4, f"{avg_f:.1f}" if fats else "—", align=center)
            cell(ws0, r, 5, f"{avg_m:.1f}" if muscles else "—", align=center)
            r += 1

        # Итоговая строка
        total_ath = sum(
            len({row["athlete_id"] for row in by_team[t]
                 if row.get("athlete_id")})
            for t in all_teams
        )
        cell(ws0, r, 1, "ИТОГО", font=Font(bold=True, size=11))
        cell(ws0, r, 2, total_ath, font=bold_font, align=center)
        for ci in range(3, 6):
            cell(ws0, r, ci, "—", align=center)
        r += 1

        self._add_logo_to_ws(ws0, anchor="G3", size=90)

        for ci in range(1, 6):
            ws0.column_dimensions[chr(64 + ci)].width = 18

        # ---- Листы по командам: матрица (игрок × дата, подстолбцы=метрики) ----
        bc_metrics = ["Вес", "% Жира", "Мышцы", "Вода", "BMR"]
        metric_cols = ["weight_kg", "body_fat_pct", "muscle_mass_kg", "body_water_pct", "bmr_kcal"]

        for t in all_teams:
            team_rows = by_team[t]
            ws = wb.create_sheet(title=t[:31])
            ws.merge_cells("A1:L1")
            cell(ws, 1, 1, f"⚖️ Состав тела — {t}", font=title_font, border=None)
            r = 2
            ws.merge_cells(f"A{r}:L{r}")
            cell(ws, r, 1,
                 f"Период: {period_names.get(days, f'{days} дн.')} | "
                 f"Дата отчёта: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                 font=subtitle_font, border=None)
            r = 4

            # Собираем данные: игрок → {дата → row}
            athlete_days = {}
            all_dates = []
            for row in team_rows:
                name = row.get("full_name", "?")
                d = str(row.get("record_date", ""))
                athlete_days.setdefault(name, {})[d] = row
                if d not in all_dates:
                    all_dates.append(d)
            all_dates.sort()

            # Двухуровневая шапка: строка дат (SR_DATE), строка метрик (SR_METR)
            SR_DATE = r
            SR_METR = r + 1
            ws.merge_cells(start_row=SR_DATE, start_column=1, end_row=SR_METR, end_column=1)
            cell(ws, SR_DATE, 1, "Игрок", font=hf, fill=hfl, align=hdr_align)
            col = 2
            for dt in all_dates:
                ws.merge_cells(start_row=SR_DATE, start_column=col,
                               end_row=SR_DATE, end_column=col + len(bc_metrics) - 1)
                cell(ws, SR_DATE, col, dt, font=hf, fill=hfl, align=hdr_align)
                for k, m in enumerate(bc_metrics):
                    cell(ws, SR_METR, col + k, m,
                         font=Font(bold=True, size=8, color="FFFFFF"),
                         fill=hfl, align=hdr_align)
                col += len(bc_metrics)
            ws.merge_cells(start_row=SR_DATE, start_column=col,
                           end_row=SR_METR, end_column=col)
            cell(ws, SR_DATE, col, "Ср.", font=hf, fill=hfl, align=hdr_align)
            SRCOL = col
            r = SR_METR + 1

            ws.freeze_panes = f"B{r}"

            if not athlete_days:
                ws.merge_cells(start_row=r, start_column=1,
                               end_row=r, end_column=SRCOL)
                cell(ws, r, 1, "Нет данных за данный период.",
                     font=Font(italic=True, color="808080"), border=None)
                r += 1
            else:
                for name in sorted(athlete_days.keys()):
                    daysmap = athlete_days[name]
                    col = 2
                    avg_vals = {mc: [] for mc in metric_cols}
                    for dt in all_dates:
                        row_data = daysmap.get(dt)
                        if row_data is None:
                            for k in range(len(bc_metrics)):
                                c = ws.cell(row=r, column=col + k, value="")
                                c.border = thin_border
                                c.fill = gray_fill
                            col += len(bc_metrics)
                            continue
                        for k, mc in enumerate(metric_cols):
                            v = row_data.get(mc)
                            if v is not None:
                                v = float(v)
                                avg_vals[mc].append(v)
                                cell(ws, r, col + k, round(v, 1), align=center)
                            else:
                                cell(ws, r, col + k, "", align=center)
                        col += len(bc_metrics)
                    # Средние в колонке «Ср.»
                    avgs = []
                    for mc in metric_cols:
                        if avg_vals[mc]:
                            avgs.append(f"{sum(avg_vals[mc]) / len(avg_vals[mc]):.1f}")
                        else:
                            avgs.append("—")
                    cell(ws, r, SRCOL, ", ".join(avgs),
                         align=center, font=bold_font)
                    cell(ws, r, 1, name, font=bold_font)
                    r += 1

            # Ширина колонок
            ws.column_dimensions["A"].width = 22
            for ci in range(2, min(SRCOL + 2, 60)):
                cl = get_column_letter(ci)
                if ws.column_dimensions[cl].width < 5.2:
                    ws.column_dimensions[cl].width = 5.2
            ws.column_dimensions[get_column_letter(SRCOL)].width = 22

        # ---- Лист «Легенда» ----
        ws_leg = wb.create_sheet("Легенда")
        ws_leg.column_dimensions["A"].width = 28
        ws_leg.column_dimensions["B"].width = 95
        _lr = 1
        _c = ws_leg.cell(_lr, 1, "ЧТО ЗНАЧИТ КАЖДЫЙ ПОКАЗАТЕЛЬ (СОСТАВ ТЕЛА)")
        _c.font = Font(bold=True, size=13, color="1F4E79"); _lr += 1
        _c = ws_leg.cell(_lr, 1, "Расшифровка показателей биоимпедансного анализа (GARLYN Bodyscan Master).")
        _c.font = Font(size=9, italic=True, color="808080"); _lr += 1
        _c = ws_leg.cell(_lr, 1, "⚠️ Данные скорректированы +3% к % жира (Potter et al., 2021) для приближения к DXA.")
        _c.font = Font(size=9, italic=True, color="CC6600"); _lr += 2
        legend_rows = [
            ("ВЕС (кг)", "Масса тела в килограммах. Контроль динамики (рост/снижение) важен для планирования нагрузки."),
            ("% ЖИРА (body fat %)", "Процент жировой ткани. Норма для юных спортсменов: 10-18% (м), 18-26% (ж). Скорректировано +3% (Potter et al., 2021). GARLYN систематически занижает жир на ~3-4% относительно DXA."),
            ("МЫШЕЧНАЯ МАССА (кг)", "Суммарная масса скелетных мышц. Рост = прогресс силы. Снижение при перегрузке/болезни."),
            ("ВОДА (body water %)", "Общий объём воды в теле. Норма: 55-65% (м), 50-60% (ж). Снижение <55% = обезвоживание (риск травм, снижение восстановления)."),
            ("BMR (ккал)", "Базовый метаболизм — количество калорий, которые организм тратит в покое. Рассчитан из безжировой массы. Для баскетболистов 15-18: ~1600-2200 ккал."),
            ("КОСТИ (bone mass кг)", "Минеральная плотность костей. Стабильна у молодых; резкое снижение — красный флаг. Норма Ca: 1300 мг/сут + витамин D 600-1000 МЕ."),
            ("ВИСЦЕРАЛЬНЫЙ ЖИР", "Жир вокруг органов (индекс 1-12). ≤4 — норма, 5-9 — повышен, ≥10 — высокий риск метаболических нарушений."),
            ("ПОДКОЖНЫЙ ЖИР (%)", "Процент подкожного жира. Коррелирует с общим % жира, но не заменяет его."),
            ("БЕЗЖИРОВАЯ МАССА (кг)", "Fat-free mass (FFM) = вес - жир. Включает мышцы, кости, воду, органы."),
            ("БЕЛОК (%)", "Процент белка в теле. Белок — строительный материал для мышц. Норма: 16-20%."),
            ("AMR (ккал)", "Активный метаболизм = BMR × коэффициент активности. Для баскетболистов: ~2800-3500 ккал."),
        ]
        for _col1, _col2 in legend_rows:
            _c = ws_leg.cell(_lr, 1, _col1); _c.font = Font(bold=True, size=10)
            _c = ws_leg.cell(_lr, 2, _col2)
            _c.font = Font(size=10)
            _c.alignment = Alignment(wrap_text=True, vertical="top")
            _lr += 1

        _lr += 1
        _c = ws_leg.cell(_lr, 1, "ЦВЕТОВОЕ КОДИРОВАНИЕ")
        _c.font = Font(bold=True, size=11, color="1F4E79"); _lr += 1
        color_legend = [
            ("🟢 Зелёный", "Показатель в пределах нормы для возраста и пола"),
            ("🟡 Жёлтый", "Пограничное значение — стоит обратить внимание"),
            ("🔴 Красный", "Отклонение от нормы — требуется консультация врача"),
        ]
        for _col1, _col2 in color_legend:
            _c = ws_leg.cell(_lr, 1, _col1); _c.font = Font(bold=True, size=10)
            _c = ws_leg.cell(_lr, 2, _col2)
            _c.font = Font(size=10)
            _lr += 1

        _lr += 1
        _c = ws_leg.cell(_lr, 1, "ИСТОЧНИКИ")
        _c.font = Font(bold=True, size=11, color="1F4E79"); _lr += 1
        sources = [
            "1. Looney DP et al. (2024) Reliability, biological variability, and accuracy of MF-BIA. Frontiers in Nutrition, 11:1491931",
            "2. Dupertuis YM et al. (2025) BIA instruments: how do they differ. Curr Opin Clin Nutr Metab Care, 28(5):379-387",
            "3. Iblasi R et al. (2025) Pre-season and in-season body composition by BIA in football. Frontiers in Nutrition, 12:1657855",
            "4. Potter AW et al. (2021) +3% body fat correction for BIA vs DXA",
            "5. ESPEN guidelines for BIA measurement standardization (Dupertuis et al., 2025)",
        ]
        for src in sources:
            _c = ws_leg.cell(_lr, 1, src)
            _c.font = Font(size=9, color="666666")
            ws_leg.merge_cells(start_row=_lr, start_column=1, end_row=_lr, end_column=2)
            _lr += 1

        # ---- Сохраняем и отправляем ----
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        await q.message.reply_document(
            document=output,
            filename=f"bc_report_{date.today().strftime('%Y%m%d')}.xlsx",
            caption=f"⚖️ Excel-отчёт по составу тела за {period_names.get(days, f'{days} дн.')}"
        )

    # ==================== ОТЧЁТ ПО УМНЫМ ЧАСАМ (EXCEL) ====================

    async def report_watches_choose_period(self, update, ctx):
        """Выбор периода для отчёта по данным умных часов."""
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        d = q.data
        state = self.get_state(uid)

        if d == "report_watches":
            # Напрямую из меню отчётов — показываем выбор команды
            is_admin = self._is_full_access(uid)
            if is_admin:
                teams = set(ACTIVE_TEAMS) & set(a["team"] for a in self.db.get_all_athletes() if a.get("team"))
            else:
                teams = self._coach_teams(uid) & set(ACTIVE_TEAMS)

            buttons = []
            if is_admin:
                buttons.append([("👥 Все команды", "report_watch_team_all")])
            for t in sorted(teams):
                buttons.append([(f"🏀 {t}", f"report_watch_team_{t}")])
            buttons.append([("🔙 Назад", "report_export_menu")])

            await q.edit_message_text(
                "⌚ *Отчёт по часам — выбор команды*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nВыбери команду:",
                reply_markup=self.kb(buttons), parse_mode="Markdown"
            )
            return

        if d.startswith("report_watch_team_"):
            team = d.replace("report_watch_team_", "")
            if not self._is_full_access(uid):
                if team != "all" and team not in self._coach_teams(uid):
                    await q.edit_message_text("🔒 Нет доступа к этой команде.",
                                               reply_markup=self.kb([[(f"🔙 Назад", "report_export_menu")]]))
                    return
            state["data"]["watch_report_team"] = team

            buttons = [
                [("📅 Последние 7 дней", "report_watch_period_7")],
                [("📅 Последние 30 дней", "report_watch_period_30")],
                [("📅 Последние 90 дней", "report_watch_period_90")],
                [("🔙 Назад", "report_watches")]
            ]
            display_team = "Все команды" if team == "all" else team
            await q.edit_message_text(
                f"⌚ *Отчёт по часам*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"Команда: *{display_team}*\n\nЗа какой период выгрузить данные?",
                reply_markup=self.kb(buttons), parse_mode="Markdown"
            )

    async def generate_watch_xlsx_report(self, update, ctx):
        """Генерация Excel-отчёта по данным умных часов."""
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid):
            return

        days = int(q.data.replace("report_watch_period_", ""))
        state = self.get_state(uid)
        team = state.get("data", {}).get("watch_report_team", "all")
        if not self._is_full_access(uid):
            if team == "all" or team not in self._coach_teams(uid):
                await q.edit_message_text("🔒 Нет доступа к этой команде.",
                                           reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]))
                return
        team_filter = None if team == "all" else team

        period_names = {7: "7 дней", 30: "30 дней", 90: "90 дней"}

        # ---- Получаем данные ----
        watch_rows = self.db.get_watch_data_period(days, team_filter)
        if team_filter is None:
            _active = set(ACTIVE_TEAMS)
            watch_rows = [r for r in watch_rows if r.get("team") in _active]

        if not watch_rows:
            await q.edit_message_text("❌ Нет данных часов за выбранный период.",
                                       reply_markup=self.kb([[(f"🔙 Назад", "report_watches")]]))
            return

        # Группируем по командам
        rows_by_team = {}
        for row in watch_rows:
            t = row.get("team", "?")
            rows_by_team.setdefault(t, []).append(row)

        all_teams = sorted(rows_by_team.keys())

        # ---- Стилы (как в generate_xlsx_report) ----
        wb = Workbook()
        hf = Font(bold=True, size=11, color="FFFFFF")
        hfl = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=10, color="808080")
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def cell(ws, row, col, value, font=None, fill=None, align=None, border=thin_border):
            c = ws.cell(row=row, column=col, value=value)
            if font: c.font = font
            if fill is not None: c.fill = fill
            if align: c.alignment = align
            if border: c.border = border
            return c

        # ---- ЛИСТ СВОДКА ----
        ws0 = wb.active
        ws0.title = "Сводка"
        ws0.merge_cells("A1:F1")
        cell(ws0, 1, 1, "⌚ ЧБК — СВОДНЫЙ ОТЧЁТ ПО ЧАСАМ", font=title_font, border=None)
        ws0.merge_cells("A2:F2")
        cell(ws0, 2, 1, f"Период: {period_names.get(days, f'{days} дн.')} | Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
             font=subtitle_font, border=None)

        r = 4
        sum_headers = ["Команда", "Спортсменов", "Ср. пульс", "Ср. сон (ч)", "Ср. шаги", "Записей"]
        for ci, h in enumerate(sum_headers, 1):
            cell(ws0, r, ci, h, font=hf, fill=hfl, align=hdr_align)
        r += 1

        for t in all_teams:
            rows = rows_by_team.get(t, [])
            n_athletes = len(set(row["athlete_id"] for row in rows))
            hr_vals = [row["resting_hr"] for row in rows if row.get("resting_hr") is not None]
            sleep_vals = [row["sleep_hours"] for row in rows if row.get("sleep_hours") is not None]
            steps_vals = [row["steps"] for row in rows if row.get("steps") is not None]

            avg_hr = round(sum(hr_vals) / len(hr_vals)) if hr_vals else None
            avg_sleep = round(sum(sleep_vals) / len(sleep_vals), 1) if sleep_vals else None
            avg_steps = round(sum(steps_vals) / len(steps_vals)) if steps_vals else None

            cell(ws0, r, 1, t, font=bold_font)
            cell(ws0, r, 2, n_athletes, align=center)
            cell(ws0, r, 3, avg_hr if avg_hr else "—", align=center)
            cell(ws0, r, 4, avg_sleep if avg_sleep else "—", align=center)
            cell(ws0, r, 5, avg_steps if avg_steps else "—", align=center)
            cell(ws0, r, 6, len(rows), align=center)
            r += 1

        # Итого
        cell(ws0, r, 1, "ИТОГО", font=Font(bold=True, size=11))
        cell(ws0, r, 2, sum(len(set(row["athlete_id"] for row in rows_by_team.get(t, []))) for t in all_teams), font=bold_font, align=center)
        cell(ws0, r, 6, len(watch_rows), font=bold_font, align=center)
        for ci in range(3, 6):
            cell(ws0, r, ci, "—", align=center)
        r += 1

        self._add_logo_to_ws(ws0, anchor="H3", size=90)

        for ci in range(1, 7):
            ws0.column_dimensions[chr(64 + ci)].width = 18

        # ---- ЛИСТЫ ПО КОМАНДАМ: матрица Игрок-День-Показатели ----
        from openpyxl.utils import get_column_letter as _gcl

        metrics = ["Пульс", "Сон", "Шаги", "Стресс", "SpO2", "HRV"]
        metric_keys = ["resting_hr", "sleep_hours", "steps", "stress", "spo2", "hrv"]

        for t in all_teams:
            rows = rows_by_team.get(t, [])
            ws = wb.create_sheet(title=t[:31])

            r = 1
            ws.merge_cells("A1:N1")
            cell(ws, r, 1, f"⌚ ЧБК — Данные часов: {t}", font=title_font, border=None)
            r += 1
            ws.merge_cells(f"A{r}:N{r}")
            cell(ws, r, 1,
                 f"Период: {period_names.get(days, f'{days} дн.')} | Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Записей: {len(rows)}",
                 font=subtitle_font, border=None)
            r += 1

            # Собираем по игроку: athlete_id -> {date_str: row}
            athlete_map = {}
            all_dates = []
            for row in rows:
                aid = row["athlete_id"]
                ds = str(row.get("record_date", ""))
                athlete_map.setdefault(aid, {"name": row.get("full_name", "?"), "days": {}})
                athlete_map[aid]["days"][ds] = row
                if ds not in all_dates:
                    all_dates.append(ds)
            all_dates.sort()

            # Двухуровневая шапка: даты (объединённые) + метрики под каждой датой
            SR_DATE = r
            SR_METR = r + 1

            # Игрок — объединён
            ws.merge_cells(start_row=SR_DATE, start_column=1, end_row=SR_METR, end_column=1)
            cell(ws, SR_DATE, 1, "Игрок", font=hf, fill=hfl, align=hdr_align)

            col = 2
            date_start_cols = {}
            for di, dt in enumerate(all_dates):
                try:
                    dt_short = dt[5:10].replace("-", ".")  # MM.DD
                except Exception:
                    dt_short = dt
                date_start_cols[dt] = col
                ws.merge_cells(start_row=SR_DATE, start_column=col,
                               end_row=SR_DATE, end_column=col + len(metrics) - 1)
                cell(ws, SR_DATE, col, dt_short, font=Font(bold=True, size=9, color="FFFFFF"),
                     fill=PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
                     align=hdr_align)
                for mi, mname in enumerate(metrics):
                    cell(ws, SR_METR, col + mi, mname, font=Font(bold=True, size=8, color="FFFFFF"),
                         fill=PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),
                         align=hdr_align)
                col += len(metrics)

            # Средняя по игроку
            ws.merge_cells(start_row=SR_DATE, start_column=col,
                           end_row=SR_DATE, end_column=col + len(metrics) - 1)
            cell(ws, SR_DATE, col, "Ср.", font=Font(bold=True, size=9, color="FFFFFF"),
                 fill=PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                 align=hdr_align)
            for mi, mname in enumerate(metrics):
                cell(ws, SR_METR, col + mi, mname, font=Font(bold=True, size=8, color="FFFFFF"),
                     fill=PatternFill(start_color="8EAADB", end_color="8EAADB", fill_type="solid"),
                     align=hdr_align)
            avg_col = col
            col += len(metrics)

            r = SR_METR + 1

            # Строки игроков
            zebra = False
            for aid in sorted(athlete_map.keys(), key=lambda x: athlete_map[x]["name"]):
                info = athlete_map[aid]
                row_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid") if zebra else None
                cell(ws, r, 1, info["name"], font=Font(bold=True, size=10), fill=row_fill)

                sums = {mk: [] for mk in metric_keys}
                for dt in all_dates:
                    base_col = date_start_cols[dt]
                    drow = info["days"].get(dt)
                    for mi, mk in enumerate(metric_keys):
                        val = drow.get(mk) if drow else None
                        c = ws.cell(row=r, column=base_col + mi, value=val)
                        c.border = thin_border
                        c.alignment = center
                        c.font = normal_font
                        if row_fill:
                            c.fill = row_fill
                        if val is not None:
                            sums[mk].append(float(val))
                            # Цветовая индикация для пульса
                            if mk == "resting_hr":
                                if val > 70:
                                    c.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                elif val > 60:
                                    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                                else:
                                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                            elif mk == "stress" and val is not None:
                                if val > 50:
                                    c.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                elif val > 30:
                                    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                                else:
                                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                            elif mk == "spo2" and val is not None:
                                if val < 95:
                                    c.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                elif val < 97:
                                    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                                else:
                                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                # Средние значения
                for mi, mk in enumerate(metric_keys):
                    avg_val = round(sum(sums[mk]) / len(sums[mk]), 1) if sums[mk] else None
                    c = ws.cell(row=r, column=avg_col + mi, value=avg_val)
                    c.border = thin_border
                    c.alignment = center
                    c.font = Font(bold=True, size=10)
                    if row_fill:
                        c.fill = row_fill

                # Граница правее средних
                cell(ws, r, avg_col + len(metrics), "", border=thin_border, fill=row_fill)

                r += 1
                zebra = not zebra

            # Ширины колонок
            ws.column_dimensions["A"].width = 22
            for ci in range(2, col + 1):
                cl = _gcl(ci)
                if ws.column_dimensions[cl].width < 8:
                    ws.column_dimensions[cl].width = 8

            self._add_logo_to_ws(ws, anchor=f"{_gcl(col + 2)}{SR_DATE}", size=80)

        # ---- Лист «Легенда» ----
        ws_leg = wb.create_sheet("Легенда")
        ws_leg.column_dimensions["A"].width = 28
        ws_leg.column_dimensions["B"].width = 95
        lr = 1
        c = ws_leg.cell(lr, 1, "ПОКАЗАТЕЛИ УМНЫХ ЧАСОВ")
        c.font = Font(bold=True, size=13, color="1F4E79"); lr += 1
        c = ws_leg.cell(lr, 1, "Расшифровка показателей в отчёте по часам.")
        c.font = Font(size=9, italic=True, color="808080"); lr += 2
        legend_rows = [
            ("ПУЛЬС ПОКОЯ (RHR)", "Уд/мин, измеряется утром. Чем ниже — тем тренированнее. Рост на 5-10 к норме = перегрузка/болезнь. >70 — внимание, >80 — риск."),
            ("СОН (ч)", "Продолжительность сна в часах. Норма 7-9 ч. Менее 6 ч — недосып, снижение восстановления."),
            ("ШАГИ", "Количество шагов за день. Норма для спортсмена >8000. Менее 5000 — малоподвижный день."),
            ("СТРЕСС", "Уровень стресса по данным часов (0-100). >50 — высокий, 30-50 — повышенный, <30 — норма."),
            ("SpO2", "Насыщение крови кислородом (%). Норма 97-100. Менее 95 — пониженное, нужен контроль."),
            ("HRV", "Вариабельность сердечного ритма (мс). Чем выше — тем лучше восстановление."),
        ]
        for col1, col2 in legend_rows:
            c = ws_leg.cell(lr, 1, col1); c.font = Font(bold=True, size=10)
            c = ws_leg.cell(lr, 2, col2); c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
            lr += 1
        lr += 1
        c = ws_leg.cell(lr, 1, "ЦВЕТА"); c.font = Font(bold=True, size=11); lr += 1
        for col_val, lab in [("E2EFDA", "Зелёный — норма"), ("FFF2CC", "Жёлтый — внимание"),
                             ("F8D7DA", "Красный — критично")]:
            c = ws_leg.cell(lr, 1, "●")
            c.fill = PatternFill(start_color=col_val, end_color=col_val, fill_type="solid")
            c.font = Font(bold=True, size=10)
            c = ws_leg.cell(lr, 2, lab); c.font = Font(size=10)
            lr += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        await q.message.reply_document(
            document=output,
            filename=f"watch_report_{date.today().strftime('%Y%m%d')}.xlsx",
            caption=f"⌚ Excel-отчёт по часам за {period_names.get(days, f'{days} дн.')}"
        )

    def _xl_score_fill(self, val):
        """Возврат заливки по шкале 1-7 (7=отлично). None для пустых."""
        if val is None:
            return None
        if val >= 6.0:
            return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # зелёный
        if val >= 4.0:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # жёлтый
        return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")      # красный

    def _xl_hr_fill(self, val):
        """Заливка по пульсу покоя: >70 критично, 61-70 внимание, <=60 норма."""
        if val is None:
            return None
        if val > 70:
            return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        if val > 60:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def _xl_srpe_fill(self, val):
        """Заливка по sRPE (1-10): низкая/средняя/высокая нагрузка."""
        if val is None:
            return None
        if val >= 8:
            return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        if val >= 5:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def _xl_pain_fill(self, val):
        """Заливка по боли NRS (0-10): 0=нет боли, 5+=красный флаг."""
        if val is None:
            return None
        if val >= 5:
            return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        if val >= 1:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def _xl_hooper_fill(self, val):
        """Заливка по Hooper (3-21): >=17 отлично, 12-16 средне, <12 низко."""
        if val is None:
            return None
        if val >= 17:
            return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        if val >= 12:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        return PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    async def set_gender_menu(self, update, ctx):
        """Меню выбора пола при регистрации — для девушек включаем цикл."""
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "set_gender"

        await q.edit_message_text(
            "👤 *Выбери пол:*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Для девушек будет добавлен вопрос о дне цикла в ежедневный опрос.",
            reply_markup=self.kb([
                [("♂ Мужской", "gender_male")],
                [("♀ Женский", "gender_female")],
            ]), parse_mode="Markdown"
        )

    async def save_gender(self, update, ctx):
        q = update.callback_query
        await q.answer()
        gender = q.data.replace("gender_", "")
        state = self.get_state(q.from_user.id)
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        self.db.update_athlete_gender(athlete["id"], gender)

        text = f"✅ Пол сохранён: {'♀ Женский' if gender == 'female' else '♂ Мужской'}"
        if gender == "female":
            from config import MENSTRUAL_ENABLED_GROUPS
            if athlete.get("age_group") in MENSTRUAL_ENABLED_GROUPS:
                text += "\n\n📅 Теперь в ежедневном опросе будет вопрос о дне цикла."

        await q.edit_message_text(text, reply_markup=self.kb([[(f"🔙 Главное меню", "main_menu")]]), parse_mode="Markdown")

    async def show_cycle_info(self, update, ctx):
        """Показать информацию о текущей фазе цикла с научной базой."""
        q = update.callback_query
        await q.answer()
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if not athlete:
            return

        # Берём последний день цикла из опросов
        cycle_history = self.db.get_cycle_history(athlete["id"], 30)
        if not cycle_history:
            await q.edit_message_text(
                "📅 *Менструальный цикл*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "Нет данных о цикле. Заполняй день цикла в ежедневном опросе, и здесь появится информация.",
                reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown"
            )
            return

        last = cycle_history[0]
        cd = last.get("cycle_day")
        cl = last.get("cycle_length") or athlete.get("cycle_length_default", CYCLE_LENGTH_MEDIAN)

        if cd is None:
            await q.edit_message_text(
                "📅 Ещё нет данных о дне цикла. Укажи его в опросе.",
                reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]), parse_mode="Markdown"
            )
            return

        # Используем новый модуль cycle_medicine для детальной информации
        info = get_cycle_phase_info(cd, cl)

        text = (
            f"📅 *Менструальный цикл*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👤 {athlete['full_name']}\n"
            f"📋 {athlete['age_group']}\n\n"
        )

        if info:
            text += info["text"]
        else:
            text += "⚠️ Не удалось определить фазу.\n\n"

        text += "\n\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n"

        # Возрастные рекомендации
        age_info = AGE_SPECIFIC.get(athlete.get("age_group", ""))
        if age_info:
            text += f"\n*Для {athlete['age_group']}:*\n{age_info.training_note}\n\n🚩 {age_info.concerns}"

        # История последних дней
        if len(cycle_history) > 1:
            text += "\n*Последние записи:*\n"
            for entry in cycle_history[:7]:
                d = entry.get("survey_date", "?")
                day = entry.get("cycle_day", "—")
                ph = entry.get("cycle_phase", "—")
                text += f"  📅 {d}: день {day} ({ph})\n"

        await q.edit_message_text(
            text,
            reply_markup=self.kb([[(f"🔙 Назад", "main_menu")]]),
            parse_mode="Markdown"
        )

    async def ask_cycle_day(self, update, ctx):
        """Спрашиваем день цикла у девушек после опроса."""
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        if update.effective_user:
            user_id = update.effective_user.id
        elif q:
            user_id = q.from_user.id
        else:
            user_id = None
        if not user_id:
            return
        state = self.get_state(user_id)
        athlete = state.get("data", {}).get("athlete")
        if not athlete:
            return

        state["step"] = "survey_cycle"

        # Берём последний cycle_day и длину цикла
        history = self.db.get_cycle_history(athlete["id"], 60)
        last_day = history[0]["cycle_day"] if history else None
        cl = athlete.get("cycle_length_default", 28)

        # Подсказка по фазе
        from cycle_medicine import get_cycle_phase_info
        next_day = (last_day + 1) if last_day else 1
        if next_day > cl:
            next_day = 1
        phase_info = get_cycle_phase_info(next_day, cl)
        phase_text = f" → предполагается {phase_info['phase'].name_ru.lower()}" if phase_info else ""

        hint = f"\n\n💡 Предп. день: {next_day}{phase_text}" if last_day else ""

        buttons = [
            [(f"🩸 1-5", "cycle_1"), (f"6-10", "cycle_6")],
            [(f"11-15", "cycle_11"), (f"16-20", "cycle_16")],
            [(f"21-25", "cycle_21"), (f"26+", "cycle_26")],
            [(f"✏️ Ввести число", "cycle_custom"), (f"❌ Нет данных", "cycle_none")],
        ]

        text = (
            f"📅 *День цикла* (цикл ~{cl} дн.){hint}\n\n"
            f"Какой сегодня день цикла?\n\n"
            f"*Как посчитать:* день 1 — это первый день месячных. "
            f"Сегодняшний день = сколько дней прошло с их начала.\n"
            f"*Пример:* месячные начались в понедельник, сегодня среда — значит день 3.\n\n"
            f"Если ещё не ведёшь дневник — просто выбери примерный день или нажми «Нет данных»."
        )

        if q:
            await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        elif update.message:
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def _ask_cycle_day_custom(self, update, ctx):
        """Запросить день цикла ручным вводом числа (кнопка «Ввести число»)."""
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        athlete = state.get("data", {}).get("athlete")
        cl = (athlete.get("cycle_length_default") or 28) if athlete else 28
        state["step"] = "survey_cycle"
        await q.edit_message_text(
            f"📅 *День цикла* (цикл ~{cl} дн.)\n\n"
            f"Введи число — какой сегодня день цикла (1-{cl}).\n"
            f"*Подсказка:* день 1 = первый день месячных."
        )

    async def set_cycle_day(self, update, ctx):
        """Сохранить день цикла из кнопок."""
        q = update.callback_query
        await q.answer()
        data = q.data
        user_id = q.from_user.id
        state = self.get_state(user_id)
        athlete = state.get("data", {}).get("athlete")
        if not athlete or state.get("step") != "survey_cycle":
            # Потеря состояния — предлагаем начать заново
            await q.edit_message_text(
                "❌ *Опрос прерван* (возможно бот перезагружался).\n\nХотите начать заново?",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")],
                                      [(f"🏠 Главное меню", "main_menu")]]),
                parse_mode="Markdown"
            )
            return

        from cycle_medicine import get_cycle_phase
        if data == "cycle_none":
            await self._ask_complaints(update, ctx)
            return

        ranges = {"cycle_1": 3, "cycle_6": 8, "cycle_11": 13,
                  "cycle_16": 18, "cycle_21": 23, "cycle_26": 28}
        cycle_day = ranges.get(data, 1)
        cycle_length = athlete.get("cycle_length_default", 28)
        phase_key, phase_info = get_cycle_phase(cycle_day, cycle_length)

        state["data"]["cycle_day"] = cycle_day
        state["data"]["cycle_length"] = cycle_length
        state["data"]["cycle_phase"] = phase_info.name_ru if phase_info else ""
        await self._ask_complaints(update, ctx)

    async def _ask_complaints(self, update, ctx):
        """Спросить про жалобы перед завершением опроса."""
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        if update.effective_user:
            user_id = update.effective_user.id
        elif q:
            user_id = q.from_user.id
        else:
            user_id = None
        if not user_id:
            return
        state = self.get_state(user_id)
        state["step"] = "survey_complaints"

        buttons = [
            [("💬 Ввести жалобу", "complaint_text")],
            [("✅ Нет жалоб", "complaint_none")],
        ]
        text = "💬 *Есть ли жалобы?*\n\nЧто беспокоит? Боли, дискомфорт, самочувствие?"

        if q:
            await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        elif update.message:
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def complaint_text(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "complaint_text_input"
        await q.edit_message_text("📝 Опишите жалобу (отправьте текстовым сообщением):")

    async def complaint_none(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["data"]["complaints"] = ""
        await self._finish_survey(update, q.from_user.id, state)

    async def _ask_cycle_length(self, update, ctx):
        """Спросить длину цикла при первом заполнении."""
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        if update.effective_user:
            user_id = update.effective_user.id
        elif q:
            user_id = q.from_user.id
        else:
            user_id = None
        if not user_id:
            return
        state = self.get_state(user_id)
        state["step"] = "survey_cycle_length"

        buttons = [
            [("21 день", "cycle_len_21"), ("24 дня", "cycle_len_24")],
            [("28 дней (норма)", "cycle_len_28"), ("30 дней", "cycle_len_30")],
            [("35 дней", "cycle_len_35"), ("✏️ Ввести число", "cycle_len_custom")],
            [("Не знаю", "cycle_len_28")],
        ]
        text = (
            "📅 *Длина менструального цикла*\n\n"
            "Менструальный цикл — это дни от первого дня месячных до начала следующих.\n"
            "Обычно он длится 21–35 дней (в среднем 28).\n\n"
            "Помогает отслеживать фазы и лучше понимать самочувствие. "
            "Если ты ещё не ведёшь дневник и не знаешь свою длину — не переживай, выбери 28 (среднее значение), "
            "позже это можно будет скорректировать.\n\n"
            "*Сколько дней обычно длится твой цикл?*"
        )
        if q:
            await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        elif update.message:
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def _cycle_length_chosen(self, update, ctx):
        """Сохранить длину цикла и спросить день цикла."""
        q = update.callback_query
        await q.answer()
        data = q.data
        cycle_length = int(data.replace("cycle_len_", ""))
        user_id = q.from_user.id
        state = self.get_state(user_id)
        athlete = state.get("data", {}).get("athlete")
        if not athlete:
            await q.edit_message_text(
                "❌ *Опрос прерван*. Начните заново.",
                reply_markup=self.kb([[(f"🔄 Пройти опрос", "do_survey")]])
            )
            return
        self.db.update_athlete_gender(athlete["id"], athlete.get("gender", "female"), cycle_length)
        athlete["cycle_length_default"] = cycle_length
        await self.ask_cycle_day(update, ctx)

    async def _ask_cycle_length_custom(self, update, ctx):
        """Запросить длину цикла ручным вводом числа (кнопка «Ввести число»)."""
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "survey_cycle_length"
        await q.edit_message_text(
            "📅 *Длина менструального цикла*\n\n"
            "Введи число — сколько дней длится твой цикл (21-35):"
        )

    async def _apply_cycle_length(self, user_id, length):
        """Сохранить длину цикла (кнопки или ручной ввод) и перейти к дню цикла."""
        state = self.get_state(user_id)
        athlete = state.get("data", {}).get("athlete")
        if not athlete:
            return False
        self.db.update_athlete_gender(athlete["id"], athlete.get("gender", "female"), length)
        athlete["cycle_length_default"] = length
        return True

    # ==================== КОНСУЛЬТАЦИЯ ====================

    async def consultation_start(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "consult_wait_complaints"
        state["data"] = {}

        text = (
            "📅 *Запись на консультацию к врачу*\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Опиши, что тебя беспокоит:"
        )
        buttons = [
            [("📝 Описать жалобы", "consult_text")],
            [("🚫 Без жалоб", "consult_no")],
            [("🏠 Отмена", "main_menu")],
        ]
        await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def consultation_text(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["step"] = "consult_text_input"
        await q.edit_message_text("📝 Напиши свои жалобы (что беспокоит):")

    async def consultation_no_complaints(self, update, ctx):
        q = update.callback_query
        await q.answer()
        state = self.get_state(q.from_user.id)
        state["data"]["consult_complaints"] = "Нет жалоб"
        await self._show_calendar(update, ctx, "consult")

    async def _show_calendar(self, update, ctx, prefix="survey"):
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        today = date.today()
        buttons = []
        row = []
        for i in range(1, 15):
            d = today + timedelta(days=i)
            row.append((d.strftime("%d.%m"), f"{prefix}_{d.isoformat()}"))
            if len(row) == 4:
                buttons.append(row)
                row = []
        if row:
            buttons.append(row)
        buttons.append([("🏠 Отмена", "main_menu")])

        text = "📅 Выберите дату консультации:"
        if q:
            await q.edit_message_text(text, reply_markup=self.kb(buttons))
        else:
            await update.message.reply_text(text, reply_markup=self.kb(buttons))

    async def consultation_date(self, update, ctx):
        q = update.callback_query
        await q.answer()
        d = q.data
        parts = d.split("_", 1)
        if len(parts) == 2:
            date_str = parts[1]
            state = self.get_state(q.from_user.id)
            state["data"]["consult_date"] = date_str
            await self._save_consultation(update, ctx)
        else:
            await q.edit_message_text("❌ Ошибка даты. Попробуйте снова.")

    async def _save_consultation(self, update, ctx):
        q = update.callback_query
        user = update.effective_user
        state = self.get_state(user.id)
        athlete = self.db.get_athlete_by_telegram_id(user.id)
        if not athlete:
            await q.edit_message_text("❌ Ошибка. Начните с /start")
            return

        complaints = state.get("data", {}).get("consult_complaints", "Не указано")
        consult_date = state.get("data", {}).get("consult_date", "Не определена")

        # Уведомление врачу
        admin_text = (
            f"📅 *Новая запись на консультацию*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👤 *{athlete['full_name']}*\n"
            f"🏀 Команда: *{athlete.get('team', '?')}*\n"
            f"📊 Возраст: *{athlete.get('age_group', '?')}*\n\n"
            f"📝 Жалобы: {complaints}\n"
            f"📅 Дата: {consult_date}\n\n"
            f"🆔 ID: [{user.id}](tg://user?id={user.id})"
        )

        for admin_id in self._full_access_ids():
            btn_name = self._first_name(athlete)
            url_btn = InlineKeyboardButton(f"💬 Написать {btn_name}", url=f"tg://user?id={user.id}")
            markup = InlineKeyboardMarkup([[url_btn]])
            try:
                await ctx.bot.send_message(
                    admin_id, admin_text,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
            except Exception as e:
                logger.error(f"Send consult admin error: {e}")

        text = (
            f"✅ *Запись на консультацию создана!*\n\n"
            f"📝 Жалобы: {complaints}\n"
            f"📅 Дата: {consult_date}\n\n"
            f"Врач свяжется с тобой! 💪"
        )
        self.clear_state(user.id)
        try:
            await q.edit_message_text(text, reply_markup=self.kb([[(f"🏠 Главное меню", "main_menu")]]), parse_mode="Markdown")
        except Exception:
            pass

    async def show_help(self, update, ctx):
        text = (
            f"❓ *Помощь*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📝 Опрос — 1-2 минуты ежедневно\n"
            f"👤 Профиль — статистика и тренды\n"
            f"📈 Графики — sparkline трендов\n"
            f"📋 Анкета — анкетирование спортсмена\n"
            f"🏆 Достижения — награды за серию\n"
            f"⌚ Данные часов — импорт CSV/JSON\n"
            f"🔄 Заново — сброс и перерегистрация\n\n"
            f"*/start* — Главное меню\n"
        )
        buttons = [[(f"🏠 Главное меню", "main_menu")]]
        if update.callback_query:
            await update.callback_query.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        else:
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    # ==================== АНКЕТИРОВАНИЕ ====================

    async def start_questionnaire(self, update, ctx):
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if not athlete:
            await q.edit_message_text("❌ Сначала зарегистрируйся!", reply_markup=self.kb([[(f"📝 Регистрация", "start_reg")]]))
            return
        if self.db.has_questionnaire(athlete["id"]) and not self.db.has_incomplete_questionnaire(athlete["id"]):
            await q.edit_message_text(
                "📋 *Анкета уже заполнена.*\n\nОбновить её сейчас? Данные обновятся для врача.",
                parse_mode="Markdown",
                reply_markup=self.kb([
                    [(f"✅ Обновить анкету", "q_restart")],
                    [(f"🏠 Главное меню", "main_menu")]
                ])
            )
            return

        # Если есть незавершённая анкета — предложить продолжить с того места
        incomplete = self.db.has_incomplete_questionnaire(athlete["id"])
        state = self.get_state(user_id)

        if incomplete and state.get("q_data"):
            # продолжаем
            last_step = state.get("step", "q_age")
            await self._q_resume_questionnaire(update, ctx, state)
            return

        # начинаем или продолжаем с сохранённого прогресса
        progress = self.db.get_questionnaire_progress(athlete["id"])
        if progress:
            # переносим сохранённые данные в q_data
            saved = {k: v for k, v in progress.items() if k not in ("id", "athlete_id", "completed_at")}
            state["q_data"] = saved
            # определяем следующий шаг
            await self._q_resume_questionnaire(update, ctx, state)
            return

        state["step"] = "q_age"
        state["q_data"] = {}
        await q.edit_message_text("📋 *Анкета баскетболиста*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nБлок 1: Общие данные\n\nСколько тебе полных лет?", parse_mode="Markdown")

    async def questionnaire_restart(self, update, ctx):
        """Обновление анкеты: очистить старые ответы и начать заново."""
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if not athlete:
            return
        # Сбрасываем прогресс в БД (старая анкета перезапишется при сохранении)
        try:
            self.db.conn.execute("DELETE FROM questionnaires WHERE athlete_id = ?", (athlete["id"],))
            self.db.conn.commit()
        except Exception as e:
            logger.warning(f"q_restart delete: {e}")
        self.clear_state(user_id)
        state = self.get_state(user_id)
        state["step"] = "q_age"
        state["q_data"] = {}
        await q.edit_message_text(
            "📋 *Обновление анкеты*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\nБлок 1: Общие данные\n\nСколько тебе полных лет?",
            parse_mode="Markdown"
        )

    async def _q_resume_questionnaire(self, update, ctx, state):
        """Продолжить анкету с того шага, где остановился."""
        q = update.callback_query
        data = state.get("q_data", {})
        # Определяем следующий незаполненный шаг
        step_order = [
            ("q_age", "age"), ("q_phone", "phone"), ("q_birth_date", "birth_date"), ("q_gender", "gender"), ("q_position", "position"),
            ("q_level", "level"), ("q_experience", "experience"), ("q_height_weight", "height"),
            ("q_trauma_12m", "trauma_12m"), ("q_zones", "zones"),
            ("q_pain_now", "pain_now"), ("q_chronic", "chronic"),
            ("q_surgery", "surgery"), ("q_meds", "meds"), ("q_allergies", "allergies"),
            ("q_train_count", "train_count"), ("q_train_duration", "train_duration"),
            ("q_season", "season"), ("q_form", "form_score"),
            ("q_sleep_score", "sleep_score"), ("q_warmup", "warmup"),
            ("q_recovery", "recovery"), ("q_water", "water"), ("q_diet", "diet"),
            ("q_pre_meal", "pre_meal"), ("q_supplements", "supplements"),
            ("q_motivation", "motivation"), ("q_stress", "stress"),
            ("q_match_state", "match_state"), ("q_reinjury", "reinjury_fear"),
        ]
        for step, key in step_order:
            if key not in data or data[key] in (None, ""):
                state["step"] = step
                await self._q_ask_step(update, ctx, step)
                return
        # Всё заполнено (кроме текстовых goal/wish) — спрашиваем цель
        if not data.get("goal"):
            state["step"] = "q_goal"
            await q.edit_message_text("📋 *Блок 6: Цели*\n\nКакую главную цель ставишь? (напиши текстом)", parse_mode="Markdown")
            return
        if not data.get("wish"):
            state["step"] = "q_wish"
            await q.edit_message_text("📋 *Блок 6: Цели*\n\nЕсть что-то, с чем нужна помощь?\n(можно пропустить — напиши «-»)", parse_mode="Markdown")
            return
        # завершено
        state["step"] = "q_wish"
        await q.edit_message_text("📋 *Блок 6: Цели*\n\nЕсть что-то, с чем нужна помощь?\n(можно пропустить — напиши «-»)", parse_mode="Markdown")

    async def _q_show_supp_buttons(self, update, ctx, selected):
        """Мультивыбор спортпита: кнопки тоглят выбранное (✅/⬜), Другое + Готово."""
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        supp_options = ["Протеин", "Креатин", "Хондропротекторы", "Изотоники", "Кофеин", "Ничего"]
        # приведём selected к списку
        if isinstance(selected, str):
            sel_list = [selected] if selected and selected != "Ничего" else []
        elif isinstance(selected, list):
            sel_list = selected
        else:
            sel_list = []

        def mark(opt):
            return "✅ " + opt if opt in sel_list else "⬜ " + opt

        buttons = [
            [(f"{mark('Протеин')}", "q_supp_toggle_Протеин"), (f"{mark('Креатин')}", "q_supp_toggle_Креатин")],
            [(f"{mark('Хондропротекторы')}", "q_supp_toggle_Хондропротекторы"), (f"{mark('Изотоники')}", "q_supp_toggle_Изотоники")],
            [(f"{mark('Кофеин')}", "q_supp_toggle_Кофеин"), (f"{mark('Ничего')}", "q_supp_toggle_Ничего")],
            [("✏️ Другое (введу сам)", "q_supp_other")],
            [("✅ Готово", "q_supp_done")],
        ]
        text = (
            "📋 *Блок 4: Питание*\n\nСпортпит / добавки?\n"
            "Можно выбрать несколько. Нажми «Готово», когда закончишь."
        )
        if sel_list:
            text += f"\n\nВыбрано: {', '.join(sel_list)}"
        if q:
            await q.edit_message_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")
        else:
            await update.message.reply_text(text, reply_markup=self.kb(buttons), parse_mode="Markdown")

    async def _q_ask_step(self, update, ctx, step):
        """Показать вопрос для конкретного шага анкеты."""
        q = update.callback_query
        questions = {
            "q_age": ("📋 *Блок 1: Общие данные*\n\nСколько тебе полных лет?", None),
            "q_phone": ("📋 *Блок 1: Общие данные*\n\nВведи номер телефона:", None),
            "q_birth_date": ("📋 *Блок 1: Общие данные*\n\nВведи дату рождения (ДД.ММ.ГГГГ):", None),
            "q_gender": ("📋 *Блок 1*\n\nВаш пол?", [[("М", "q_gender_М"), ("Ж", "q_gender_Ж")]]),
            "q_position": ("📋 *Блок 1*\n\nИгровая позиция?", [[("PG","q_position_PG"),("SG","q_position_SG"),("SF","q_position_SF")],[("PF","q_position_PF"),("C","q_position_C"),("Универсал","q_position_Универсал")]]),
            "q_level": ("📋 *Блок 1*\n\nУровень игры?", [[("Любитель","q_level_Любитель")],[("Полупрофессионал","q_level_Полупрофессионал")],[("Профессионал","q_level_Профессионал")],[("Ветеран","q_level_Ветеран")]]),
            "q_experience": ("📋 *Блок 1*\n\nКак давно играешь?", [[("менее 1 года","q_exp_менее 1 года")],[("1–3 года","q_exp_1–3 года")],[("3–7 лет","q_exp_3–7 лет")],[("7–15 лет","q_exp_7–15 лет")],[("более 15 лет","q_exp_более 15 лет")]]),
            "q_trauma_12m": ("📋 *Блок 2: Травмы*\n\nБыли травмы за 12 мес?", [[("Да","q_trauma_Да"),("Нет","q_trauma_Нет")]]),
            "q_zones": ("📋 *Блок 2*\n\nЗоны, которые беспокоили:", [[("Голеностопы","q_zones_Голеностопы"),("Колени","q_zones_Колени")],[("Бёдра","q_zones_Бёдра"),("Поясница","q_zones_Поясница")],[("Кисти","q_zones_Кисти"),("Плечи","q_zones_Плечи")],[("Шея","q_zones_Шея"),("Ничего","q_zones_Ничего")]]),
            "q_pain_now": ("📋 *Блок 2*\n\nЕсть боль сейчас?", [[("Да","q_pain_Да"),("Нет","q_pain_Нет")]]),
            "q_chronic": ("📋 *Блок 2*\n\nРецидивирующая травма?", [[("Да","q_chronic_Да"),("Нет","q_chronic_Нет")]]),
            "q_surgery": ("📋 *Блок 2*\n\nБыли операции из-за спорта?", [[("Да","q_surgery_Да"),("Нет","q_surgery_Нет")]]),
            "q_meds": ("📋 *Блок 2*\n\nПринимаешь лекарства/БАДы?", [[("Да","q_meds_Да"),("Нет","q_meds_Нет")]]),
            "q_allergies": ("📋 *Блок 2*\n\nЕсть аллергии?", [[("Да","q_allergies_Да"),("Нет","q_allergies_Нет")]]),
            "q_train_count": ("📋 *Блок 3: Режим*\n\nСколько тренировок в неделю?", [[(str(i),f"q_tcount_{i}") for i in range(0,11)]]),
            "q_train_duration": ("📋 *Блок 3*\n\nДлительность тренировки?", [[(str(i),f"q_tdur_{i}") for i in [30,45,60,90,120,180]]]),
            "q_season": ("📋 *Блок 3*\n\nСезон или межсезонье?", [[("Игровой сезон","q_season_Игровой сезон")],[("Предсезонка","q_season_Предсезонка")],[("Межсезонье","q_season_Межсезонье")],[("Пауза/отпуск","q_season_Пауза/отпуск")]]),
            "q_form": ("📋 *Блок 3*\n\nФизическая форма (1-10)?", [[(str(i),f"q_form_{i}") for i in range(1,6)],[(str(i),f"q_form_{i}") for i in range(6,11)]]),
            "q_sleep_score": ("📋 *Блок 3*\n\nКачество сна (1-10)?", [[(str(i),f"q_sleep_{i}") for i in range(1,6)],[(str(i),f"q_sleep_{i}") for i in range(6,11)]]),
            "q_warmup": ("📋 *Блок 3*\n\nРазминка/заминка?", [[("Да, всегда","q_warm_Да, всегда"),("Только разминку","q_warm_Только разминку")],[("Только заминку","q_warm_Только заминку"),("Нет","q_warm_Нет")],[("Не регулярно","q_warm_Не регулярно")]]),
            "q_recovery": ("📋 *Блок 3*\n\nВосстановительные дни?", [[("Да, регулярно","q_rec_Да, регулярно"),("Иногда","q_rec_Иногда")],[("Нет","q_rec_Нет"),("Не знаю","q_rec_Не знаю")]]),
            "q_water": ("📋 *Блок 4: Питание*\n\nСколько воды (л) в день?", [[("1л","q_water_1"),("1.5л","q_water_1.5")],[("2л","q_water_2"),("2.5л","q_water_2.5")],[("3л","q_water_3"),("3.5л+","q_water_3.5")]]),
            "q_diet": ("📋 *Блок 4*\n\nПлан питания?", [[("Да, с диетологом","q_diet_Да, с диетологом"),("Да, сам","q_diet_Да, сам")],[("Частично","q_diet_Частично"),("Нет","q_diet_Нет")],[("Не задумывался","q_diet_Не задумывался")]]),
            "q_pre_meal": ("📋 *Блок 4*\n\nЗа сколько ешь перед тренировкой?", [[("За 1-2ч","q_meal_За 1-2ч"),("За 3-4ч","q_meal_За 3-4ч")],[("Менее чем за час","q_meal_Менее чем за час"),("Натощак","q_meal_Натощак")]]),
            "q_supplements": ("📋 *Блок 4*\n\nСпортпит?", [[("Протеин","q_supp_Протеин"),("Креатин","q_supp_Креатин")],[("Хондропротекторы","q_supp_Хондропротекторы"),("Изотоники","q_supp_Изотоники")],[("Кофеин","q_supp_Кофеин"),("Ничего","q_supp_Ничего")],[("Другое","q_supp_Другое")]]),
            "q_motivation": ("📋 *Блок 5: Психология*\n\nМотивация (1-10)?", [[(str(i),f"q_motiv_{i}") for i in range(1,6)],[(str(i),f"q_motiv_{i}") for i in range(6,11)]]),
            "q_stress": ("📋 *Блок 5*\n\nВнешние факторы мешают?", [[("Да, сильно","q_stress_Да, сильно"),("Периодически","q_stress_Периодически")],[("Нет","q_stress_Нет"),("Не влияют","q_stress_Не влияют")]]),
            "q_match_state": ("📋 *Блок 5*\n\nСостояние перед матчами?", [[("Спокойная уверенность","q_match_Спокойная уверенность")],[("Лёгкое волнение","q_match_Лёгкое волнение")],[("Сильная тревога","q_match_Сильная тревога")],[("По-разному","q_match_По-разному")]]),
            "q_reinjury": ("📋 *Блок 5*\n\nБоишься повторной травмы?", [[("Да, постоянно","q_reinj_Да, постоянно"),("Иногда","q_reinj_Иногда")],[("Нет","q_reinj_Нет"),("Не было травм","q_reinj_Не было травм")]]),
        }
        if step in questions:
            if step == "q_supplements":
                # мультивыбор спортпита
                state = self.get_state(q.from_user.id)
                cur = state.get("q_data", {}).get("supplements", [])
                await self._q_show_supp_buttons(update, ctx, cur)
                return
            text, buttons = questions[step]
            kb = self.kb(buttons) if buttons else None
            await q.edit_message_text(text, reply_markup=kb, parse_mode="Markdown")
        else:
            state = self.get_state(q.from_user.id)
            state["step"] = step
            # текстовые шаги
            text_map = {
                "q_trauma_12m_detail": "📋 *Блок 2*\n\nКакие травмы и как лечил?",
                "q_pain_now_detail": "📋 *Блок 2*\n\nГде и как давно?",
                "q_chronic_detail": "📋 *Блок 2*\n\nКакая? Как часто обостряется?",
                "q_surgery_detail": "📋 *Блок 2*\n\nКакие операции?",
                "q_surgery_date": "📋 *Блок 2*\n\nКогда была операция (год или дата)?",
                "q_meds_detail": "📋 *Блок 2*\n\nКакие именно?",
                "q_allergies_detail": "📋 *Блок 2*\n\nНа что аллергия?",
            }
            await q.edit_message_text(text_map.get(step, "Продолжим"), parse_mode="Markdown")

    async def handle_questionnaire_answer(self, update, ctx):
        q = update.callback_query
        await q.answer()
        user_id = q.from_user.id
        state = self.get_state(user_id)
        step = state.get("step", "")
        data = state.get("q_data", {})
        athlete = self.db.get_athlete_by_telegram_id(user_id)
        if not athlete: return

        if step == "q_age":
            data["age"] = q.data.replace("q_age_", "")
            state["step"] = "q_gender"
            await q.edit_message_text("📋 *Блок 1: Общие данные*\n\nВаш пол?", reply_markup=self.kb([[("М", "q_gender_М"), ("Ж", "q_gender_Ж")]]), parse_mode="Markdown")
            return

        if step == "q_gender":
            data["gender"] = q.data.replace("q_gender_", "")
            state["step"] = "q_position"
            await q.edit_message_text("📋 *Блок 1: Общие данные*\n\nВаша основная игровая позиция?", reply_markup=self.kb([
                [("PG", "q_position_PG"), ("SG", "q_position_SG"), ("SF", "q_position_SF")],
                [("PF", "q_position_PF"), ("C", "q_position_C"), ("Универсал", "q_position_Универсал")],
            ]), parse_mode="Markdown")
            return

        if step == "q_position":
            data["position"] = q.data.replace("q_position_", "")
            state["step"] = "q_level"
            await q.edit_message_text("📋 *Блок 1: Общие данные*\n\nУровень игры?", reply_markup=self.kb([
                [("Любитель", "q_level_Любитель")],
                [("Полупрофессионал", "q_level_Полупрофессионал")],
                [("Профессионал", "q_level_Профессионал")],
                [("Ветеран", "q_level_Ветеран")],
            ]), parse_mode="Markdown")
            return

        if step == "q_level":
            data["level"] = q.data.replace("q_level_", "")
            state["step"] = "q_experience"
            await q.edit_message_text("📋 *Блок 1: Общие данные*\n\nКак давно играешь в баскетбол?", reply_markup=self.kb([
                [("менее 1 года", "q_exp_менее 1 года")],
                [("1–3 года", "q_exp_1–3 года")],
                [("3–7 лет", "q_exp_3–7 лет")],
                [("7–15 лет", "q_exp_7–15 лет")],
                [("более 15 лет", "q_exp_более 15 лет")],
            ]), parse_mode="Markdown")
            return

        if step == "q_experience":
            data["experience"] = q.data.replace("q_exp_", "")
            state["step"] = "q_height_weight"
            await q.edit_message_text(
                "📋 *Блок 1: Общие данные*\n\n"
                "📏 *Введи рост и вес через пробел:*\n"
                "Например: `185 82`\n\n"
                "Если не знаешь — напиши «-»",
                parse_mode="Markdown"
            )
            return

        if step == "q_trauma_12m":
            data["trauma_12m"] = q.data.replace("q_trauma_", "")
            if data["trauma_12m"] == "Да":
                state["step"] = "q_trauma_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nКакие травмы и как лечил?", parse_mode="Markdown")
            else:
                data["trauma_12m_detail"] = ""
                state["step"] = "q_zones"
                await self._q_show_zones(update, ctx)
            return

        if step == "q_zones":
            data["zones"] = q.data.replace("q_zones_", "")
            state["step"] = "q_pain_now"
            await q.edit_message_text("📋 *Блок 2: Травмы*\n\nЕсть боль или дискомфорт прямо сейчас?", reply_markup=self.kb([[("Да", "q_pain_Да"), ("Нет", "q_pain_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_pain_now":
            data["pain_now"] = q.data.replace("q_pain_", "")
            if data["pain_now"] == "Да":
                state["step"] = "q_pain_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nГде и как давно?", parse_mode="Markdown")
            else:
                data["pain_now_detail"] = ""
                state["step"] = "q_chronic"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nЕсть рецидивирующая (повторяющаяся) травма?", reply_markup=self.kb([[("Да", "q_chronic_Да"), ("Нет", "q_chronic_Нет")]]), parse_mode="Markdown")
            return

        if step == "q_chronic":
            data["chronic"] = q.data.replace("q_chronic_", "")
            if data["chronic"] == "Да":
                state["step"] = "q_chronic_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nКакая? Как часто обостряется?", parse_mode="Markdown")
            else:
                data["chronic_detail"] = ""
                await self._q_continue_block2(update, ctx)
            return

        if step == "q_surgery":
            data["surgery"] = q.data.replace("q_surgery_", "")
            if data["surgery"] == "Да":
                state["step"] = "q_surgery_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nКакие были операции?", parse_mode="Markdown")
            else:
                await self._q_continue_block2_meds(update, ctx)
            return

        if step == "q_meds":
            data["meds"] = q.data.replace("q_meds_", "")
            if data["meds"] == "Да":
                state["step"] = "q_meds_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nКакие именно?", parse_mode="Markdown")
            else:
                data["meds_detail"] = ""
                state["step"] = "q_allergies"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nЕсть аллергии?", reply_markup=self.kb([[("Да", "q_allergies_Да"), ("Нет", "q_allergies_Нет")]]), parse_mode="Markdown")

        if step == "q_allergies":
            data["allergies"] = q.data.replace("q_allergies_", "")
            if data["allergies"] == "Да":
                state["step"] = "q_allergies_detail"
                await q.edit_message_text("📋 *Блок 2: Травмы*\n\nНа что аллергия? (что вызывает, как проявляется)", parse_mode="Markdown")
            else:
                data["allergies_detail"] = ""
                await self._q_save_then_block3(update, ctx)
            return
    
        # Блок 3: тренировочный режим
        if step == "q_train_count":
            state["q_data"]["train_count"] = int(q.data.replace("q_tcount_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_train_duration"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nСколько минут длится тренировка/игра?", reply_markup=self.kb([[(str(i), f"q_tdur_{i}") for i in [30,45,60,90,120,180]]]), parse_mode="Markdown")
            return

        if step == "q_train_duration":
            state["q_data"]["train_duration"] = int(q.data.replace("q_tdur_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_season"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nСейчас сезон или межсезонье?", reply_markup=self.kb([[("Игровой сезон", "q_season_Игровой сезон")], [("Предсезонка", "q_season_Предсезонка")], [("Межсезонье", "q_season_Межсезонье")], [("Пауза/отпуск", "q_season_Пауза/отпуск")]]), parse_mode="Markdown")
            return

        if step == "q_season":
            state["q_data"]["season"] = q.data.replace("q_season_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_form"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nОцени физическую форму (1-10):", reply_markup=self.kb([[(str(i), f"q_form_{i}") for i in range(1,6)],[(str(i), f"q_form_{i}") for i in range(6,11)]]), parse_mode="Markdown")
            return

        if step == "q_form":
            state["q_data"]["form_score"] = int(q.data.replace("q_form_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_sleep_score"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nОцени качество сна (1-10):", reply_markup=self.kb([[(str(i), f"q_sleep_{i}") for i in range(1,6)],[(str(i), f"q_sleep_{i}") for i in range(6,11)]]), parse_mode="Markdown")
            return

        if step == "q_sleep_score":
            state["q_data"]["sleep_score"] = int(q.data.replace("q_sleep_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_warmup"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nДелаешь разминку/заминку?", reply_markup=self.kb([[("Да, всегда", "q_warm_Да, всегда"), ("Только разминку", "q_warm_Только разминку")], [("Только заминку", "q_warm_Только заминку"), ("Нет", "q_warm_Нет")], [("Не регулярно", "q_warm_Не регулярно")]]), parse_mode="Markdown")
            return

        if step == "q_warmup":
            state["q_data"]["warmup"] = q.data.replace("q_warm_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_recovery"
            await q.edit_message_text("📋 *Блок 3: Режим*\n\nЕсть восстановительные дни?", reply_markup=self.kb([[("Да, регулярно", "q_rec_Да, регулярно"), ("Иногда", "q_rec_Иногда")], [("Нет", "q_rec_Нет"), ("Не знаю", "q_rec_Не знаю")]]), parse_mode="Markdown")
            return

        if step == "q_recovery":
            state["q_data"]["recovery"] = q.data.replace("q_rec_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_water"
            await q.edit_message_text("📋 *Блок 4: Питание*\n\nСколько воды (л) в день?", reply_markup=self.kb([[("1л", "q_water_1"), ("1.5л", "q_water_1.5")], [("2л", "q_water_2"), ("2.5л", "q_water_2.5")], [("3л", "q_water_3"), ("3.5л+", "q_water_3.5")]]), parse_mode="Markdown")
            return

        if step == "q_water":
            state["q_data"]["water"] = float(q.data.replace("q_water_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_diet"
            await q.edit_message_text("📋 *Блок 4: Питание*\n\nПлан питания?", reply_markup=self.kb([[("Да, с диетологом", "q_diet_Да, с диетологом"), ("Да, сам", "q_diet_Да, сам")], [("Частично", "q_diet_Частично"), ("Нет", "q_diet_Нет")], [("Не задумывался", "q_diet_Не задумывался")]]), parse_mode="Markdown")
            return

        if step == "q_diet":
            state["q_data"]["diet"] = q.data.replace("q_diet_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_pre_meal"
            await q.edit_message_text("📋 *Блок 4: Питание*\n\nЗа сколько обычно ешь перед тренировкой или игрой?", reply_markup=self.kb([[("За 1-2ч", "q_meal_За 1-2ч"), ("За 3-4ч", "q_meal_За 3-4ч")], [("Менее чем за час", "q_meal_Менее чем за час"), ("Натощак", "q_meal_Натощак")]]), parse_mode="Markdown")
            return

        if step == "q_pre_meal":
            state["q_data"]["pre_meal"] = q.data.replace("q_meal_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_supplements"
            cur = state.get("q_data", {}).get("supplements", [])
            await self._q_show_supp_buttons(update, ctx, cur)
            return

        if step == "q_supplements":
            supp_options = ["Протеин", "Креатин", "Хондропротекторы", "Изотоники", "Кофеин", "Ничего"]

            # Кнопка "Готово" -> сохранить и перейти к блоку 5
            if q.data == "q_supp_done":
                if not state["q_data"].get("supplements"):
                    state["q_data"]["supplements"] = "Ничего"
                # автoсохранение
                try:
                    await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
                except Exception as e:
                    logger.error(f"Q autosave: {e}")
                state["step"] = "q_motivation"
                await q.edit_message_text("📋 *Блок 5: Психология*\n\nМотивация (1-10):", reply_markup=self.kb([[(str(i), f"q_motiv_{i}") for i in range(1,6)],[(str(i), f"q_motiv_{i}") for i in range(6,11)]]), parse_mode="Markdown")
                return

            # Кнопка "Другое" -> ввести свой вариант
            if q.data == "q_supp_other":
                state["step"] = "q_supp_other_input"
                await q.edit_message_text("📋 *Блок 4: Питание*\n\nНапиши свой вариант спортпита:", parse_mode="Markdown")
                return

            # Тоггл выбранного варианта (q_supp_toggle_<name>)
            if q.data.startswith("q_supp_toggle_"):
                option = q.data.replace("q_supp_toggle_", "")
                cur = state["q_data"].get("supplements", [])
                if isinstance(cur, str):
                    cur = [cur] if cur and cur != "Ничего" else []
                cur = list(cur) if isinstance(cur, list) else []
                if option == "Ничего":
                    # если выбрали Ничего — очищаем остальные
                    cur = ["Ничего"]
                else:
                    if "Ничего" in cur:
                        cur.remove("Ничего")
                    if option in cur:
                        cur.remove(option)
                    else:
                        cur.append(option)
                state["q_data"]["supplements"] = cur
                # автoсохранение
                try:
                    await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
                except Exception as e:
                    logger.error(f"Q autosave: {e}")
                # перерисовать кнопки с отметками
                await self._q_show_supp_buttons(update, ctx, cur)
                return

            # Обратная совместимость: единичный выбор q_supp_X (не используется в новой UI)
            state["q_data"]["supplements"] = q.data.replace("q_supp_", "")
            # автoсохранение
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_motivation"
            await q.edit_message_text("📋 *Блок 5: Психология*\n\nМотивация (1-10):", reply_markup=self.kb([[(str(i), f"q_motiv_{i}") for i in range(1,6)],[(str(i), f"q_motiv_{i}") for i in range(6,11)]]), parse_mode="Markdown")
            return

        if step == "q_supp_other_input":
            # Сюда данные приходят через handle_text (см. ниже ветку q_supp_other_input в handle_text)
            pass

        if step == "q_motivation":
            state["q_data"]["motivation"] = int(q.data.replace("q_motiv_", ""))
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_stress"
            await q.edit_message_text("📋 *Блок 5: Психология*\n\nВнешние факторы мешают?", reply_markup=self.kb([[("Да, сильно", "q_stress_Да, сильно"), ("Периодически", "q_stress_Периодически")], [("Нет", "q_stress_Нет"), ("Не влияют", "q_stress_Не влияют")]]), parse_mode="Markdown")
            return

        if step == "q_stress":
            state["q_data"]["stress"] = q.data.replace("q_stress_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_match_state"
            await q.edit_message_text("📋 *Блок 5: Психология*\n\nСостояние перед матчами?", reply_markup=self.kb([[("Спокойная уверенность", "q_match_Спокойная уверенность")], [("Лёгкое волнение", "q_match_Лёгкое волнение")], [("Сильная тревога", "q_match_Сильная тревога")], [("По-разному", "q_match_По-разному")]]), parse_mode="Markdown")
            return

        if step == "q_match_state":
            state["q_data"]["match_state"] = q.data.replace("q_match_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_reinjury"
            await q.edit_message_text("📋 *Блок 5: Психология*\n\nБоишься повторной травмы?", reply_markup=self.kb([[("Да, постоянно", "q_reinj_Да, постоянно"), ("Иногда", "q_reinj_Иногда")], [("Нет", "q_reinj_Нет"), ("Не было травм", "q_reinj_Не было травм")]]), parse_mode="Markdown")
            return

        if step == "q_reinjury":
            state["q_data"]["reinjury_fear"] = q.data.replace("q_reinj_", "")
            # автoсохранение прогресса
            try:
                await self._db_run(self.db.save_questionnaire, athlete["id"], state.get("q_data", {}))
            except Exception as e:
                logger.error(f"Q autosave: {e}")
            state["step"] = "q_goal"
            await q.edit_message_text("📋 *Блок 6: Цели*\n\nКакую главную цель ставишь? (напиши текстом)", parse_mode="Markdown")
            return

        return

    async def _q_show_zones(self, update, ctx):
        q = update.callback_query
        await q.edit_message_text("📋 *Блок 2: Травмы*\n\nЗоны, которые беспокоили за 3 месяца?", reply_markup=self.kb([
            [("Голеностопы", "q_zones_Голеностопы"), ("Колени", "q_zones_Колени")],
            [("Бёдра/пах", "q_zones_Бёдра"), ("Поясница", "q_zones_Поясница")],
            [("Кисти/пальцы", "q_zones_Кисти"), ("Плечи", "q_zones_Плечи")],
            [("Шея", "q_zones_Шея"), ("Ничего", "q_zones_Ничего")],
        ]), parse_mode="Markdown")

    async def _q_continue_block2(self, update, ctx):
        q = update.callback_query
        state = self.get_state(q.from_user.id)
        state["step"] = "q_surgery"
        await q.edit_message_text("📋 *Блок 2: Травмы*\n\nБыли операции, из-за спорта?", reply_markup=self.kb([[("Да", "q_surgery_Да"), ("Нет", "q_surgery_Нет")]]), parse_mode="Markdown")

    async def _q_continue_block2_meds(self, update, ctx):
        q = update.callback_query
        state = self.get_state(q.from_user.id)
        state["q_data"]["surgery_detail"] = ""
        state["q_data"]["surgery_date"] = ""
        state["step"] = "q_meds"
        await q.edit_message_text("📋 *Блок 2: Травмы*\n\nПринимаешь лекарства, БАДы или физиотерапию?", reply_markup=self.kb([[("Да", "q_meds_Да"), ("Нет", "q_meds_Нет")]]), parse_mode="Markdown")

    async def _q_continue_meds(self, update, ctx, state):
        state["step"] = "q_meds"
        await update.message.reply_text("📋 *Блок 2: Травмы*\n\nПринимаешь лекарства, БАДы или физиотерапию?", reply_markup=self.kb([[("Да", "q_meds_Да"), ("Нет", "q_meds_Нет")]]), parse_mode="Markdown")

    async def _q_save_then_block3(self, update, ctx):
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        if q:
            user_id = q.from_user.id
        else:
            user_id = update.effective_user.id
        state = self.get_state(user_id)
        state["step"] = "q_train_count"
        text_q = "📋 *Блок 3: Тренировочный режим*\n\nСколько раз в неделю тренируешься?"
        kb = self.kb([[(str(i), f"q_tcount_{i}") for i in range(0,11)]])
        if q:
            await q.edit_message_text(text_q, reply_markup=kb, parse_mode="Markdown")
        else:
            await update.message.reply_text(text_q, reply_markup=kb, parse_mode="Markdown")

    async def _q_finish_and_save(self, update, ctx):
        """Сохранить анкету в БД и показать финал"""
        q = update.callback_query
        state = self.get_state(q.from_user.id)
        data = state.get("q_data", {})
        athlete = self.db.get_athlete_by_telegram_id(q.from_user.id)
        if athlete:
            await self._db_run(self.db.save_questionnaire, athlete["id"], data)
            self.db.complete_questionnaire(athlete["id"])
        self.clear_state(q.from_user.id)
        await self._finish_questionnaire_offer(update, ctx)

    async def _finish_questionnaire_offer(self, update, ctx):
        """После анкеты: предложить подписаться на канал «Вне лимита» и перейти к опросу."""
        q = update.callback_query if hasattr(update, 'callback_query') and update.callback_query else None
        text = (
            "✅ *Анкета заполнена!*\n\n"
            "Благодарю за прохождение анкеты!\n\n"
            "В нашем Telegram-канале «Вне лимита» — много актуальной информации о восстановлении, "
            "питании и профилактике травм. Рекомендую подписаться и познакомиться с материалами:\n\n"
            "📢 Нажми кнопку «Подписаться на канал», вернись в чат — и продолжишь с опросом."
        )
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📢 Подписаться на канал", url="https://t.me/vnelimita")],
            [InlineKeyboardButton("✅ Перейти к опросу", callback_data="do_survey")],
            [InlineKeyboardButton("🏠 Главное меню", callback_data="main_menu")],
        ])
        if q:
            await q.edit_message_text(text, parse_mode="Markdown", reply_markup=keyboard)
        else:
            await update.message.reply_text(text, parse_mode="Markdown", reply_markup=keyboard)

    async def show_questionnaire_list(self, update, ctx, page=0):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id): return
        athletes = self.db.get_all_athletes()
        btns = []
        for a in athletes:
            has = self.db.has_questionnaire(a["id"])
            mark = "✅" if has else "⬜"
            btns.append([(f"{mark} {a['id']}. {a['full_name']}", f"qview_{a['id']}")])
        btns.append([("🔙", "admin_manage")])
        await q.edit_message_text("📋 Анкеты", reply_markup=self.kb(btns))

    async def show_questionnaire_detail(self, update, ctx):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id): return
        a_id = int(q.data.replace("qview_", ""))
        athlete = self.db.get_athlete_by_id(a_id)
        if not athlete:
            await q.edit_message_text("Не найден")
            return
        qd = self.db.get_questionnaire(a_id)
        if not qd:
            await q.edit_message_text(f"Анкета для {athlete['full_name']} ещё не заполнена.", reply_markup=self.kb([[(f"🔙", "questionnaire_list")]]))
            return
        fields = [
            ("Возраст", qd.get("age")),
            ("Пол", qd.get("gender")),
            ("Телефон", qd.get("phone")),
            ("Дата рождения", qd.get("birth_date")),
            ("Позиция", qd.get("position")),
            ("Уровень", qd.get("level")),
            ("Стаж", qd.get("experience")),
            ("Рост/Вес", f"{qd.get('height','?')}см/{qd.get('weight','?')}кг"),
            ("Травмы 12м", qd.get("trauma_12m")),
            ("Боль сейчас", qd.get("pain_now")),
            ("Хроническое", qd.get("chronic")),
            ("Операции", qd.get("surgery")),
            ("Операции: что", qd.get("surgery_detail")),
            ("Операции: дата", qd.get("surgery_date")),
            ("Лекарства", qd.get("meds")),
            ("Аллергии", qd.get("allergies")),
            ("Аллергии: на что", qd.get("allergies_detail")),
            ("Трен/нед", qd.get("train_count")),
            ("Мотивация", qd.get("motivation")),
            ("Цель", qd.get("goal","—")),
        ]
        text = f"📋 *Анкета: {athlete['full_name']}*\n━\n\n"
        for label, val in fields:
            if val:
                text += f"*{label}:* {val}\n"

        # ---- Карточка спортсмена: последний опрос + тренд ----
        imported = "from openpyxl import Workbook"  # noqa
        last_wellness = self.db.get_last_wellness(a_id, 7)
        if last_wellness:
            text += f"\n📊 *Последний опрос ({last_wellness[0]['survey_date']}):*\n"
            w0 = last_wellness[0]
            text += f"😴 Сон: {w0.get('sleep_score','—')} | 😩 Утомл: {w0.get('fatigue_score','—')}\n"
            text += f"🤕 Боль: {w0.get('muscle_soreness','—')} | ❤️ Пульс: {w0.get('resting_hr','—')}\n"
            if w0.get('sRPE_score') is not None:
                text += f"💪 Тренировка: {'Да' if w0.get('had_training') else 'Нет'} | sRPE: {w0.get('sRPE_score')}\n"
            if w0.get('cycle_phase'):
                text += f"🔄 Фаза цикла: {w0.get('cycle_phase')}\n"
            if w0.get('complaints'):
                text += f"💬 Жалобы: {w0.get('complaints')}\n"

            # Тренд Hooper за неделю (спарклайн, 3 шкалы: сон+утомление+боль)
            dates = []
            vals = []
            for w in reversed(last_wellness):
                h = sum(filter(None, [w.get('sleep_score'), w.get('fatigue_score'),
                                       w.get('muscle_soreness')]))
                dates.append(w['survey_date'][5:])
                vals.append(h)
            if len(vals) >= 2:
                text += f"\n📈 *Hooper за неделю:*\n"
                for i, h in enumerate(vals):
                    if h >= 17: e = "🟢"
                    elif h >= 12: e = "🟡"
                    else: e = "🔴"
                    text += f"{e}"
                text += "\n"
                text += " ".join(d[-2:] for d in dates) + "\n"
        else:
            text += "\n*Нет опросов за последние 7 дней.*\n"

        buttons = [[(f"🔙", "questionnaire_list"), (f"📊 Отчёт", "admin_report")]]
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=self.kb(buttons))

    async def athlete_list(self, update, ctx, page=0):
        q = update.callback_query
        await q.answer()
        uid = q.from_user.id
        if not self._is_admin_or_coach(uid): return
        athletes = self._scoped_athletes(uid)
        banned = set(x["id"] for x in self.db.get_banned_athletes())
        pp = 15
        total = len(athletes)
        s = page * pp
        e = min(s + pp, total)
        pa = athletes[s:e]
        text = f"👥 Список ({total}) Стр.{page+1}/{(total-1)//pp+1}\n\n"
        prev_team = athletes[s - 1].get("team") if s > 0 else None
        for a in pa:
            t = a.get("team") or "Без команды"
            if t != prev_team:
                text += f"\n🏀 *{t}*\n"
                prev_team = t
            m = "🔒 " if a["id"] in banned else ""
            u = f" @{a['username']}" if a.get("username") else ""
            text += f"{m}{a['id']}. {a['full_name']}{u}\n"
        btns = []
        for a in pa:
            btns.append([(f"🏥 Рекомендации: {a['full_name']}", f"recs_{a['id']}")])
            btns.append([(f"⚖️ Состав тела: {a['full_name']}", f"bc_view_{a['id']}")])
        nav = []
        if page > 0: nav.append(("⬅️", f"athlete_page_{page-1}"))
        if e < total: nav.append(("➡️", f"athlete_page_{page+1}"))
        if nav: btns.append(nav)
        btns.append([("🔙", "admin_manage")])
        await q.edit_message_text(text, reply_markup=self.kb(btns))

    async def show_ban_menu(self, update, ctx, page=0):
        q = update.callback_query
        await q.answer()
        if not self._is_full_access(q.from_user.id): return
        athletes = self.db.get_all_athletes()
        banned = set(x["id"] for x in self.db.get_banned_athletes())
        pp = 10
        total = len(athletes)
        s = page * pp
        e = min(s + pp, total)
        pa = athletes[s:e]
        btns = []
        for a in pa:
            if a["id"] in banned:
                btns.append([(f"✅ {a['id']}. {a['full_name']} Разблок.", f"unban_{a['id']}")])
            else:
                btns.append([(f"🔴 {a['id']}. {a['full_name']} Заблок.", f"ban_{a['id']}")])
        nav = []
        if page > 0: nav.append(("⬅️", f"ban_page_{page-1}"))
        if e < total: nav.append(("➡️", f"ban_page_{page+1}"))
        if nav: btns.append(nav)
        btns.append([("🔙", "admin_manage")])
        await q.edit_message_text(f"🔒 Блокировка Стр.{page+1}/{(total-1)//pp+1}\n\n", reply_markup=self.kb(btns))

    # ==================== ЗАПУСК ====================

    async def _send_daily_reminder(self, context):
        """Ежедневная рассылка напоминаний спортсменам."""
        if REMINDER_HOUR is None:
            return

        # Дедуп: одна рассылка в день — рестарты не спамят повторно
        import pytz
        from datetime import datetime as _dt
        _today = _dt.now(pytz.timezone(REMINDER_TZ)).strftime("%Y-%m-%d")
        if self.db.get_setting("reminder_sent_date") == _today:
            logger.info("Напоминание уже отправлено сегодня — пропускаю")
            return

        if TESTING:
            logger.info("TESTING MODE — напоминания и сводки не отправляются")
            return

        bot = context.bot

        athletes = self.db.get_all_athletes()
        # Batch: один запрос вместо 3N
        survey_map = self.db.get_has_survey_today_map()
        sent = 0
        skipped = 0
        for a in athletes:
            if survey_map.get(a["id"], False):
                skipped += 1
                continue
            if self.db.is_coach(a["telegram_id"]):
                skipped += 1
                continue
            try:
                await bot.send_message(
                    chat_id=a["telegram_id"],
                    text=(
                        "🏀 *ЧБК — Напоминание об опросе!*\n"
                        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                        f"👋 Привет, {self._first_name(a)}!\n\n"
                        "📝 *Ежедневный опрос ещё не пройден!*\n"
                        "Всего 1 минута — и у доктора будут данные о твоём состоянии.\n\n"
                        f"🔥 Твоя серия: *{a.get('survey_streak', 0)} дней*\n\n"
                        "👉 Нажми /start чтобы открыть меню"
                    ),
                    parse_mode="Markdown"
                )
                sent += 1
            except Exception as e:
                logger.error(f"Reminder error for {a['full_name']} (tg={a['telegram_id']}): {e}")

        logger.info(f"Reminder sent to {sent}/{len(athletes)} ({skipped} already surveyed)")

        # Сводка врачу — используем тот же batch (тренеры не спортсмены)
        try:
            _ath = [a for a in athletes if not self.db.is_coach(a["telegram_id"])]
            passed = [a for a in _ath if survey_map.get(a["id"], False)]
            not_passed = [a for a in _ath if not survey_map.get(a["id"], False)]
            summary = (
                f"📊 *Сводка по опросам на сегодня*\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"✅ Прошли ({len(passed)}):\n"
            )
            if passed:
                for a in passed[:15]:
                    summary += f"  • {a['full_name']} ({a['team']})\n"
            else:
                summary += "  — ещё никто не прошёл\n"
            summary += f"\n❌ Не прошли ({len(not_passed)}):\n"
            if not_passed:
                for a in not_passed[:20]:
                    summary += f"  • {a['full_name']} ({a['team']})\n"
            else:
                summary += "  — все прошли! 🎉\n"
            if len(not_passed) > 20:
                summary += f"  • … и ещё {len(not_passed)-20}\n"

            for admin_id in self._full_access_ids():
                try:
                    await bot.send_message(chat_id=admin_id, text=summary, parse_mode="Markdown")
                except Exception as e:
                    logger.error(f"Admin summary error: {e}")
        except Exception as e:
            logger.error(f"Doctor summary error: {e}")

        self.db.set_setting("reminder_sent_date", _today)


    async def _check_trend_alerts(self, context):
        """Ежедневная проверка трендов: 3 дня подряд отклонение → алерт врачу."""
        if TESTING:
            return
        bot = context.bot
        try:
            # Batch: данные за 3 дня для всех
            n_days_map = self.db.get_last_n_days_all(3)
            if not n_days_map:
                return

            athletes = self.db.get_all_athletes()
            athlete_map = {a["id"]: a for a in athletes}
            alerts = []

            for aid, days_data in n_days_map.items():
                if len(days_data) < 3:
                    continue
                a = athlete_map.get(aid, {})
                name = a.get("full_name", f"#{aid}")
                team = a.get("team", "?")

                # Паттерн 1: Пульс 3 дня подряд выше baseline + 1σ
                bl = self.db.get_individual_baseline(aid, 30)
                if bl and bl.get("median_hr") is not None and bl.get("std_hr", 0) is not None:
                    hr_thr = bl["median_hr"] + max(1.5 * (bl["std_hr"] or 0), bl["median_hr"] * 0.10)
                    hr_high = [d.get("resting_hr") and d["resting_hr"] > hr_thr for d in days_data]
                    if all(hr_high):
                        hrs = [d["resting_hr"] for d in days_data]
                        alerts.append(f"❤️ *{name}* ({team}): пульс 3 дня подряд выше нормы ({hrs[0]}→{hrs[1]}→{hrs[2]}, норма ~{int(bl['median_hr'])})")

                # Паттерн 2: Сон 3 дня подряд ниже baseline - 1σ
                if bl and bl.get("median", {}).get("sleep") is not None:
                    sleep_thr = bl["median"]["sleep"] - max(1.5 * (bl.get("std", {}).get("sleep", 0) or 0), 1.0)
                    sleep_low = [d.get("sleep_score") and d["sleep_score"] < sleep_thr for d in days_data]
                    if all(sleep_low):
                        sleeps = [d["sleep_score"] for d in days_data]
                        alerts.append(f"😴 *{name}* ({team}): сон 3 дня подряд ниже нормы ({sleeps[0]}→{sleeps[1]}→{sleeps[2]})")

                # Паттерн 3: Готовность 3 дня подряд < 5
                readiness = [d.get("readiness") for d in days_data]
                if all(r is not None and r < 5 for r in readiness):
                    alerts.append(f"🎯 *{name}* ({team}): готовность 3 дня подряд < 5 ({readiness[0]}→{readiness[1]}→{readiness[2]})")

                # Паттерн 4: Боль NRS 2 дня подряд >= 5
                pain = [d.get("pain_nrs") for d in days_data[-2:]]
                if all(p is not None and p >= 5 for p in pain):
                    loc = days_data[-1].get("pain_location", "")
                    alerts.append(f"🤕 *{name}* ({team}): боль NRS >= 5 два дня подряд ({pain[0]}→{pain[1]}) {loc}")

                # Паттерн 5: Пропуск опроса 2 дня подряд
                dates_present = [d.get("survey_date") for d in days_data]
                if len(dates_present) < 2:
                    alerts.append(f"⚪ *{name}* ({team}): нет опроса за 2+ дня")

            # Отправка алертов врачам
            if alerts:
                alert_text = "🚨 *Trend-алерты за сегодня:*\n━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                alert_text += "\n\n".join(alerts[:15])
                if len(alerts) > 15:
                    alert_text += f"\n\n… и ещё {len(alerts) - 15}"

                for admin_id in self._full_access_ids():
                    try:
                        await bot.send_message(chat_id=admin_id, text=alert_text, parse_mode="Markdown")
                    except Exception as e:
                        logger.error(f"Trend alert send error: {e}")

                logger.info(f"Trend alerts sent: {len(alerts)} patterns detected")
            else:
                logger.info("Trend alerts: no patterns detected")

        except Exception as e:
            logger.error(f"Trend alerts error: {e}", exc_info=True)


    async def _send_weekly_reports(self, context):
        """Воскресный персональный отчёт-картинка каждому спортсмену."""
        if TESTING:
            logger.info("TESTING MODE — воскресные отчёты не отправляются")
            return
        bot = context.bot
        athletes = self.db.get_all_athletes()
        sent = 0
        failed = 0
        for a in athletes:
            try:
                # Пропускаем тех, у кого нет данных за неделю (нет смысла в отчёте)
                stats = self.db.get_athlete_stats(a["id"], 7)
                if not stats.get("total_days"):
                    continue
                trend = [v for _, v in self.db.get_trend_data(a["id"], "sleep", 7)]
                prev = self.db.get_athlete_stats_window(a["id"], 7, 7)
                baseline = self.db.get_individual_baseline(a["id"], 30)
                trends = {
                    "readiness": [v for _, v in self.db.get_trend_data(a["id"], "readiness", 7)],
                    "hr": [v for _, v in self.db.get_trend_data(a["id"], "hr", 7)],
                }
                pain = self.db.get_week_pain_summary(a["id"], 7)
                from weekly_report import generate_weekly_report
                # ранг без эмодзи (PIL не рисует цветные эмодзи)
                _rank = {
                    30: "Золото", 14: "Серебро", 7: "Бронза", 3: "Старт",
                }
                plain_rank = "Новичок"
                for th, name in sorted(_rank.items(), reverse=True):
                    if a.get("survey_streak", 0) >= th:
                        plain_rank = name
                        break
                path = generate_weekly_report(
                    full_name=a["full_name"],
                    age_group=a.get("age_group", "?"),
                    team=a.get("team", "?"),
                    streak=a.get("survey_streak", 0),
                    rank=plain_rank,
                    total=a.get("total_surveys", 0),
                    stats=stats,
                    trend_7d=trend,
                    prev=prev,
                    baseline=baseline,
                    trends=trends,
                    pain=pain,
                )
                with open(path, "rb") as f:
                    await bot.send_photo(
                        chat_id=a["telegram_id"],
                        photo=f,
                        caption=(
                            f"📊 *Твой недельный отчёт*, {self._first_name(a)}!\n"
                            f"Серия: 🔥 {a.get('survey_streak', 0)} "
                            f"{'день' if a.get('survey_streak', 0) == 1 else ('дня' if a.get('survey_streak', 0) % 10 in (2, 3, 4) and a.get('survey_streak', 0) % 100 not in (12, 13, 14) else 'дней')}\n\n"
                            f"Увидимся в понедельник! 💪"
                        ),
                        parse_mode="Markdown"
                    )
                sent += 1
            except Exception as e:
                failed += 1
                logger.error(f"Weekly report error for {a.get('full_name', '?')} (tg={a.get('telegram_id')}): {e}")
        logger.info(f"Weekly reports sent={sent} failed={failed} of {len(athletes)}")

    @staticmethod
    async def error_handler(update, context):
        """Global error handler."""
        logger.error("Exception: %s", context.error, exc_info=context.error)
        if update and update.effective_message:
            try:
                await update.effective_message.reply_text(
                    "\u26a0\ufe0f \u041f\u0440\u043e\u0438\u0437\u043e\u0448\u043b\u0430 \u043e\u0448\u0438\u0431\u043a\u0430. \u041f\u043e\u043f\u0440\u043e\u0431\u0443\u0439\u0442\u0435 \u043f\u043e\u0437\u0436\u0435 \u0438\u043b\u0438 \u043d\u0430\u043f\u0438\u0448\u0438\u0442\u0435 /start"
                )
            except Exception:
                pass

    def run(self):
        app = (
            Application.builder()
            .token(BOT_TOKEN)
        )
        if PROXY_URL:
            app = app.proxy(PROXY_URL).get_updates_proxy(PROXY_URL)
        app = app.build()

        # Кэшируем бота — не создавать Bot(token=...) в _send_admin (ЛАКМУС №7)
        self._bot = app.bot

        # Привязываем очередь заданий, чтобы обработчики могли перепланировать рассылку
        self.job_queue = app.job_queue

        # Global error handler
        app.add_error_handler(self.error_handler)

        if TESTING:
            logger.warning('TESTING MODE — notifications disabled')

        # Периодическая очистка неактивных сессий (защита от утечки памяти)
        try:
            app.job_queue.run_repeating(self._cleanup_stale_sessions, interval=1800, first=600, name="ttl_cleanup")
        except Exception as e:
            logger.warning(f"TTL cleanup not scheduled: {e}")

        app.add_handler(CommandHandler("start", self.cmd_start))
        app.add_handler(CommandHandler("help", self.show_help))
        app.add_handler(CallbackQueryHandler(self.callback_handler))
        app.add_handler(MessageHandler(filters.Document.ALL, self.handle_document))
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, self.handle_text))

        # Планируем ежедневную рассылку
        # Читаем время из БД или используем значение по умолчанию
        try:
            saved_hour = self.db.get_setting("reminder_hour")
            saved_tz = self.db.get_setting("reminder_tz", "Asia/Yekaterinburg")
            # Пустая строка = напоминания выключены (кнопка «off»). Не вызывать int('').
            if saved_hour not in (None, ""):
                global REMINDER_HOUR, REMINDER_TZ
                try:
                    REMINDER_HOUR = int(saved_hour)
                    REMINDER_TZ = saved_tz
                    logger.info(f"Время из БД: {REMINDER_HOUR}:00, TZ: {REMINDER_TZ}")
                except (TypeError, ValueError):
                    logger.warning(f"Некорректное reminder_hour в БД: {saved_hour!r} — считаю выключенным")
                    REMINDER_HOUR = None
            else:
                # В БД пусто — оставляем дефолт из config (будет настроено ниже), но не крашимся
                logger.info("reminder_hour в БД пуст — используется дефолт из config")
        except:
            pass

        if REMINDER_HOUR is not None:
            import pytz
            from datetime import datetime as dt
            local_tz = pytz.timezone(REMINDER_TZ)
            now_local = dt.now(local_tz)
            target_local = now_local.replace(hour=REMINDER_HOUR, minute=REMINDER_MINUTE, second=0, microsecond=0)
            target_utc = target_local.astimezone(pytz.UTC)
            logger.info(f"Напоминание по {REMINDER_TZ} в {REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d} = UTC {target_utc.hour:02d}:{target_utc.minute:02d}")
            
            # Если время ещё не прошло сегодня — планируем через run_daily
            # Убираем старые задания с этим именем, чтобы рестарты не плодили дубли
            for old in self.job_queue.get_jobs_by_name("daily_reminder"):
                old.schedule_removal()
            app.job_queue.run_daily(
                self._send_daily_reminder,
                time=target_utc.time(),
                days=tuple(range(7)),
                name="daily_reminder"
            )
            logger.info(f"Ежедневные напоминания настроены на {REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d}")
            
            # Если время уже прошло сегодня — отправляем сразу
            if now_local.hour >= REMINDER_HOUR and now_local.minute >= REMINDER_MINUTE:
                logger.info("Время напоминания уже прошло сегодня — отправляю сразу")
                # Запускаем через 3 секунды, чтобы бот успел инициализироваться
                app.job_queue.run_once(self._send_daily_reminder, 3, name="initial_reminder")

        # Trend-алерты (проверка 10:15, через 15 мин после напоминаний)
        try:
            import pytz as _pytz2
            from datetime import datetime as _dt2
            trend_tz = _pytz2.timezone("Asia/Yekaterinburg")
            trend_target = trend_tz.localize(_dt2(1970, 1, 1, 10, 15))
            for old in self.job_queue.get_jobs_by_name("trend_alerts"):
                old.schedule_removal()
            app.job_queue.run_daily(
                self._check_trend_alerts,
                time=trend_target.timetz(),
                days=tuple(range(7)),
                name="trend_alerts"
            )
            logger.info("Trend-алерты настроены на 10:15 по Челябинску")
        except Exception as e:
            logger.warning(f"Trend alerts not scheduled: {e}")

        # Воскресный персональный отчёт-картинка (каждое воскресенье 18:00 по ЕКБ)
        try:
            import pytz as _pytz
            from datetime import datetime as _dt_cls
            sunday_tz = _pytz.timezone("Asia/Yekaterinburg")
            sunday_target = sunday_tz.localize(_dt_cls(1970, 1, 4, 18, 0))  # воскресенье
            for old in self.job_queue.get_jobs_by_name("weekly_report"):
                old.schedule_removal()
            app.job_queue.run_daily(
                self._send_weekly_reports,
                time=sunday_target.timetz(),
                days=(6,),  # Sunday (0=Mon, 6=Sun)
                name="weekly_report"
            )
            logger.info("Воскресные отчёты настроены на 18:00 по Челябинску")
        except Exception as e:
            logger.warning(f"Weekly report not scheduled: {e}")

        # Понедельничная сводка для тренеров (09:00 по ЕКБ)
        try:
            monday_target = _pytz.timezone("Asia/Yekaterinburg").localize(_dt_cls(1970, 1, 5, 9, 0))  # понедельник
            for old in self.job_queue.get_jobs_by_name("weekly_coach_summary"):
                old.schedule_removal()
            app.job_queue.run_daily(
                self._send_weekly_coach_summary,
                time=monday_target.timetz(),
                days=(0,),  # Monday (0=Mon)
                name="weekly_coach_summary"
            )
            logger.info("Сводка тренеров настроена на понедельник 09:00 по Челябинску")
        except Exception as e:
            logger.warning(f"Coach summary not scheduled: {e}")

        logger.info("Бот v3.0 запущен!")
        app.run_polling(drop_pending_updates=False)


if __name__ == "__main__":
    bot = SportHealthBot()
    bot.run()

