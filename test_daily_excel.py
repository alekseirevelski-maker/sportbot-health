import sys, io
sys.path.insert(0, "/opt/telegram-bots/sport-health")
from database import Database
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

db = Database()

# Test new method
rows = db.get_wellness_by_period(7, None)
print(f"get_wellness_by_period(7) -> {len(rows)} rows")
if rows:
    print("Sample row:", rows[0])

# Test per team
rows_chbk = db.get_wellness_by_period(7, "ЧБК-2")
print(f"get_wellness_by_period(7,'ЧБК-2') -> {len(rows_chbk)} rows")

# Simulate report generation for team
from datetime import datetime, date
HR_NORMS = {
    "U14": {"min": 55, "max": 75}, "U15": {"min": 50, "max": 70},
    "U16": {"min": 50, "max": 70}, "U17": {"min": 45, "max": 65},
    "U18": {"min": 45, "max": 65}, "U19": {"min": 42, "max": 62},
    "U21": {"min": 42, "max": 62}, "Pro": {"min": 40, "max": 60},
}

if rows_chbk:
    wb = Workbook()
    ws = wb.active
    ws.title = "ЧБК-2"
    headers = ["ФИО","Дата","Сон","Стресс","Утомл","Боль","Настр","Пульс","Трен","sRPE","Фаза","Жалобы"]
    for ci,h in enumerate(headers,1):
        ws.cell(row=1,column=ci,value=h).font=Font(bold=True)
    r=2
    for row in rows_chbk:
        ws.cell(row=r,column=1,value=row.get("full_name"))
        ws.cell(row=r,column=2,value=row.get("survey_date"))
        ws.cell(row=r,column=3,value=row.get("sleep_score"))
        ws.cell(row=r,column=4,value=row.get("stress_score"))
        ws.cell(row=r,column=5,value=row.get("fatigue_score"))
        ws.cell(row=r,column=6,value=row.get("muscle_soreness"))
        ws.cell(row=r,column=7,value=row.get("mood_score"))
        ws.cell(row=r,column=8,value=row.get("resting_hr"))
        ws.cell(row=r,column=9,value="Да" if row.get("had_training") else "Нет")
        ws.cell(row=r,column=10,value=row.get("sRPE_score"))
        ws.cell(row=r,column=11,value=row.get("cycle_phase"))
        ws.cell(row=r,column=12,value=row.get("complaints"))
        r+=1
    out=io.BytesIO()
    wb.save(out)
    print(f"Team report generated OK: {len(out.getvalue())} bytes")
